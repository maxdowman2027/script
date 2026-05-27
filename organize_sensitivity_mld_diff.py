#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
整理 sensitivity 结果 CSV：在 bandwidth / coding / wifi_format / cur_degree、
rate、rx_chan 相同配置下，将 mld_en=0 与 mld_en=1 两行合并为一行。

- 去掉 mld_en 列
- sensitivity_dbm 拆为 sensitivity_dbm_mld_en0、sensitivity_dbm_mld_en1
- 新增 sensitivity_dbm_mld_diff = sensitivity_dbm_mld_en0 - sensitivity_dbm_mld_en1
- 同时写出带差值列填充色的 xlsx（差值大偏绿，负值偏红，中间黄/橙过渡）

默认读取 output/sensitivity_out/result 下 *_result.csv，写出到
output/sensitivity_out/result/organized/（或 --output_dir 指定）。
"""

from __future__ import annotations

import argparse
import glob
import os
from typing import List, Sequence, Tuple

import numpy as np
import pandas as pd

CONFIG_MATCH_COLS = ("bandwidth", "coding", "wifi_format", "cur_degree")
PAIR_EXTRA_COLS = ("rate", "rx_chan")

SENS_MLD_EN0_COL = "sensitivity_dbm_mld_en0"
SENS_MLD_EN1_COL = "sensitivity_dbm_mld_en1"
DIFF_COL = "sensitivity_dbm_mld_diff"

# 合并键之外、随 mld_en 变化的列（不写入宽表）
DROP_ON_MERGE = frozenset(
    {
        "mld_en",
        "sensitivity_dbm",
        "testcase_folder",
        "rx_session_dir",
    }
)

DEFAULT_INPUT_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "output",
    "sensitivity_out",
    "result",
)


def _normalize_mld_en(series: pd.Series) -> pd.Series:
    def _one(v):
        if pd.isna(v):
            return np.nan
        s = str(v).strip()
        if s in ("0", "1"):
            return s
        try:
            i = int(float(s))
            if i in (0, 1):
                return str(i)
        except (TypeError, ValueError):
            pass
        return s

    return series.map(_one)


def merge_mld_sensitivity_rows(
    df: pd.DataFrame,
    config_cols: Sequence[str] = CONFIG_MATCH_COLS,
    extra_cols: Sequence[str] = PAIR_EXTRA_COLS,
) -> pd.DataFrame:
    """
    Pivot long -> wide: one row per (config + rate + rx_chan) with both mld_en values.

    Only keeps groups that have both mld_en=0 and mld_en=1.
    Other columns are taken from the mld_en=0 row (when present), else first row in group.
    """
    required = list(config_cols) + list(extra_cols) + ["mld_en", "sensitivity_dbm"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"CSV missing columns: {missing}")

    work = df.copy()
    work.columns = [str(c).strip() for c in work.columns]
    work["_mld_norm"] = _normalize_mld_en(work["mld_en"])
    work["_sens_num"] = pd.to_numeric(work["sensitivity_dbm"], errors="coerce")

    key_cols = list(config_cols) + list(extra_cols)

    wide = (
        work.dropna(subset=["_mld_norm"])
        .pivot_table(
            index=key_cols,
            columns="_mld_norm",
            values="_sens_num",
            aggfunc="first",
        )
    )
    if "0" not in wide.columns or "1" not in wide.columns:
        return pd.DataFrame()

    out = wide[["0", "1"]].copy()
    out.columns = [SENS_MLD_EN0_COL, SENS_MLD_EN1_COL]
    out[DIFF_COL] = out[SENS_MLD_EN0_COL] - out[SENS_MLD_EN1_COL]
    out = out.reset_index()

    meta_cols = [
        c
        for c in work.columns
        if c not in key_cols
        and c not in DROP_ON_MERGE
        and not c.startswith("_")
    ]
    if meta_cols:
        meta0 = work[work["_mld_norm"] == "0"].drop_duplicates(subset=key_cols, keep="first")
        if meta0.empty:
            meta0 = work.drop_duplicates(subset=key_cols, keep="first")
        meta0 = meta0[key_cols + [c for c in meta_cols if c in meta0.columns]]
        out = out.merge(meta0, on=key_cols, how="left")

    # 稳定列顺序：配置键 -> 元数据 -> 灵敏度三列
    front = list(config_cols) + list(extra_cols)
    meta_ordered = [c for c in out.columns if c not in front and c not in (
        SENS_MLD_EN0_COL,
        SENS_MLD_EN1_COL,
        DIFF_COL,
    )]
    tail = [SENS_MLD_EN0_COL, SENS_MLD_EN1_COL, DIFF_COL]
    out = out[front + meta_ordered + tail]
    return out


def _mld_diff_fill(diff_value: float):
    """差值 = mld_en0 - mld_en1：越大越绿，负值越红，中间黄/橙过渡（单位 dB）。"""
    from openpyxl.styles import PatternFill

    if diff_value >= 2.0:
        return PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    if diff_value >= 1.0:
        return PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    if diff_value >= -1.0:
        return PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    if diff_value >= -2.0:
        return PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    return PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


def write_mld_wide_outputs(df_out: pd.DataFrame, csv_path: str, xlsx_path: str) -> None:
    """Write CSV and XLSX; color sensitivity_dbm_mld_diff column in XLSX."""
    import openpyxl

    parent = os.path.dirname(os.path.abspath(csv_path))
    if parent:
        os.makedirs(parent, exist_ok=True)

    df_out.to_csv(csv_path, index=False)

    df_out.to_excel(xlsx_path, index=False, engine="openpyxl")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    diff_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == DIFF_COL:
            diff_col_idx = idx
            break

    if diff_col_idx is not None:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=diff_col_idx)
            if isinstance(cell.value, (int, float)) and not (
                isinstance(cell.value, float) and np.isnan(cell.value)
            ):
                cell.fill = _mld_diff_fill(float(cell.value))

    wb.save(xlsx_path)
    wb.close()


def process_csv_file(
    input_path: str,
    output_csv_path: str,
    output_xlsx_path: str,
    config_cols: Sequence[str] = CONFIG_MATCH_COLS,
    extra_cols: Sequence[str] = PAIR_EXTRA_COLS,
) -> Tuple[int, int, int]:
    """Returns (input_rows, output_rows, unpaired_input_rows_estimate)."""
    df = pd.read_csv(input_path)
    n_in = len(df)
    df_out = merge_mld_sensitivity_rows(df, config_cols=config_cols, extra_cols=extra_cols)
    n_out = len(df_out)

    write_mld_wide_outputs(df_out, output_csv_path, output_xlsx_path)
    return n_in, n_out, n_in - 2 * n_out


def collect_input_files(input_dir: str, pattern: str) -> List[str]:
    path = os.path.join(input_dir, pattern)
    return sorted(f for f in glob.glob(path) if os.path.isfile(f))


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Merge mld_en=0/1 rows into one wide row per "
            "bandwidth, coding, wifi_format, cur_degree, rate, rx_chan."
        )
    )
    parser.add_argument("--input_dir", default=DEFAULT_INPUT_DIR)
    parser.add_argument("--pattern", default="*_result.csv")
    parser.add_argument(
        "--output_dir",
        default=None,
        help="Output directory (default: <input_dir>/organized)",
    )
    parser.add_argument("--combined", default=None, help="Optional combined output CSV")
    parser.add_argument("--inplace", action="store_true")
    args = parser.parse_args()

    input_dir = os.path.abspath(args.input_dir)
    output_dir = (
        input_dir
        if args.inplace
        else os.path.abspath(args.output_dir or os.path.join(input_dir, "organized"))
    )

    files = collect_input_files(input_dir, args.pattern)
    if not files:
        print(f"No files matching {args.pattern} in {input_dir}")
        return

    combined_parts: List[pd.DataFrame] = []
    print(f"Input: {input_dir} ({len(files)} file(s))")
    if not args.inplace:
        print(f"Output: {output_dir}")

    for fp in files:
        base = os.path.basename(fp)
        stem, _ = os.path.splitext(base)
        if args.inplace:
            csv_fp = os.path.splitext(fp)[0] + "_mld_wide.csv"
            xlsx_fp = os.path.splitext(fp)[0] + "_mld_wide.xlsx"
        else:
            csv_fp = os.path.join(output_dir, f"{stem}_mld_wide.csv")
            xlsx_fp = os.path.join(output_dir, f"{stem}_mld_wide.xlsx")

        n_in, n_out, n_drop = process_csv_file(fp, csv_fp, xlsx_fp)
        print(
            f"  {base}: {n_in} rows -> {n_out} merged rows "
            f"({n_drop} input rows not in full 0+1 pairs); csv+xlsx"
        )

        if args.combined:
            combined_parts.append(pd.read_csv(csv_fp))

    if args.combined and combined_parts:
        comb_path = os.path.abspath(args.combined)
        parent = os.path.dirname(comb_path)
        if parent:
            os.makedirs(parent, exist_ok=True)
        pd.concat(combined_parts, ignore_index=True).to_csv(comb_path, index=False)
        print(f"Combined: {comb_path} ({sum(len(p) for p in combined_parts)} rows)")


if __name__ == "__main__":
    main()
