import openpyxl

def check_colored_columns():
    file_path = "D:/chip_test/dev/xian_test/Xian-Esp-Test-Scripts/py_script_fpga_tx_wifi7/Log/wifi_tx_rls4/no_he_2/crc_fail_result.xlsx"

    try:
        wb = openpyxl.load_workbook(file_path)
        sheet_name = wb.sheetnames[0]  # 使用第一个Sheet
        ws = wb[sheet_name]

        priority_columns = ['tx_power_set(dBm)', 'evm', 'evm_nss0', 'evm_nss1']
        found_columns = []

        print(f"在 {sheet_name} 中检查列名颜色：")
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            cell_font = ws.cell(row=1, column=col_idx).font

            if cell_value in priority_columns:
                found_columns.append(cell_value)
                color = cell_font.color.rgb if cell_font.color else "无颜色"
                print(f"列 {col_idx} ({cell_value})：字体颜色 = {color}")
            else:
                color = cell_font.color.rgb if cell_font.color else "无颜色"
                if color != "无颜色":
                    print(f"列 {col_idx} ({cell_value})：字体颜色 = {color}")

        print()
        print("找到的重点列：")
        for col in found_columns:
            print(f"  - {col}")

        print()
        print(f"总共找到 {len(found_columns)} 个重点列")
        if len(found_columns) == len(priority_columns):
            print("所有重点列都已标红")
        else:
            missing = [col for col in priority_columns if col not in found_columns]
            print(f"缺少以下重点列：{', '.join(missing)}")

        wb.close()
    except Exception as e:
        print(f"Error reading file: {e}")

if __name__ == "__main__":
    check_colored_columns()