import openpyxl

def check_fill_colors():
    file_path = "D:/chip_test/dev/xian_test/Xian-Esp-Test-Scripts/py_script_fpga_tx_wifi7/Log/wifi_tx_rls4/no_he/merged_result.xlsx"

    try:
        # 打开工作簿
        wb = openpyxl.load_workbook(file_path)

        # 获取第一个Sheet
        sheet_name = "channel11_BCC"
        ws = wb[sheet_name]

        # 查找wifi_format列的索引
        wifi_format_index = None
        for idx, cell in enumerate(ws[1]):
            if cell.value == "wifi_format":
                wifi_format_index = idx
                break

        if wifi_format_index is not None:
            print(f"在 {sheet_name} 中找到wifi_format列，索引为: {wifi_format_index}")

            # 遍历前50行，检查填充色
            colors_used = {}
            for row_idx in range(2, min(51, ws.max_row + 1)):
                cell_value = ws.cell(row=row_idx, column=wifi_format_index + 1).value
                row_fill = None
                for col_idx in range(1, ws.max_column + 1):
                    cell_fill = ws.cell(row=row_idx, column=col_idx).fill
                    if cell_fill.start_color.index != "00000000":
                        row_fill = cell_fill.start_color.index
                        break

                if cell_value and row_fill:
                    if cell_value not in colors_used:
                        colors_used[cell_value] = row_fill
                    print(f"行 {row_idx}: 格式='{cell_value}', 颜色='{row_fill}'")

            print(f"\n检测到的格式与颜色对应关系:")
            for format_name, color in colors_used.items():
                print(f"格式: '{format_name}', 颜色: '{color}'")
        else:
            print(f"在 {sheet_name} 中未找到wifi_format列")

        wb.close()

    except Exception as e:
        print(f"读取文件失败: {e}")

if __name__ == "__main__":
    check_fill_colors()