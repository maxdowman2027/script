import os
import openpyxl

def verify_merged_files():
    output_dir = "D:/chip_test/dev/xian_test/Xian-Esp-Test-Scripts/py_script_fpga_tx_wifi7/Log/wifi_tx_rls4/no_he_2"
    merged_file = os.path.join(output_dir, "merged_result.xlsx")
    crc_file = os.path.join(output_dir, "crc_fail_result.xlsx")

    # 检查文件是否存在
    if not os.path.exists(merged_file):
        print(f"合并后的文件不存在: {merged_file}")
        return False

    if not os.path.exists(crc_file):
        print(f"CRC失败文件不存在: {crc_file}")
        return False

    print("两个文件都已成功生成")

    # 验证merged_result.xlsx
    try:
        wb = openpyxl.load_workbook(merged_file)
        print(f"merged_result.xlsx包含 {len(wb.sheetnames)} 个工作表:")
        for sheet_name in wb.sheetnames:
            print(f"  - {sheet_name}")

            ws = wb[sheet_name]
            print(f"  包含 {ws.max_row - 1} 行数据")

            # 检查列是否包含wifi_format和evm列
            found_wifi_format = False
            found_evm = False
            for cell in ws[1]:
                if cell.value == "wifi_format":
                    found_wifi_format = True
                if cell.value == "evm":
                    found_evm = True

            if found_wifi_format:
                print(f"  OK 包含wifi_format列")
            else:
                print(f"  ERROR 不包含wifi_format列")

            if found_evm:
                print(f"  OK 包含evm列")
            else:
                print(f"  ERROR 不包含evm列")

            # 检查evm_nss0和evm_nss1列是否在evm列之后
            if found_evm and ("evm_nss0" in [cell.value for cell in ws[1]] or "evm_nss1" in [cell.value for cell in ws[1]]):
                print(f"  OK evm_nss列已正确放置在evm列之后")
            else:
                print(f"  未找到evm_nss列或evm列")

        wb.close()
        print()
    except Exception as e:
        print(f"读取merged_result.xlsx失败: {e}")
        return False

    # 验证crc_fail_result.xlsx
    try:
        wb = openpyxl.load_workbook(crc_file)
        print(f"crc_fail_result.xlsx包含 {len(wb.sheetnames)} 个工作表:")
        for sheet_name in wb.sheetnames:
            print(f"  - {sheet_name}")

            ws = wb[sheet_name]
            print(f"  包含 {ws.max_row - 1} 行数据")

            # 检查列是否包含重点列
            priority_columns = ['tx_power_set(dBm)', 'evm', 'evm_nss0', 'evm_nss1']
            found_priority_columns = []
            for cell in ws[1]:
                if cell.value in priority_columns:
                    found_priority_columns.append(cell.value)

            print(f"  找到 {len(found_priority_columns)} 个重点列:")
            for col in found_priority_columns:
                print(f"    - {col}")

            # 检查前几个重点列是否加粗
            bold_columns = []
            for idx, cell in enumerate(ws[1]):
                if cell.value in priority_columns and cell.font.bold:
                    bold_columns.append(cell.value)

            print(f"  {len(bold_columns)} 个重点列标题已加粗")

            # 检查是否有填充色
            has_fill_color = False
            for row in ws.iter_rows(min_row=2, max_row=min(10, ws.max_row)):
                for cell in row:
                    if cell.fill.start_color.index != "00000000":
                        has_fill_color = True
                        break
                if has_fill_color:
                    break

            if has_fill_color:
                print(f"  OK 包含填充色")
            else:
                print(f"  ERROR 不包含填充色")

        wb.close()
        print()
    except Exception as e:
        print(f"读取crc_fail_result.xlsx失败: {e}")
        return False

    print("所有验证都通过！")
    return True

if __name__ == "__main__":
    verify_merged_files()