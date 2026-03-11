
import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill
from openpyxl import Workbook

# 读取CSV文件
file_path = r'D:/users/gxu/spur_scan/260310/scan_spur_data/xtal_duty_disable/spur_scan_result_2G_coef.csv'
df = pd.read_csv(file_path)

# 检查pwr 4、pwr 5、pwr 6列是否存在（对应Excel中的G、H、I列）
required_columns = ['pwr 4', 'pwr 5', 'pwr 6']
for col in required_columns:
    if col not in df.columns:
        raise ValueError(f'列 {col} 不存在于CSV文件中')

# 计算每一行的统计量
def detect_outliers(row):
    # 只处理数值类型的数据
    valid_values = []
    valid_columns = []
    for col in required_columns:
        val = row[col]
        # 检查值是否是有效的数值（不是['no_spur']或#VALUE!）
        if isinstance(val, str) and (val.startswith("['") or val == '#VALUE!'):
            continue
        try:
            float_val = float(val)
            valid_values.append(float_val)
            valid_columns.append(col)
        except:
            continue

    outliers = []
    if len(valid_values) >= 2:  # 需要至少2个有效值才能计算偏差
        mean_val = np.mean(valid_values)
        # 使用绝对偏差判断（超过0.3 dB认为是异常）
        for col, val in zip(valid_columns, valid_values):
            if abs(val - mean_val) > 0.3:  # 0.3 dB的绝对偏差
                outliers.append(col)
    return outliers

df['异常列'] = df.apply(detect_outliers, axis=1)

# 保存为Excel文件
output_file = file_path.replace('.csv', '_异常检测.xlsx')
df.to_excel(output_file, index=False)

# 使用openpyxl加载并添加颜色
from openpyxl import load_workbook
wb = load_workbook(output_file)
ws = wb.active

# 定义红色填充
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

# 找到G、H、I列的列索引
col_indices = {col: chr(ord('A') + i) for i, col in enumerate(df.columns)}

# 遍历每一行并标记异常值
for row_idx in range(2, ws.max_row + 1):
    outliers = df.loc[row_idx - 2, '异常列']
    for col in outliers:
        cell = ws[f'{col_indices[col]}{row_idx}']
        cell.fill = red_fill

wb.save(output_file)
print(f'检测完成！结果已保存至: {output_file}')
