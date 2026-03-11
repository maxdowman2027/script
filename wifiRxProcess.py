import pandas as pd
import matplotlib.pyplot as plt
import os
import openpyxl
from matplotlib.pyplot import MultipleLocator
from matplotlib.backends.backend_pdf import PdfPages
import glob
import math
import datetime
import time
import fnmatch
import re
from openpyxl import load_workbook, Workbook

def wifi_rx_plot(foldername):
    """
    执行wifiRxPlot.py的功能：
    - 读取指定文件夹下的CSV数据
    - 计算PER和EVM
    - 绘制灵敏度曲线
    - 计算灵敏度指标
    - 保存到Excel和PDF
    """
    # 确保路径以分隔符结尾
    if not foldername.endswith(os.sep):
        foldername = foldername + os.sep

    sens_accuracy = 100
    PAK_NUM = 1000
    current_time = time.localtime()
    sens_path = foldername + 'sens_' + str(current_time.tm_year) + str(current_time.tm_mon) + str(current_time.tm_mday) + '_' + \
                str(current_time.tm_hour) + str(current_time.tm_min) + str(current_time.tm_sec) + '.xlsx'
    print(f"保存路径: {sens_path}")

    mybook = openpyxl.Workbook()
    mybook.save(sens_path)
    writer = pd.ExcelWriter(sens_path, mode='a', engine="openpyxl")

    subfolder_list = next(os.walk(foldername))[1]
    for subfolder_name in subfolder_list:
        pp = PdfPages(os.path.join(foldername, subfolder_name + '_sensitivity.pdf'))
        testcase_list = os.listdir(os.path.join(foldername, subfolder_name))

        for testcase in testcase_list:
            testpath = foldername + subfolder_name + '/' + testcase
            os.chdir(testpath)
            my_files = sorted(glob.glob('*.csv'), key=os.path.getmtime)
            df = pd.DataFrame()
            pak_num = PAK_NUM

            for i in range(len(my_files)):
                df = df.append(pd.read_csv(testpath + '/' + my_files[i], index_col=False))

            if 'acr' in testcase or 'ACI' in testcase:
                chan_list = df[' rx_chan'].unique()
            else:
                chan_list = df[' rx_chan'].unique()

            for chan in chan_list:
                legend_convert = []
                column_convert = []
                x1 = plt.figure(dpi=64, figsize=(11, 18))

                if 'acr' in testcase or 'ACI' in testcase:
                    df_chan = df[df[' rx_chan'] == chan]
                    df_chan['per'] = df_chan[' rxnum'].map(lambda x: 1 - min(x, pak_num) / pak_num)
                else:
                    df_chan = df[df[' rx_chan'] == chan]
                    df_chan['per'] = df_chan[' rxnum'].map(lambda x: 1 - min(x, pak_num) / pak_num)

                if 'acr' in testcase or 'ACI' in testcase:
                    table = pd.pivot_table(df_chan, index=[" acr"], columns=["rate"], values=["per"])
                else:
                    table = pd.pivot_table(df_chan, index=[" rfpwr"], columns=["rate"], values=["per"])

                if ' evm0' in df_chan.columns:
                    if 'acr' in testcase or 'ACI' in testcase:
                        table_evm = pd.pivot_table(df_chan, index=[" acr"], columns=["rate"], values=[" evm0"])
                    else:
                        table_evm = pd.pivot_table(df_chan, index=[" rfpwr"], columns=["rate"], values=[" evm0"])
                else:
                    table_evm = pd.DataFrame()

                columnlist = table.columns
                for column in columnlist:
                    column_convert.append(column[1])
                table.columns = column_convert
                if not table_evm.empty:
                    table_evm.columns = column_convert

                sens_result = list()
                for column in column_convert:
                    per4pow = table[column]
                    per = per4pow.values
                    pow = per4pow.index

                    for i in range(0, len(per)):
                        if '11b' in testcase:
                            if 'acr' in testcase or 'ACI' in testcase:
                                if per[i] > 0.08:
                                    if per[i-1] == 0:
                                        delta_per = (math.log10(per[i]) - 0.0001) / sens_accuracy
                                    else:
                                        delta_per = (math.log10(per[i]) - math.log10(per[i-1])) / sens_accuracy
                                    per_sens = math.log10(per[i])
                                    pow_sens = pow[i]
                                    for j in range(0, sens_accuracy):
                                        per_sens = per_sens - delta_per
                                        pow_sens = pow_sens - 1 / sens_accuracy
                                        if per_sens <= -1.096:
                                            pow_sens_result = pow_sens
                                            break
                                    break
                            else:
                                pow_sens_result = 0
                                if per[i] < 0.08:
                                    if per[i] == 0:
                                        delta_per = (math.log10(per[i - 1]) + 10) / sens_accuracy
                                        per_sens = -10
                                    else:
                                        delta_per = (math.log10(per[i-1]) - math.log10(per[i])) / sens_accuracy
                                        per_sens = math.log10(per[i])
                                    pow_sens = pow[i]
                                    for j in range(0, sens_accuracy):
                                        per_sens = per_sens + delta_per
                                        pow_sens = pow_sens - 1 / sens_accuracy
                                        if per_sens >= -1.096:
                                            pow_sens_result = pow_sens
                                            break
                                    break
                        else:
                            if 'acr' in testcase or 'ACI' in testcase:
                                if per[i] > 0.1:
                                    if per[i-1] == 0:
                                        delta_per = (math.log10(per[i]) + 2) / sens_accuracy
                                    else:
                                        delta_per = (math.log10(per[i]) - math.log10(per[i-1])) / sens_accuracy
                                    per_sens = math.log10(per[i])
                                    pow_sens = pow[i]
                                    for j in range(0, sens_accuracy):
                                        per_sens = per_sens - delta_per
                                        pow_sens = pow_sens - 1 / sens_accuracy
                                        if per_sens <= -1:
                                            pow_sens_result = pow_sens
                                            break
                                    break
                            else:
                                pow_sens_result = 0
                                if per[i] < 0.1:
                                    if per[i] == 0:
                                        if per[i - 1] == 0:
                                            delta_per = 0
                                        else:
                                            delta_per = (math.log10(per[i - 1]) + 10) / sens_accuracy
                                        per_sens = -10
                                    else:
                                        delta_per = (math.log10(per[i-1]) - math.log10(per[i])) / sens_accuracy
                                        per_sens = math.log10(per[i])
                                    pow_sens = pow[i]
                                    for j in range(0, sens_accuracy):
                                        per_sens = per_sens + delta_per
                                        pow_sens = pow_sens - 1 / sens_accuracy
                                        if per_sens >= -1:
                                            pow_sens_result = pow_sens
                                            break
                                    break
                    sens_result.append(round(pow_sens_result, 2))

                sens_dictionary = dict(zip(column_convert, sens_result))
                sens_df = pd.DataFrame(sens_dictionary, index=[0])
                sens_df.to_excel(writer, sheet_name=testcase + ' ' + subfolder_name + ' ' + 'chan' + str(chan), index=False)

                legendList = df_chan['rate'].unique()
                table = table[legendList]
                legend_convert = legendList

                custom_colors = ['#FF0000', '#00FF00', '#0000FF', '#FFFF00', '#FF00FF', '#00FFFF', '#800000', '#008000',
                                 '#000080', '#808000', '#F08080', '#8080F0', '#806F86', '#006066']
                plt.subplot(2, 1, 1)
                for i in range(len(table.columns)):
                    plt.semilogy(table.index, table[table.columns[i]].values, 'o-', color=custom_colors[i])

                plt.ylim([1e-4, 1])
                if 'acr' in testcase or 'ACI' in testcase:
                    plt.xlim([-16, 50])
                    plt.xlabel('ACR (dB)')
                else:
                    plt.xlim([-110, -5])
                    plt.xlabel('power (dBm)')

                plt.ylabel('PER')
                plt.legend(legend_convert, bbox_to_anchor=(1, 1), loc=2, borderaxespad=0, numpoints=1, fontsize=8)
                plt.title(testcase + ' ' + subfolder_name + ' ' + 'chan' + str(chan) + ' sensitivity')
                plt.grid()
                x_major_locator = MultipleLocator(5)
                ax = plt.gca()
                plt.subplots_adjust(right=0.8)
                ax.xaxis.set_major_locator(x_major_locator)

                if not table_evm.empty:
                    plt.subplot(2, 1, 2)
                    for i in range(len(table_evm.columns)):
                        plt.plot(table_evm.index, table_evm[table_evm.columns[i]].values, 'o-', color=custom_colors[i])
                    plt.ylim([-60, 0])
                    if 'acr' in testcase or 'ACI' in testcase:
                        plt.xlim([-16, 50])
                        plt.xlabel('ACR (dB)')
                    else:
                        plt.xlim([-110, -5])
                        plt.xlabel('power (dBm)')

                    plt.ylabel('EVM(dB)')
                    plt.legend(legend_convert, bbox_to_anchor=(1, 1), loc=2, borderaxespad=0, numpoints=1, fontsize=8)
                    plt.title(testcase + ' ' + subfolder_name + ' ' + 'chan' + str(chan) + ' EVM')
                    plt.grid()
                    x_major_locator = MultipleLocator(5)
                    ax = plt.gca()
                    plt.subplots_adjust(right=0.8)
                    ax.xaxis.set_major_locator(x_major_locator)

                pp.savefig(x1)

        pp.close()
    writer.close()
    return sens_path

def merge_all_sheets_to_one(
    root_path,
    folder_pattern,
    output_file,
    xlsx_pattern="*.xlsx",
    use_regex_folder=False,
    use_regex_xlsx=False,
    sheet_name_col="原工作表名称"
):
    """
    执行file_merge.py的功能：
    - 搜索符合条件的文件夹和XLSX文件
    - 合并所有sheet到单个输出文件
    - 新增列标注原工作表名称
    """
    if not os.path.exists(root_path):
        print(f"错误：根路径 '{root_path}' 不存在！")
        return

    target_folders = []
    for item in os.listdir(root_path):
        item_path = os.path.join(root_path, item)
        if not os.path.isdir(item_path):
            continue

        is_match = re.match(folder_pattern, item) is not None if use_regex_folder else fnmatch.fnmatch(item, folder_pattern)
        if is_match:
            target_folders.append(item_path)
            print(f"找到文件夹：{item_path}")

    if not target_folders:
        print("未找到符合条件的文件夹！")
        return

    xlsx_files = []
    for folder in target_folders:
        try:
            # 只处理最新的sens文件，避免重复
            candidate_files = []
            for file_name in os.listdir(folder):
                file_path = os.path.join(folder, file_name)
                if not os.path.isfile(file_path) or not file_name.lower().endswith(".xlsx"):
                    continue

                is_xlsx_match = re.match(xlsx_pattern, file_name) is not None if use_regex_xlsx else fnmatch.fnmatch(file_name, xlsx_pattern)
                if is_xlsx_match:
                    candidate_files.append(file_path)

            # 找到最新的文件
            if candidate_files:
                # 按修改时间排序，获取最新的
                candidate_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
                latest_file = candidate_files[0]
                xlsx_files.append(latest_file)
                print(f"找到最新XLSX文件：{latest_file}")

        except PermissionError:
            print(f"权限错误：无法访问文件夹 {folder}")
        except Exception as e:
            print(f"遍历文件夹 {folder} 时出错：{str(e)}")

    if not xlsx_files:
        print("未找到符合条件的XLSX文件！")
        return

    wb_output = Workbook()
    ws_output = wb_output.active
    ws_output.title = "合并数据"
    header_written = False
    total_rows = 0
    total_files = 0
    total_sheets = 0

    try:
        for xlsx_file in xlsx_files:
            try:
                wb_input = load_workbook(xlsx_file, read_only=True, data_only=True)
                total_files += 1
                print(f"\n处理文件：{xlsx_file}")

                for sheet_name in wb_input.sheetnames:
                    total_sheets += 1
                    ws_input = wb_input[sheet_name]
                    rows = list(ws_input.iter_rows(values_only=True))

                    if not rows:
                        print(f"  跳过空工作表：{sheet_name}")
                        continue

                    if not header_written:
                        header_row = [sheet_name_col] + list(rows[0])
                        ws_output.append(header_row)
                        header_written = True
                        data_rows = rows[1:]
                    else:
                        data_rows = rows[1:] if len(rows) > 1 else []

                    if data_rows:
                        for row in data_rows:
                            if all(cell is None for cell in row):
                                continue
                            new_row = [sheet_name] + list(row)
                            ws_output.append(new_row)
                        row_count = len(data_rows)
                        total_rows += row_count
                        print(f"  合并工作表 {sheet_name}：{row_count} 行数据")

                wb_input.close()

            except Exception as e:
                print(f"读取 {xlsx_file} 时出错：{str(e)}，跳过该文件")

        wb_output.save(output_file)
        wb_output.close()

        print(f"\n=== 合并完成 ===")
        print(f"输出文件：{output_file}")
        print(f"共处理 {len(target_folders)} 个文件夹")
        print(f"共处理 {total_files} 个XLSX文件，{total_sheets} 个工作表")
        print(f"总计合并 {total_rows} 行数据（不含表头）")
        print(f"新增列名：{sheet_name_col}")

    except PermissionError:
        print(f"权限错误：无法写入输出文件 {output_file}")
    except Exception as e:
        print(f"合并过程中出错：{str(e)}")

def merge_notch_files_to_spur(root_path, notch0_file, notch1_file, output_file):
    """
    合并notch_enable0和notch_enable1的Excel文件为spur文件
    notch_enable0的数据在上，notch_enable1的数据在下
    notch_enable0使用黄色填充，notch_enable1使用绿色填充
    """
    print(f"\n=== 合并 {notch0_file} 和 {notch1_file} ===")

    # 读取两个文件
    if not os.path.exists(notch0_file):
        print(f"错误：文件 {notch0_file} 不存在！")
        return
    if not os.path.exists(notch1_file):
        print(f"错误：文件 {notch1_file} 不存在！")
        return

    df0 = pd.read_excel(notch0_file)
    df1 = pd.read_excel(notch1_file)

    # 合并数据
    merged_df = pd.concat([df0, df1], ignore_index=True)

    # 使用openpyxl引擎保存文件
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    merged_df.to_excel(writer, index=False, sheet_name='合并数据')
    writer.save()
    writer.close()

    print(f"合并完成！输出文件：{output_file}")
    print(f"数据行数：{len(merged_df)} 行（notch0: {len(df0)} 行，notch1: {len(df1)} 行）")

    # 设置单元格填充颜色
    wb = load_workbook(output_file)
    ws = wb.active

    # 定义填充样式
    from openpyxl.styles import PatternFill

    # 黄色填充（notch_enable0）
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    # 绿色填充（notch_enable1）
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    # 获取列数
    col_count = ws.max_column

    # 为notch_enable0的数据设置黄色填充（第2行到第len(df0)+1行）
    for row in range(2, len(df0) + 2):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = yellow_fill

    # 为notch_enable1的数据设置绿色填充（第len(df0)+2行到最后一行）
    for row in range(len(df0) + 2, ws.max_row + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = green_fill

    # 保存修改后的文件
    wb.save(output_file)
    print("颜色填充已设置完成！")
    print(f"notch_enable0数据使用黄色填充（第2-{len(df0)+1}行）")
    print(f"notch_enable1数据使用绿色填充（第{len(df0)+2}-{ws.max_row}行）")


def main():
    """
    主函数：
    1. 处理notch_enable0和notch_enable1文件夹
    2. 执行wifiRxPlot功能
    3. 执行file_merge功能
    4. 合并生成spur文件
    """
    # 请根据实际路径修改
    root_path = r"D:\users\gxu\spur_scan\2G_high_mcs\40m\vht"
    prefix = "2G_40m_vht"

    # 处理notch_enable0
    notch0_folder = os.path.join(root_path, "notch_enable0")
    if os.path.exists(notch0_folder):
        print(f"\n=== 处理 {notch0_folder} ===")
        wifi_rx_plot(notch0_folder)

        output0_file = os.path.join(root_path, f"{prefix}_notch_enable0.xlsx")
        merge_all_sheets_to_one(
            root_path=root_path,
            folder_pattern="notch_enable0",
            output_file=output0_file,
            xlsx_pattern="*sens_*.xlsx",
            use_regex_folder=False,
            use_regex_xlsx=False,
            sheet_name_col="原工作表名称"
        )
    else:
        print(f"警告：文件夹 {notch0_folder} 不存在！")

    # 处理notch_enable1
    notch1_folder = os.path.join(root_path, "notch_enable1")
    if os.path.exists(notch1_folder):
        print(f"\n=== 处理 {notch1_folder} ===")
        wifi_rx_plot(notch1_folder)

        output1_file = os.path.join(root_path, f"{prefix}_notch_enable1.xlsx")
        merge_all_sheets_to_one(
            root_path=root_path,
            folder_pattern="notch_enable1",
            output_file=output1_file,
            xlsx_pattern="*sens_*.xlsx",
            use_regex_folder=False,
            use_regex_xlsx=False,
            sheet_name_col="原工作表名称"
        )
    else:
        print(f"警告：文件夹 {notch1_folder} 不存在！")

    # 合并生成spur文件
    notch0_file = os.path.join(root_path, f"{prefix}_notch_enable0.xlsx")
    notch1_file = os.path.join(root_path, f"{prefix}_notch_enable1.xlsx")
    spur_file = os.path.join(root_path, f"{prefix}_spur.xlsx")

    if os.path.exists(notch0_file) and os.path.exists(notch1_file):
        merge_notch_files_to_spur(root_path, notch0_file, notch1_file, spur_file)
    else:
        print("\n警告：无法合并notch文件，因为其中一个或两个文件不存在！")

if __name__ == "__main__":
    main()
