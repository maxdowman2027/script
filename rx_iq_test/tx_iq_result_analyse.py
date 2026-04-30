#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
处理指定路径下的 wifi_tx_cali_xxx 格式的CSV文件：
1. 将所有CSV文件合并到一个Excel文件中，按带宽（bw）分为20、40、80、160四个sheet
2. 对diff_pwr大于指定阈值的单元格设置红色填充色
3. 统计异常值的数量和占比，并保存到另一个Excel文件中
"""

import os
import glob
import pandas as pd
from openpyxl.styles import PatternFill

# ================= 配置区域 =================
# 在这里修改配置
SEARCH_DIRECTORY = r"D:\users\gxu\E22\tx_iq\260421"  # 要搜索的根目录
FILE_PATTERN = "wifi_tx_cali_*.csv"  # 文件名匹配模式（支持通配符）
DIFF_PWR_THRESHOLD = -45  # diff_pwr列的阈值（值大于此数的行将被标记）
OUTPUT_MERGED_FILE = r"D:\users\gxu\scripts\output\tx_iq_260421\merged_tx_iq_data.xlsx"  # 合并数据输出文件
OUTPUT_STATS_FILE = r"D:\users\gxu\scripts\output\tx_iq_260421\tx_iq_statistics.xlsx"  # 统计结果输出文件
# ===========================================


def find_csv_files(directory: str, pattern: str = "*.csv") -> list:
    """
    递归查找指定目录下符合模式的CSV文件

    Args:
        directory: 要搜索的目录
        pattern: 文件名匹配模式，默认为"*.csv"

    Returns:
        符合条件的文件路径列表
    """
    csv_files = []
    # 使用glob递归查找符合模式的文件
    for file_path in glob.glob(os.path.join(directory, "**", pattern), recursive=True):
        if os.path.isfile(file_path):
            csv_files.append(file_path)
    return csv_files


def merge_and_analyze_csv_files(csv_files: list, threshold: float, merged_output: str, stats_output: str):
    """
    合并CSV文件，按带宽分组到不同的sheet，标记异常值，并统计结果

    Args:
        csv_files: 要处理的CSV文件列表
        threshold: diff_pwr列的阈值
        merged_output: 合并数据输出文件路径
        stats_output: 统计结果输出文件路径
    """
    # 按带宽分组存储数据
    bw_data = {
        20: [],
        40: [],
        80: [],
        160: []
    }

    print("开始处理 %d 个CSV文件..." % len(csv_files))

    for file_path in csv_files:
        print("正在处理: %s" % file_path)
        try:
            # 读取CSV文件，处理最后一列的多余逗号
            import csv
            data = []
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # 移除行尾的逗号
                    processed_row = {}
                    for key, value in row.items():
                        processed_key = key.strip() if isinstance(key, str) else str(key).strip()
                        if isinstance(value, str):
                            processed_value = value.strip().rstrip(',')
                        elif value is None:
                            processed_value = ''
                        else:
                            processed_value = str(value).strip().rstrip(',')
                        processed_row[processed_key] = processed_value
                    data.append(processed_row)

            df = pd.DataFrame(data)

            # 转换数值列到适当的类型
            for col in ['tx_iq_on', 'bw', 'freqMhz', 'freqCw', 'tone_freq', 'filt_len', 'sig_pwr', 'ori_pwr', 'mir_pwr', 'diff_pwr']:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')

            # 去除所有列名的前后空格
            df.columns = [col.strip() for col in df.columns]

            # 检查是否包含必要的列
            required_columns = ["bw", "diff_pwr"]
            has_required_columns = True
            for col in required_columns:
                col_found = False
                for df_col in df.columns:
                    if df_col.strip().lower() == col:
                        col_found = True
                        break
                if not col_found:
                    has_required_columns = False
                    print("警告: 文件 %s 不包含 %s 列，已跳过" % (file_path, col))
                    break

            if not has_required_columns:
                continue

            # 添加来源文件信息
            df["source_file"] = os.path.basename(file_path)
            df["full_path"] = file_path

            # 按带宽分组
            for bw in bw_data.keys():
                filtered_df = df[df["bw"] == bw]
                if not filtered_df.empty:
                    bw_data[bw].append(filtered_df)

        except Exception as e:
            print("错误: 无法读取文件 %s: %s" % (file_path, e))
            continue

    # 合并每个带宽的数据
    merged_data = {}
    for bw in bw_data.keys():
        if bw_data[bw]:
            merged_data[bw] = pd.concat(bw_data[bw], ignore_index=True)
            print("带宽 %dMHz: %d 行数据" % (bw, len(merged_data[bw])))
        else:
            merged_data[bw] = pd.DataFrame()

    # 保存合并后的数据到Excel文件
    os.makedirs(os.path.dirname(merged_output), exist_ok=True)

    with pd.ExcelWriter(merged_output, engine="openpyxl") as writer:
        # 检查是否有数据需要保存
        has_data = False
        for bw in [20, 40, 80, 160]:
            if not merged_data[bw].empty:
                has_data = True
                sheet_name = "BW%d" % bw
                merged_data[bw].to_excel(writer, sheet_name=sheet_name, index=False)

                # 获取工作表
                worksheet = writer.sheets[sheet_name]

                # 查找diff_pwr列的索引
                diff_pwr_index = None
                for idx, col in enumerate(merged_data[bw].columns):
                    if col.strip().lower() == "diff_pwr":
                        diff_pwr_index = idx
                        break

                if diff_pwr_index is not None:
                    # 定义红色填充色
                    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # 红色

                    # 遍历所有行，添加填充色
                    for row in range(2, len(merged_data[bw]) + 2):  # Excel行号从2开始（跳过表头）
                        cell_value = merged_data[bw].iloc[row - 2, diff_pwr_index]
                        try:
                            value = float(cell_value)
                            if value > threshold:
                                cell = worksheet.cell(row=row, column=diff_pwr_index + 1)
                                cell.fill = red_fill
                        except (ValueError, TypeError):
                            continue

        # 如果没有数据，创建一个空的工作表
        if not has_data:
            pd.DataFrame().to_excel(writer, sheet_name="NoData", index=False)

    print("合并数据已保存到: %s" % merged_output)

    # 统计结果
    stats_data = []
    total_all_rows = 0
    total_anomaly_rows = 0

    for bw in [20, 40, 80, 160]:
        if not merged_data[bw].empty:
            total_rows = len(merged_data[bw])
            valid_values = merged_data[bw]["diff_pwr"].dropna()
            invalid_rows = total_rows - len(valid_values)
            anomaly_rows = sum(valid_values > threshold)
            anomaly_percentage = (anomaly_rows / total_rows) * 100

            stats_data.append({
                "带宽 (MHz)": bw,
                "总行数": total_rows,
                "有效行数": len(valid_values),
                "无效行数": invalid_rows,
                "diff_pwr > %d 的行数" % threshold: anomaly_rows,
                "占比 (%)": "%.2f" % anomaly_percentage
            })

            total_all_rows += total_rows
            total_anomaly_rows += anomaly_rows

    # 计算总体统计
    if total_all_rows > 0:
        overall_percentage = (total_anomaly_rows / total_all_rows) * 100
        stats_data.append({
            "带宽 (MHz)": "总体",
            "总行数": total_all_rows,
            "有效行数": "-",
            "无效行数": "-",
            "diff_pwr > %d 的行数" % threshold: total_anomaly_rows,
            "占比 (%)": "%.2f" % overall_percentage
        })

    # 收集所有 diff_pwr 大于阈值的数据
    anomaly_data = []
    for bw in [20, 40, 80, 160]:
        if not merged_data[bw].empty:
            filtered_df = merged_data[bw][merged_data[bw]["diff_pwr"] > threshold]
            if not filtered_df.empty:
                anomaly_data.append(filtered_df)

    # 合并异常数据
    if anomaly_data:
        anomaly_df = pd.concat(anomaly_data, ignore_index=True)
    else:
        anomaly_df = pd.DataFrame()

    # 计算所有点中 diff_pwr 的最大值
    all_diff_pwr_values = []
    for bw in [20, 40, 80, 160]:
        if not merged_data[bw].empty:
            all_diff_pwr_values.extend(merged_data[bw]["diff_pwr"].dropna().tolist())

    max_diff_pwr = None
    if all_diff_pwr_values:
        max_diff_pwr = max(all_diff_pwr_values)

    # 计算超过门限的 diff_pwr 值的平均值
    avg_anomaly_diff_pwr = None
    if not anomaly_df.empty:
        avg_anomaly_diff_pwr = anomaly_df["diff_pwr"].mean()

    # 更新统计数据
    if max_diff_pwr is not None:
        for stat in stats_data:
            if stat["带宽 (MHz)"] == "总体":
                stat["diff_pwr 最大值"] = max_diff_pwr
                if avg_anomaly_diff_pwr is not None:
                    stat["异常值平均值"] = avg_anomaly_diff_pwr
                else:
                    stat["异常值平均值"] = "-"
            else:
                # 计算每个带宽的 diff_pwr 最大值
                bw = stat["带宽 (MHz)"]
                if not merged_data[bw].empty:
                    stat["diff_pwr 最大值"] = merged_data[bw]["diff_pwr"].max()
                    # 计算每个带宽的异常值平均值
                    if bw in [20, 40, 80, 160]:
                        bw_anomaly_data = merged_data[bw][merged_data[bw]["diff_pwr"] > threshold]
                        if not bw_anomaly_data.empty:
                            stat["异常值平均值"] = bw_anomaly_data["diff_pwr"].mean()
                        else:
                            stat["异常值平均值"] = "-"
                else:
                    stat["diff_pwr 最大值"] = "-"
                    stat["异常值平均值"] = "-"

    # 保存统计结果
    os.makedirs(os.path.dirname(stats_output), exist_ok=True)

    stats_df = pd.DataFrame(stats_data)
    with pd.ExcelWriter(stats_output, engine="openpyxl") as writer:
        stats_df.to_excel(writer, sheet_name="Statistics", index=False)
        anomaly_df.to_excel(writer, sheet_name="AnomalyData", index=False)

    print("统计结果已保存到: %s" % stats_output)

    # 打印统计信息
    print("\n=== 统计结果 ===")
    for stat in stats_data:
        if stat["带宽 (MHz)"] == "总体":
            print("\n总体统计:")
        else:
            print("\n带宽 %dMHz:" % stat["带宽 (MHz)"])
        print("  总行数: %d" % stat["总行数"])
        if stat["带宽 (MHz)"] != "总体":
            print("  有效行数: %d" % stat["有效行数"])
            print("  无效行数: %d" % stat["无效行数"])
        print("  diff_pwr > %d 的行数: %d" % (threshold, stat["diff_pwr > %d 的行数" % threshold]))
        print("  占比: %s%%" % stat["占比 (%)"])


def main():
    print("配置信息:")
    print("  搜索目录: %s" % SEARCH_DIRECTORY)
    print("  文件模式: %s" % FILE_PATTERN)
    print("  阈值: %d" % DIFF_PWR_THRESHOLD)
    print("  合并数据输出文件: %s" % OUTPUT_MERGED_FILE)
    print("  统计结果输出文件: %s" % OUTPUT_STATS_FILE)
    print()

    # 查找符合条件的CSV文件
    csv_files = find_csv_files(SEARCH_DIRECTORY, FILE_PATTERN)

    if not csv_files:
        print("在 %s 中未找到符合模式 %s 的CSV文件" % (SEARCH_DIRECTORY, FILE_PATTERN))
        return

    print("找到 %d 个符合条件的CSV文件" % len(csv_files))

    # 合并、分析和统计数据
    merge_and_analyze_csv_files(csv_files, DIFF_PWR_THRESHOLD, OUTPUT_MERGED_FILE, OUTPUT_STATS_FILE)

    print("\n所有处理完成!")


if __name__ == "__main__":
    main()