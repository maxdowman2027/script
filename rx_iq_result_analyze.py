#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
递归检索指定目录下的CSV文件，处理两个任务：
1. 汇总符合文件名匹配模式的文件，将diff_pwr列小于指定值的行添加填充色并统计占比
2. 统计所有最边缘点，将diff_pwr列小于指定值的行添加填充色并统计占比
"""

import os
import glob
import pandas as pd
from openpyxl.styles import PatternFill


# ================= 配置区域 =================
# 在这里修改配置
SEARCH_DIRECTORY = r"D:\users\gxu\rx_iq\E22\regression_v2_0414"  # 要搜索的根目录
FILE_PATTERN = "rx_iq_cal_res_*.csv"  # 文件名匹配模式（支持通配符）
DIFF_PWR_THRESHOLD = 45  # diff_pwr列的阈值（值小于此数的行将被标记）
OUTPUT_FILE_ALL = r"D:\users\gxu\scripts\output\all_rows_analysis.xlsx"  # 所有行分析输出文件
OUTPUT_FILE_EDGE = r"D:\users\gxu\scripts\output\edge_points_analysis.xlsx"  # 边缘点分析输出文件
# ===========================================


def find_csv_files(directory: str, pattern: str = "*.csv") -> list:
    """
    递归查找指定目录下符合模式的CSV文件

    Args:
        directory: 要搜索的目录
        pattern: 文件名匹配模式，默认为"*.csv"

    Returns:
        符合条件的文件路径列表
    """
    csv_files = []
    # 使用glob递归查找符合模式的文件
    for file_path in glob.glob(os.path.join(directory, "**", pattern), recursive=True):
        if os.path.isfile(file_path):
            csv_files.append(file_path)
    return csv_files


def extract_all_rows(csv_files: list, threshold: float, output_file: str):
    """
    从CSV文件中提取所有行，将diff_pwr列小于阈值的行添加填充色，并统计占比

    Args:
        csv_files: 要处理的CSV文件列表
        threshold: diff_pwr列的阈值
        output_file: 输出文件路径
    """
    all_data = []

    print(f"开始处理 {len(csv_files)} 个CSV文件...")

    for file_path in csv_files:
        print(f"正在处理: {file_path}")
        try:
            # 读取CSV文件
            df = pd.read_csv(file_path)

            # 去除所有列名的前后空格
            df.columns = [col.strip() for col in df.columns]

            # 检查是否包含diff_pwr列（忽略大小写和空格）
            has_diff_pwr = False
            diff_pwr_column = None
            for col in df.columns:
                if col.strip().lower() == "diff_pwr":
                    has_diff_pwr = True
                    diff_pwr_column = col
                    break

            if not has_diff_pwr:
                print(f"警告: 文件 {file_path} 不包含 diff_pwr 列，已跳过")
                continue

            # 添加来源文件信息
            df["source_file"] = os.path.basename(file_path)
            df["full_path"] = file_path

            all_data.append(df)
        except Exception as e:
            print(f"错误: 无法读取文件 {file_path}: {e}")
            continue

    if not all_data:
        print("未找到包含diff_pwr列的有效数据")
        return

    # 合并所有数据
    merged_df = pd.concat(all_data, ignore_index=True)

    # 保存到Excel文件
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        merged_df.to_excel(writer, sheet_name="AllData", index=False)

        # 获取工作表
        worksheet = writer.sheets["AllData"]

        # 查找diff_pwr列的索引
        diff_pwr_index = None
        for idx, col in enumerate(merged_df.columns):
            if col.strip().lower() == "diff_pwr":
                diff_pwr_index = idx
                break

        if diff_pwr_index is not None:
            # 定义填充色（使用RGB颜色代码）
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # 红色
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # 绿色

            # 遍历所有行，添加填充色
            for row in range(2, len(merged_df) + 2):  # Excel行号从2开始（跳过表头）
                cell_value = merged_df.iloc[row - 2, diff_pwr_index]
                try:
                    value = float(cell_value)
                    if value < threshold:
                        cell = worksheet.cell(row=row, column=diff_pwr_index + 1)
                        cell.fill = red_fill
                    else:
                        cell = worksheet.cell(row=row, column=diff_pwr_index + 1)
                        cell.fill = green_fill
                except (ValueError, TypeError):
                    continue

        # 统计小于阈值的占比
        try:
            total_rows = len(merged_df)
            valid_values = merged_df[diff_pwr_column].dropna()
            invalid_rows = total_rows - len(valid_values)
            less_threshold_rows = sum(valid_values < threshold)
            less_threshold_percentage = (less_threshold_rows / total_rows) * 100

            # 添加统计信息
            stats_df = pd.DataFrame({
                "统计项": [
                    "总行数",
                    "有效行数",
                    "无效行数",
                    f"diff_pwr < {threshold} 的行数",
                    f"diff_pwr < {threshold} 的占比"
                ],
                "数值": [
                    total_rows,
                    len(valid_values),
                    invalid_rows,
                    less_threshold_rows,
                    f"{less_threshold_percentage:.2f}%"
                ]
            })

            stats_df.to_excel(writer, sheet_name="Statistics", index=False)

            print(f"处理完成! 共处理 {total_rows} 行数据")
            print(f"diff_pwr < {threshold} 的行数: {less_threshold_rows}")
            print(f"占比: {less_threshold_percentage:.2f}%")

        except Exception as e:
            print(f"统计信息计算错误: {e}")

    print(f"结果已保存到: {output_file}")


def extract_edge_points(csv_files: list, threshold: float, output_file: str):
    """
    从CSV文件中提取最边缘点，将diff_pwr列小于阈值的行添加填充色，并统计占比

    Args:
        csv_files: 要处理的CSV文件列表
        threshold: diff_pwr列的阈值
        output_file: 输出文件路径
    """
    all_data = []

    print(f"开始处理 {len(csv_files)} 个CSV文件...")

    for file_path in csv_files:
        print(f"正在处理: {file_path}")
        try:
            # 读取CSV文件
            df = pd.read_csv(file_path)

            # 去除所有列名的前后空格
            df.columns = [col.strip() for col in df.columns]

            # 检查是否包含必要的列
            required_columns = ["bw", "tone_freq", "diff_pwr"]
            has_required_columns = True
            for col in required_columns:
                col_found = False
                for df_col in df.columns:
                    if df_col.strip().lower() == col:
                        col_found = True
                        break
                if not col_found:
                    has_required_columns = False
                    print(f"警告: 文件 {file_path} 不包含 {col} 列，已跳过")
                    break

            if not has_required_columns:
                continue

            # 提取最边缘点
            edge_points = []
            unique_bw_values = df["bw"].unique()

            for bw in unique_bw_values:
                # 筛选当前带宽的数据
                bw_data = df[df["bw"] == bw]

                # 找到最大和最小的tone_freq
                max_freq = bw_data["tone_freq"].max()
                min_freq = bw_data["tone_freq"].min()

                # 添加边缘点
                edge_points.append(bw_data[bw_data["tone_freq"] == max_freq])
                edge_points.append(bw_data[bw_data["tone_freq"] == min_freq])

            if edge_points:
                edge_df = pd.concat(edge_points, ignore_index=True)

                # 添加来源文件信息
                edge_df["source_file"] = os.path.basename(file_path)
                edge_df["full_path"] = file_path

                all_data.append(edge_df)

        except Exception as e:
            print(f"错误: 无法读取文件 {file_path}: {e}")
            continue

    if not all_data:
        print("未找到包含有效边缘点的数据")
        return

    # 合并所有边缘点数据
    merged_df = pd.concat(all_data, ignore_index=True)

    # 保存到Excel文件
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        merged_df.to_excel(writer, sheet_name="EdgePoints", index=False)

        # 获取工作表
        worksheet = writer.sheets["EdgePoints"]

        # 查找diff_pwr列的索引
        diff_pwr_index = None
        for idx, col in enumerate(merged_df.columns):
            if col.strip().lower() == "diff_pwr":
                diff_pwr_index = idx
                break

        if diff_pwr_index is not None:
            # 定义填充色（使用RGB颜色代码）
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # 红色
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # 绿色

            # 遍历所有行，添加填充色
            for row in range(2, len(merged_df) + 2):  # Excel行号从2开始（跳过表头）
                cell_value = merged_df.iloc[row - 2, diff_pwr_index]
                try:
                    value = float(cell_value)
                    if value < threshold:
                        cell = worksheet.cell(row=row, column=diff_pwr_index + 1)
                        cell.fill = red_fill
                    else:
                        cell = worksheet.cell(row=row, column=diff_pwr_index + 1)
                        cell.fill = green_fill
                except (ValueError, TypeError):
                    continue

        # 统计小于阈值的占比
        try:
            total_rows = len(merged_df)
            valid_values = merged_df["diff_pwr"].dropna()
            invalid_rows = total_rows - len(valid_values)
            less_threshold_rows = sum(valid_values < threshold)
            less_threshold_percentage = (less_threshold_rows / total_rows) * 100

            # 添加统计信息
            stats_df = pd.DataFrame({
                "统计项": [
                    "总边缘点行数",
                    "有效行数",
                    "无效行数",
                    f"diff_pwr < {threshold} 的行数",
                    f"diff_pwr < {threshold} 的占比"
                ],
                "数值": [
                    total_rows,
                    len(valid_values),
                    invalid_rows,
                    less_threshold_rows,
                    f"{less_threshold_percentage:.2f}%"
                ]
            })

            stats_df.to_excel(writer, sheet_name="Statistics", index=False)

            print(f"处理完成! 共处理 {total_rows} 行边缘点数据")
            print(f"diff_pwr < {threshold} 的行数: {less_threshold_rows}")
            print(f"占比: {less_threshold_percentage:.2f}%")

        except Exception as e:
            print(f"统计信息计算错误: {e}")

    print(f"结果已保存到: {output_file}")


def main():
    print("配置信息:")
    print(f"  搜索目录: {SEARCH_DIRECTORY}")
    print(f"  文件模式: {FILE_PATTERN}")
    print(f"  阈值: {DIFF_PWR_THRESHOLD}")
    print(f"  所有行分析输出文件: {OUTPUT_FILE_ALL}")
    print(f"  边缘点分析输出文件: {OUTPUT_FILE_EDGE}")
    print()

    # 查找符合条件的CSV文件
    csv_files = find_csv_files(SEARCH_DIRECTORY, FILE_PATTERN)

    if not csv_files:
        print(f"在 {SEARCH_DIRECTORY} 中未找到符合模式 {FILE_PATTERN} 的CSV文件")
        return

    print(f"找到 {len(csv_files)} 个符合条件的CSV文件")

    # 处理所有行
    print("\n=== 处理所有行 ===")
    extract_all_rows(csv_files, DIFF_PWR_THRESHOLD, OUTPUT_FILE_ALL)

    # 处理边缘点
    print("\n=== 处理边缘点 ===")
    extract_edge_points(csv_files, DIFF_PWR_THRESHOLD, OUTPUT_FILE_EDGE)

    print("\n所有处理完成!")


if __name__ == "__main__":
    main()