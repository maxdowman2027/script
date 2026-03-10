import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font

# 读取两个 CSV 文件
file1_path = r'D:\users\gxu\spur_scan\260309\scan_spur_data\normal\spur_scan_result_2G_coef.csv'
file2_path = r'D:\users\gxu\spur_scan\260309\scan_spur_data\xtal_duty_disable\spur_scan_result_2G_coef.csv'

df1 = pd.read_csv(file1_path)
df2 = pd.read_csv(file2_path)

# 数据预处理
for col in ['phy_mode', 'channel', 'frequency', 'Used_Frequency', 'X_CoefFixed', 'Y_CoefFixed']:
    if col in df1.columns:
        df1[col] = df1[col].replace(['no_spur', 'invalid'], np.nan)
    if col in df2.columns:
        df2[col] = df2[col].replace(['no_spur', 'invalid'], np.nan)

pwr_cols = [col for col in df1.columns if 'pwr' in col or 'avg' in col]
for col in pwr_cols:
    if col in df1.columns:
        df1[col] = df1[col].replace(['#VALUE!', '[', ']', '', ' ', 'no_spur'], np.nan)
        df1[col] = pd.to_numeric(df1[col], errors='coerce')
    if col in df2.columns:
        df2[col] = df2[col].replace(['#VALUE!', '[', ']', '', ' ', 'no_spur'], np.nan)
        df2[col] = pd.to_numeric(df2[col], errors='coerce')

merge_columns = ['phy_mode', 'channel', 'frequency', 'Used_Frequency', 'X_CoefFixed', 'Y_CoefFixed']
merged_df = pd.merge(df1, df2, on=merge_columns, how='inner', suffixes=('_normal', '_xtal'))

# 计算每个配置下循环pwr值的差异
for mode in ['normal', 'xtal']:
    pwr1 = f'pwr 1_{mode}' if mode == 'normal' else f'pwr 4_{mode}'
    pwr2 = f'pwr 2_{mode}' if mode == 'normal' else f'pwr 5_{mode}'
    pwr3 = f'pwr 3_{mode}' if mode == 'normal' else f'pwr 6_{mode}'

    if all(col in merged_df.columns for col in [pwr1, pwr2, pwr3]):
        # 计算标准差
        merged_df[f'pwr_std_{mode}'] = merged_df[[pwr1, pwr2, pwr3]].std(axis=1, skipna=True)

        # 计算最大值和最小值的差异
        merged_df[f'pwr_range_{mode}'] = merged_df[[pwr1, pwr2, pwr3]].max(axis=1, skipna=True) - merged_df[[pwr1, pwr2, pwr3]].min(axis=1, skipna=True)

# 设定差异阈值（3dB或5dB为较大差异）
variance_threshold = 3
merged_df['has_high_variance_normal'] = merged_df['pwr_range_normal'] > variance_threshold if 'pwr_range_normal' in merged_df.columns else False
merged_df['has_high_variance_xtal'] = merged_df['pwr_range_xtal'] > variance_threshold if 'pwr_range_xtal' in merged_df.columns else False
merged_df['has_high_std_normal'] = merged_df['pwr_std_normal'] > variance_threshold/3 if 'pwr_std_normal' in merged_df.columns else False
merged_df['has_high_std_xtal'] = merged_df['pwr_std_xtal'] > variance_threshold/3 if 'pwr_std_xtal' in merged_df.columns else False

# 保存到Excel
output_file = r'D:\users\gxu\spur_scan\260309\scan_spur_data\spur_loop_variance_check.xlsx'
merged_df.to_excel(output_file, index=False, engine='openpyxl')

# 设置Excel样式
wb = openpyxl.load_workbook(output_file)
ws = wb.active

# 定义样式
yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
orange_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
bold_font = Font(bold=True)

# 标注差异较大的配置
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
    index = i
    high_variance = False

    if merged_df.iloc[index]['has_high_variance_normal']:
        ws.cell(row=i+2, column=1).font = bold_font
        for cell in row:
            cell.fill = yellow_fill
        high_variance = True

    if merged_df.iloc[index]['has_high_variance_xtal']:
        if high_variance:
            for cell in row:
                cell.fill = red_fill
        else:
            ws.cell(row=i+2, column=1).font = bold_font
            for cell in row:
                cell.fill = orange_fill

# 添加说明列
ws.cell(row=1, column=len(merged_df.columns)+1, value='说明')
ws.cell(row=2, column=len(merged_df.columns)+1, value='黄色：Normal模式循环差异大')
ws.cell(row=3, column=len(merged_df.columns)+1, value='橙色：XTAL模式循环差异大')
ws.cell(row=4, column=len(merged_df.columns)+1, value='红色：两个模式循环差异都大')
ws.cell(row=5, column=len(merged_df.columns)+1, value=f'差异阈值：> {variance_threshold}dB')

# 设置列宽
for col in range(1, ws.max_column+1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

wb.save(output_file)

print(f"循环pwr值差异检查完成，结果已保存到: {output_file}")

# 显示统计信息
print(f"\n=== 循环pwr值差异统计 ===")
if 'pwr_range_normal' in merged_df.columns:
    high_var_normal = len(merged_df[merged_df['has_high_variance_normal']])
    print(f"Normal模式循环差异> {variance_threshold}dB的配置数: {high_var_normal} ({high_var_normal/len(merged_df)*100:.1f}%)")

if 'pwr_range_xtal' in merged_df.columns:
    high_var_xtal = len(merged_df[merged_df['has_high_variance_xtal']])
    print(f"XTAL模式循环差异> {variance_threshold}dB的配置数: {high_var_xtal} ({high_var_xtal/len(merged_df)*100:.1f}%)")

both_high_var = len(merged_df[merged_df['has_high_variance_normal'] & merged_df['has_high_variance_xtal']])
if both_high_var > 0:
    print(f"两个模式循环差异都大的配置数: {both_high_var} ({both_high_var/len(merged_df)*100:.1f}%)")

print(f"\n=== 详细异常配置列表 ===")
for idx, row in merged_df[merged_df['has_high_variance_normal'] | merged_df['has_high_variance_xtal']].iterrows():
    config_info = []
    if not np.isnan(row['phy_mode']):
        config_info.append(f"phy_mode={row['phy_mode']}")
    if not np.isnan(row['channel']):
        config_info.append(f"channel={row['channel']}")
    if pd.notnull(row['frequency']):
        config_info.append(f"frequency={row['frequency']}")
    config_str = ", ".join(config_info)

    variance_info = []
    if row['has_high_variance_normal']:
        variance_info.append(f"Normal: {row['pwr_range_normal']:.1f}dB")
    if row['has_high_variance_xtal']:
        variance_info.append(f"XTAL: {row['pwr_range_xtal']:.1f}dB")
    variance_str = ", ".join(variance_info)

    print(f"{config_str}: {variance_str}")
