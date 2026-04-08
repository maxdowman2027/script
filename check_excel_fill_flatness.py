import openpyxl

def check_fill_color(excel_file):
    try:
        wb = openpyxl.load_workbook(excel_file)
        print("检查文件:", excel_file)
        print("-" * 50)

        # 遍历所有Sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\nSheet: {sheet_name}")

            # 查找wifi_format列
            wifi_format_col = None
            for idx, cell in enumerate(ws[1]):
                if cell.value == 'wifi_format':
                    wifi_format_col = idx + 1
                    break

            if wifi_format_col:
                print(f"找到wifi_format列: 第{wifi_format_col}列")

                # 检查前几行的填充色
                found_fill = False
                for row in range(2, min(10, ws.max_row + 1)):
                    cell = ws.cell(row=row, column=wifi_format_col)
                    if cell.fill.start_color.index != '00000000':
                        print(f"  第{row}行wifi_format: 填充色 {cell.fill.start_color.index}")
                        found_fill = True

                if found_fill:
                    print("填充色已成功应用")
                else:
                    print("未找到填充色")
            else:
                print("未找到wifi_format列")

        wb.close()

    except Exception as e:
        print(f"检查Excel文件时出错: {e}")

if __name__ == "__main__":
    excel_file = r"D:\chip_test\dev\xian_test\Xian-Esp-Test-Scripts\py_script_fpga_tx_wifi7\Log\wifi_tx_rls4\regression_v1.0\merged_tx_result_flatness_fail.xlsx"
    check_fill_color(excel_file)
