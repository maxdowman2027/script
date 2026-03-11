import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException

# ===================== 配置项（请根据需求修改）=====================
TARGET_PATH = r"D:\users\gxu\spur_scan\2G_high_mcs"  # 要检索的根文件夹路径（建议用r前缀避免转义）
FILE_PATTERN = "*spur_diff.xlsx"          # 文件名匹配格式（如"report_*.xlsx"匹配以report_开头的xlsx文件）
THRESHOLD_VALUE = 3            # 阈值：大于此值的数字会被标红
OVERWRITE_ORIGINAL = False       # 是否覆盖原文件（False则另存为"原文件名_标红.xlsx"）

# ===================== 核心功能函数 =====================
def mark_exceed_value_red(file_path):
    """
    处理单个xlsx文件：将大于阈值的数字单元格标红
    :param file_path: 目标文件的完整路径
    """
    try:
        # 加载工作簿（data_only=True确保读取公式计算后的实际值，而非公式本身）
        wb = load_workbook(file_path, data_only=True)
        # 定义红色字体样式
        red_font = Font(color="FF0000")  # RGB红色值
        
        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 遍历所有有数据的单元格（只遍历已使用区域，提升效率）
            for row in ws.iter_rows():
                for cell in row:
                    # 仅处理数字类型且值大于阈值的单元格
                    if isinstance(cell.value, (int, float)) and cell.value > THRESHOLD_VALUE:
                        cell.font = red_font  # 设置字体为红色
        
        # 保存文件（区分覆盖/另存为）
        if OVERWRITE_ORIGINAL:
            wb.save(file_path)
            print(f"✅ 已覆盖原文件：{file_path}")
        else:
            file_dir, file_name = os.path.split(file_path)
            name, ext = os.path.splitext(file_name)
            new_file_path = os.path.join(file_dir, f"{name}_标红{ext}")
            wb.save(new_file_path)
            print(f"✅ 已另存为标红文件：{new_file_path}")
            
        wb.close()  # 关闭工作簿释放资源
        
    except InvalidFileException:
        print(f"❌ 文件格式错误，无法处理：{file_path}")
    except PermissionError:
        print(f"❌ 权限不足，无法访问/修改文件：{file_path}")
    except Exception as e:
        print(f"❌ 处理文件失败 {file_path}：{str(e)}")

def main():
    """主函数：递归检索文件并批量处理"""
    # 拼接递归搜索的路径模式：**/ 表示匹配所有子目录
    search_pattern = os.path.join(TARGET_PATH, "**", FILE_PATTERN)
    # 关键：设置recursive=True实现递归搜索
    target_files = glob.glob(search_pattern, recursive=True)
    
    # 检查是否找到文件
    if not target_files:
        print(f"⚠️ 未在路径 {TARGET_PATH} 及其子文件夹下找到符合 {FILE_PATTERN} 格式的xlsx文件")
        return
    
    print(f"🔍 共找到 {len(target_files)} 个符合条件的文件（含子文件夹），开始处理...")
    # 批量处理每个文件
    for file in target_files:
        mark_exceed_value_red(file)
    
    print("\n🎉 所有文件处理完成！")

if __name__ == "__main__":
    # 先检查目标路径是否存在
    if not os.path.exists(TARGET_PATH):
        print(f"❌ 错误：指定的路径 {TARGET_PATH} 不存在，请检查配置项！")
    else:
        main()