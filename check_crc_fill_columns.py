import openpyxl

def check_crc_fill_columns():
    file_path = "D:/chip_test/dev/xian_test/Xian-Esp-Test-Scripts/py_script_fpga_tx_wifi7/Log/wifi_tx_rls4/no_he/crc_fail_result.xlsx"
    priority_columns = ['tx_power_set(dBm)', 'evm', 'evm_nss0', 'evm_nss1']

    try:
        # 打开工作簿
        wb = openpyxl.load_workbook(file_path)

        # 获取第一个Sheet
        sheet_name = "channel11_BCC"
        ws = wb[sheet_name]

        print(f"在 {sheet_name} 中检查列顺序和填充色:")
        print("-" * 50)

        # 显示列顺序
        print("列顺序:")
        for col_idx, cell in enumerate(ws[1]):
            print(f"{col_idx + 1}. {cell.value}")

        print("\n" + "-" * 50)

        # 显示前10行的内容
        print("前10行的内容:")
        for row_idx in range(2, min(11, ws.max_row + 1)):
            row_values = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_values.append(str(cell.value))
            print(f"行 {row_idx - 1}: {', '.join(row_values[:5])}...")

        print("\n" + "-" * 50)

        # 查找wifi_format列的索引
        wifi_format_index = None
        for idx, cell in enumerate(ws[1]):
            if cell.value == "wifi_format":
                wifi_format_index = idx
                break

        if wifi_format_index is not None:
            print(f"找到wifi_format列，索引为: {wifi_format_index}")
        else:
            print("未找到wifi_format列")

        # 检查填充色
        colors_used = {}
        for row_idx in range(2, min(11, ws.max_row + 1)):
            cell_value = ws.cell(row=row_idx, column=wifi_format_index + 1).value
            row_fill = None
            for col_idx in range(len(priority_columns) + 1, ws.max_column + 1):
                cell_fill = ws.cell(row=row_idx, column=col_idx).fill
                if cell_fill.start_color.index != "00000000" and cell_fill.start_color.index != "00FFFF99":
                    row_fill = cell_fill.start_color.index
                    break

            if cell_value and row_fill:
                if cell_value not in colors_used:
                    colors_used[cell_value] = row_fill
                print(f"行 {row_idx - 1}: 格式='{cell_value}', 颜色='{row_fill}'")

        print("\n" + "-" * 50)
        print(f"检测到的格式与颜色对应关系:")
        for format_name, color in colors_used.items():
            print(f"格式: '{format_name}', 颜色: '{color}'")

        wb.close()

    except Exception as e:
        print(f"读取文件失败: {e}")

if __name__ == "__main__":
    check_crc_fill_columns()