#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RX sensitivity calculation (same algorithm as wifiRxPlot.py), CSV output only.

Interpolates sensitivity (dBm) at PER 8% (11b) or 10% (other) vs rfpwr sweep per rate.
"""

from __future__ import annotations

import glob
import math
import os
import re
from collections import defaultdict
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

import numpy as np
import pandas as pd

# mld_en 宽表合并（与 organize_sensitivity_mld_diff.py 共用）
MLD_CONFIG_MATCH_COLS = ("bandwidth", "coding", "wifi_format", "cur_degree")
MLD_PAIR_EXTRA_COLS = ("rate", "rx_chan")
SENS_MLD_EN0_COL = "sensitivity_dbm_mld_en0"
SENS_MLD_EN1_COL = "sensitivity_dbm_mld_en1"
SENS_MLD_DIFF_COL = "sensitivity_dbm_mld_diff"
_MLD_DROP_ON_MERGE = frozenset(
    {"mld_en", "sensitivity_dbm", "testcase_folder", "rx_session_dir"}
)

_MLD_EN_RE = re.compile(r"mld_en(\d+)", re.I)
_CUR_DEGREE_RE = re.compile(r"cur_degree(\d+)", re.I)


def parse_testcase_folder_params(
    testcase_folder: Optional[str],
) -> Tuple[str, str]:
    """
    Extract mld_en and cur_degree from testcase folder basename.

    Example: wifi_txrx_test_RXSens_..._mld_en0_cur_degree45 -> ("0", "45")
    """
    if not testcase_folder:
        return "", ""
    mld = _MLD_EN_RE.search(testcase_folder)
    deg = _CUR_DEGREE_RE.search(testcase_folder)
    return (mld.group(1) if mld else "", deg.group(1) if deg else "")


def testcase_params_from_path_config(path_config: Mapping[str, Any]) -> Tuple[str, str]:
    """mld_en / cur_degree from path_config, with fallback parse on testcase_folder."""
    mld = str(path_config.get("mld_en") or "").strip()
    deg = str(path_config.get("cur_degree") or "").strip()
    if mld and deg:
        return mld, deg
    return parse_testcase_folder_params(path_config.get("testcase_folder"))


def _strip_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [str(c).strip() for c in out.columns]
    return out


def _require_columns(df: pd.DataFrame, names: Sequence[str]) -> None:
    missing = [n for n in names if n not in df.columns]
    if missing:
        raise ValueError(f"RX CSV missing columns {missing}; have {list(df.columns)}")


def infer_testcase_label(df: pd.DataFrame, testcase_folder: Optional[str], wifi_format: Optional[str]) -> str:
    """Guess 11b vs other for PER threshold (wifiRxPlot uses '11b' in testcase name)."""
    blob = " ".join(
        [
            str(testcase_folder or ""),
            str(wifi_format or ""),
            " ".join(str(x) for x in df.get("rate", pd.Series(dtype=object)).dropna().unique()),
        ]
    ).lower()
    if "11b" in blob or "dsss" in blob:
        return "11b"
    return wifi_format or testcase_folder or "default"


def calc_sensitivity_dbm(
    per: Sequence[float],
    power: Sequence[float],
    *,
    testcase: str,
    is_acr: bool,
    sens_accuracy: int = 100,
) -> float:
    """
    One rate column: interpolate power at target PER (wifiRxPlot logic).
    Returns sensitivity in dBm; 0 if not found.
    """
    per_list = list(per)
    pow_list = list(power)
    n = len(per_list)
    if n == 0:
        return 0.0

    is_11b = "11b" in testcase.lower()
    pow_sens_result = 0.0

    for i in range(n):
        if is_11b:
            if is_acr:
                if per_list[i] > 0.08:
                    if i == 0:
                        return 0.0
                    if per_list[i - 1] == 0:
                        delta_per = (math.log10(per_list[i]) - 0.0001) / sens_accuracy
                    else:
                        delta_per = (math.log10(per_list[i]) - math.log10(per_list[i - 1])) / sens_accuracy
                    per_sens = math.log10(per_list[i])
                    pow_sens = float(pow_list[i])
                    for _ in range(sens_accuracy):
                        per_sens -= delta_per
                        pow_sens -= 1.0 / sens_accuracy
                        if per_sens <= -1.096:
                            pow_sens_result = pow_sens
                            break
                    break
            else:
                if per_list[i] < 0.08:
                    if per_list[i] == 0:
                        if i == 0:
                            return 0.0
                        delta_per = (math.log10(per_list[i - 1]) + 10) / sens_accuracy
                        per_sens = -10.0
                    else:
                        if i == 0:
                            return 0.0
                        delta_per = (math.log10(per_list[i - 1]) - math.log10(per_list[i])) / sens_accuracy
                        per_sens = math.log10(per_list[i])
                    pow_sens = float(pow_list[i])
                    for _ in range(sens_accuracy):
                        per_sens += delta_per
                        pow_sens -= 1.0 / sens_accuracy
                        if per_sens >= -1.096:
                            pow_sens_result = pow_sens
                            break
                    break
        else:
            if is_acr:
                if per_list[i] > 0.1:
                    if i == 0:
                        return 0.0
                    if per_list[i - 1] == 0:
                        delta_per = (math.log10(per_list[i]) + 2) / sens_accuracy
                    else:
                        delta_per = (math.log10(per_list[i]) - math.log10(per_list[i - 1])) / sens_accuracy
                    per_sens = math.log10(per_list[i])
                    pow_sens = float(pow_list[i])
                    for _ in range(sens_accuracy):
                        per_sens -= delta_per
                        pow_sens -= 1.0 / sens_accuracy
                        if per_sens <= -1.0:
                            pow_sens_result = pow_sens
                            break
                    break
            else:
                if per_list[i] < 0.1:
                    if per_list[i] == 0:
                        if i == 0:
                            return 0.0
                        if per_list[i - 1] == 0:
                            delta_per = 0.0
                            per_sens = -10.0
                        else:
                            delta_per = (math.log10(per_list[i - 1]) + 10) / sens_accuracy
                            per_sens = -10.0
                    else:
                        if i == 0:
                            return 0.0
                        delta_per = (math.log10(per_list[i - 1]) - math.log10(per_list[i])) / sens_accuracy
                        per_sens = math.log10(per_list[i])
                    pow_sens = float(pow_list[i])
                    for _ in range(sens_accuracy):
                        per_sens += delta_per
                        pow_sens -= 1.0 / sens_accuracy
                        if per_sens >= -1.0:
                            pow_sens_result = pow_sens
                            break
                    break

    return round(pow_sens_result, 2)


def load_rx_session_dataframe(session_dir: str, csv_glob: str = "*.csv") -> pd.DataFrame:
    """Merge all RX CSV files in one directory (same as wifiRxPlot per testcase folder)."""
    paths = sorted(glob.glob(os.path.join(session_dir, csv_glob)))
    if not paths:
        raise FileNotFoundError(f"No CSV in session dir: {session_dir}")

    frames = []
    for p in paths:
        frames.append(pd.read_csv(p, index_col=False))
    df = pd.concat(frames, ignore_index=True)
    return _strip_columns(df)


def sensitivity_rows_for_session(
    session_dir: str,
    path_config: Mapping[str, Any],
    *,
    pak_num: int = 1000,
    sens_accuracy: int = 100,
    csv_glob: str = "*.csv",
) -> List[Dict[str, Any]]:
    """
    Compute one row per (rx_chan, rate) with path config + sensitivity_dbm.
    """
    df = load_rx_session_dataframe(session_dir, csv_glob=csv_glob)
    _require_columns(df, ["rx_chan", "rxnum", "rfpwr", "rate"])

    testcase = infer_testcase_label(
        df,
        path_config.get("testcase_folder"),
        path_config.get("wifi_format"),
    )
    is_acr = "acr" in testcase.lower() or "aci" in testcase.lower()

    df = df.copy()
    df["per"] = df["rxnum"].map(lambda x: 1 - min(float(x), pak_num) / pak_num)

    index_col = "acr" if is_acr else "rfpwr"
    if index_col not in df.columns and is_acr:
        raise ValueError(f"ACR testcase but column 'acr' missing in {session_dir}")

    mld_en, cur_degree = testcase_params_from_path_config(path_config)
    rows: List[Dict[str, Any]] = []
    base = {
        "band": path_config.get("band") or "",
        "phymode": path_config.get("phymd") or "",
        "bandwidth": path_config.get("bandwidth") or "",
        "coding": path_config.get("coding") or "",
        "wifi_format": path_config.get("wifi_format") or "",
        "testcase_folder": path_config.get("testcase_folder") or "",
        "mld_en": mld_en,
        "cur_degree": cur_degree,
        "config_tag": path_config.get("config_tag") or "",
        "rx_session_dir": os.path.abspath(session_dir),
        "testcase_label": testcase,
    }

    for chan in sorted(df["rx_chan"].dropna().unique(), key=lambda x: str(x)):
        df_chan = df[df["rx_chan"] == chan]
        table = pd.pivot_table(df_chan, index=[index_col], columns=["rate"], values=["per"], aggfunc="mean")
        column_convert = [col[1] if isinstance(col, tuple) else col for col in table.columns]
        table.columns = column_convert

        for rate in column_convert:
            per4pow = table[rate]
            sens = calc_sensitivity_dbm(
                per4pow.values,
                per4pow.index,
                testcase=testcase,
                is_acr=is_acr,
                sens_accuracy=sens_accuracy,
            )
            row = dict(base)
            row["rx_chan"] = chan
            row["rate"] = rate
            row["sensitivity_dbm"] = sens
            rows.append(row)

    return rows


def write_sensitivity_csv(rows: Sequence[Dict[str, Any]], out_path: str) -> None:
    """Write sensitivity result table."""
    columns = [
        "band",
        "phymode",
        "bandwidth",
        "coding",
        "wifi_format",
        "testcase_folder",
        "mld_en",
        "cur_degree",
        "config_tag",
        "rx_session_dir",
        "rx_chan",
        "testcase_label",
        "rate",
        "sensitivity_dbm",
    ]
    parent = os.path.dirname(os.path.abspath(out_path))
    if parent:
        os.makedirs(parent, exist_ok=True)
    frame = pd.DataFrame(rows)
    for col in columns:
        if col not in frame.columns:
            frame[col] = ""
    frame = frame[columns]
    frame.to_csv(out_path, index=False, encoding="utf-8-sig")


def _normalize_mld_en(series: pd.Series) -> pd.Series:
    def _one(v):
        if pd.isna(v):
            return np.nan
        s = str(v).strip()
        if s in ("0", "1"):
            return s
        try:
            i = int(float(s))
            if i in (0, 1):
                return str(i)
        except (TypeError, ValueError):
            pass
        return s

    return series.map(_one)


def merge_mld_sensitivity_rows(
    df: pd.DataFrame,
    config_cols: Sequence[str] = MLD_CONFIG_MATCH_COLS,
    extra_cols: Sequence[str] = MLD_PAIR_EXTRA_COLS,
) -> pd.DataFrame:
    """
    Long -> wide: one row per (bandwidth, coding, wifi_format, cur_degree, rate, rx_chan)
    when both mld_en=0 and mld_en=1 exist.
    """
    required = list(config_cols) + list(extra_cols) + ["mld_en", "sensitivity_dbm"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"CSV missing columns: {missing}")

    work = df.copy()
    work.columns = [str(c).strip() for c in work.columns]
    work["_mld_norm"] = _normalize_mld_en(work["mld_en"])
    work["_sens_num"] = pd.to_numeric(work["sensitivity_dbm"], errors="coerce")

    key_cols = list(config_cols) + list(extra_cols)
    wide = (
        work.dropna(subset=["_mld_norm"])
        .pivot_table(
            index=key_cols,
            columns="_mld_norm",
            values="_sens_num",
            aggfunc="first",
        )
    )
    if "0" not in wide.columns or "1" not in wide.columns:
        return pd.DataFrame()

    out = wide[["0", "1"]].copy()
    out.columns = [SENS_MLD_EN0_COL, SENS_MLD_EN1_COL]
    out[SENS_MLD_DIFF_COL] = out[SENS_MLD_EN0_COL] - out[SENS_MLD_EN1_COL]
    out = out.reset_index()

    meta_cols = [
        c
        for c in work.columns
        if c not in key_cols
        and c not in _MLD_DROP_ON_MERGE
        and not c.startswith("_")
    ]
    if meta_cols:
        meta0 = work[work["_mld_norm"] == "0"].drop_duplicates(subset=key_cols, keep="first")
        if meta0.empty:
            meta0 = work.drop_duplicates(subset=key_cols, keep="first")
        meta0 = meta0[key_cols + [c for c in meta_cols if c in meta0.columns]]
        out = out.merge(meta0, on=key_cols, how="left")

    front = list(config_cols) + list(extra_cols)
    meta_ordered = [
        c
        for c in out.columns
        if c not in front
        and c not in (SENS_MLD_EN0_COL, SENS_MLD_EN1_COL, SENS_MLD_DIFF_COL)
    ]
    tail = [SENS_MLD_EN0_COL, SENS_MLD_EN1_COL, SENS_MLD_DIFF_COL]
    return out[front + meta_ordered + tail]


def _mld_diff_fill(diff_value: float):
    from openpyxl.styles import PatternFill

    if diff_value >= 2.0:
        return PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    if diff_value >= 1.0:
        return PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    if diff_value >= -1.0:
        return PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    if diff_value >= -2.0:
        return PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    return PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


def write_mld_wide_outputs(df_out: pd.DataFrame, csv_path: str, xlsx_path: str) -> None:
    """Write wide CSV + XLSX with colored sensitivity_dbm_mld_diff column."""
    import openpyxl

    parent = os.path.dirname(os.path.abspath(csv_path))
    if parent:
        os.makedirs(parent, exist_ok=True)

    df_out.to_csv(csv_path, index=False, encoding="utf-8-sig")
    df_out.to_excel(xlsx_path, index=False, engine="openpyxl")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    diff_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == SENS_MLD_DIFF_COL:
            diff_col_idx = idx
            break
    if diff_col_idx is not None:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=diff_col_idx)
            if isinstance(cell.value, (int, float)) and not (
                isinstance(cell.value, float) and np.isnan(cell.value)
            ):
                cell.fill = _mld_diff_fill(float(cell.value))
    wb.save(xlsx_path)
    wb.close()


def write_mld_wide_from_long_rows(
    rows: Sequence[Dict[str, Any]],
    csv_path: str,
    xlsx_path: str,
) -> pd.DataFrame:
    """Merge long sensitivity rows to mld_en wide table; write csv/xlsx. Returns wide DataFrame."""
    if not rows:
        return pd.DataFrame()
    wide = merge_mld_sensitivity_rows(pd.DataFrame(rows))
    if wide.empty:
        return wide
    write_mld_wide_outputs(wide, csv_path, xlsx_path)
    return wide


def _safe_filename_part(value: Any, default: str = "na") -> str:
    text = str(value).strip() if value is not None else ""
    if not text:
        return default
    return re.sub(r"[^\w\-.]+", "_", text)[:80]


def _radar_radius_from_sensitivity(sensitivity_dbm: float) -> float:
    """Polar radius: 0 dBm sensitivity -> origin; else -sensitivity_dbm (larger = more sensitive)."""
    if sensitivity_dbm == 0.0:
        return 0.0
    return -sensitivity_dbm


def _radar_diff_point_color(diff_dbm: float) -> str:
    """Green if mld_en0 - mld_en1 >= 0 (outside zero ring), red if < 0 (inside)."""
    return "#2ca02c" if float(diff_dbm) >= 0 else "#d62728"


def _format_diff_tick_label(diff_db: float) -> str:
    """Format diff (dB) for radial reference labels."""
    if abs(diff_db) < 1e-9:
        return "0 dB"
    if diff_db < 0:
        return f"{diff_db:.1f} dB"
    return f"+{diff_db:.1f} dB"


def _mld_diff_radial_summary_ticks(
    diffs: Sequence[float],
    r0: float,
    r_lo: float,
    r_hi: float,
    *,
    merge_tol_db: float = 0.15,
) -> List[Tuple[float, str]]:
    """
    Key diff reference circles in [r_lo, r_hi].

    plot radius = r0 + diff: diff<0 inside zero ring (closer to center = more negative).
    Labels only min / median / max of measured diffs, plus 0 dB when the zero ring is visible.
    """
    d_sorted = sorted(float(d) for d in diffs)
    d_min, d_max = d_sorted[0], d_sorted[-1]
    n = len(d_sorted)
    if n % 2:
        d_med = d_sorted[n // 2]
    else:
        d_med = 0.5 * (d_sorted[n // 2 - 1] + d_sorted[n // 2])

    candidates: List[float] = []
    if r_lo <= r0 <= r_hi:
        candidates.append(0.0)
    for d in (d_min, d_med, d_max):
        if r_lo <= _mld_diff_plot_radius(r0, d) <= r_hi:
            candidates.append(d)

    picked: List[float] = []
    for d in sorted(candidates, key=lambda x: x):
        if any(abs(d - p) < merge_tol_db for p in picked):
            continue
        picked.append(d)

    return [(d, _format_diff_tick_label(d)) for d in picked]


def _apply_mld_diff_radial_ticks(
    ax,
    r0: float,
    tick_diffs: Sequence[float],
    *,
    label_position_deg: float = 40.0,
) -> None:
    """Set polar radial grid at r0+diff; labels show diff (dB) on the ring."""
    rticks = [_mld_diff_plot_radius(r0, d) for d in tick_diffs]
    if not rticks:
        return
    ax.set_rticks(rticks)
    ax.set_rlabel_position(label_position_deg)
    ax.set_yticklabels([_format_diff_tick_label(d) for d in tick_diffs], fontsize=7)
    ax.yaxis.grid(
        True,
        which="major",
        linestyle="-",
        linewidth=0.9,
        color="#888888",
        alpha=0.65,
    )


def _mld_diff_plot_radius(r0: float, diff: float) -> float:
    """Polar radius for diff (mld_en0 − mld_en1); diff=0 at r0."""
    return r0 + float(diff)


def _mld_diff_polar_layout(
    diffs: Sequence[float],
    *,
    margin_db: float = 0.5,
    min_data_span_db: float = 0.5,
) -> Tuple[float, float, float, List[float]]:
    """
    Linear map: plot_r = r0 + diff, with r0 = -d_min + pad.

    Innermost measured diff (d_min) sits at radius ``pad``; diff=0 at r0.
    Same formula for all-negative / all-positive / mixed data so points and
    reference rings share one coordinate system.
    """
    d_list = [float(d) for d in diffs]
    d_min = min(d_list)
    d_max = max(d_list)
    data_span = d_max - d_min
    pad = margin_db if data_span < 1e-9 else max(margin_db, 0.12 * data_span)

    r0 = -d_min + pad
    radii = [_mld_diff_plot_radius(r0, d) for d in d_list]
    r_lo = pad
    r_hi = pad + data_span + pad

    if r_hi - r_lo < min_data_span_db:
        mid = 0.5 * (r_lo + r_hi)
        r_lo = max(0.0, mid - min_data_span_db / 2.0)
        r_hi = mid + min_data_span_db / 2.0

    return r0, r_lo, r_hi, radii


def plot_sensitivity_mld_diff_radar(
    rows: Sequence[Dict[str, Any]],
    out_dir: str,
    *,
    diff_col: str = SENS_MLD_DIFF_COL,
) -> List[str]:
    """
    Polar charts: angle = cur_degree (°), signed radius = r0 + sensitivity_dbm_mld_diff.

    Dashed ring at r0 is diff=0 (mld_en0 − mld_en1). diff>0 outside (green, 优化);
    diff<0 inside (red, 恶化; smaller radius = more negative diff). Reference circles
    at measured min/median/max diffs (and 0 dB when visible) only.
    """
    import matplotlib.pyplot as plt
    from matplotlib.lines import Line2D

    groups: Dict[Tuple[Any, ...], List[Tuple[float, float]]] = defaultdict(list)
    for row in rows:
        deg = str(row.get("cur_degree") or "").strip()
        diff = row.get(diff_col)
        if not deg or diff in (None, ""):
            continue
        try:
            deg_f = float(deg)
            diff_f = float(diff)
        except (TypeError, ValueError):
            continue
        if math.isnan(diff_f):
            continue
        config_key = (
            row.get("band"),
            row.get("phymode"),
            row.get("bandwidth"),
            row.get("coding"),
            row.get("wifi_format"),
            row.get("rx_chan"),
            row.get("rate"),
        )
        groups[config_key].append((deg_f, diff_f))

    if not groups:
        return []

    out_dir = os.path.abspath(out_dir)
    os.makedirs(out_dir, exist_ok=True)
    written: List[str] = []

    for config_key, pts in sorted(groups.items(), key=lambda kv: str(kv[0])):
        (
            band,
            phymode,
            bandwidth,
            coding,
            wifi_format,
            rx_chan,
            rate,
        ) = config_key
        pts = sorted(pts, key=lambda x: x[0])
        if not pts:
            continue

        degrees = [p[0] for p in pts]
        diffs = [p[1] for p in pts]
        d_min = min(diffs)
        d_max = max(diffs)
        theta = np.deg2rad(degrees)
        r0, r_lo, r_hi, radius = _mld_diff_polar_layout(diffs)
        colors = [_radar_diff_point_color(d) for d in diffs]
        tick_diffs = [d for d, _lbl in _mld_diff_radial_summary_ticks(diffs, r0, r_lo, r_hi)]
        tick_radii = [_mld_diff_plot_radius(r0, d) for d in tick_diffs]

        fig, ax = plt.subplots(figsize=(9, 9), subplot_kw={"projection": "polar"})
        theta_bg = np.linspace(0, 2 * np.pi, 360)

        red_hi = min(r0, r_hi)
        if red_hi > r_lo:
            ax.fill_between(
                theta_bg,
                r_lo,
                np.full_like(theta_bg, red_hi),
                color="#ffcccc",
                alpha=0.35,
                zorder=0,
            )
        if r_hi > max(r0, r_lo):
            ax.fill_between(
                theta_bg,
                max(r0, r_lo),
                np.full_like(theta_bg, r_hi),
                color="#ccffcc",
                alpha=0.28,
                zorder=0,
            )

        for i in range(len(theta)):
            j = (i + 1) % len(theta)
            mid_diff = 0.5 * (diffs[i] + diffs[j])
            seg_color = _radar_diff_point_color(mid_diff)
            ax.plot(
                [theta[i], theta[j]],
                [radius[i], radius[j]],
                color=seg_color,
                linewidth=1.4,
                alpha=0.75,
                zorder=2,
            )
        ax.scatter(
            theta,
            radius,
            c=colors,
            s=58,
            zorder=4,
            edgecolors="black",
            linewidths=0.45,
        )

        ax.set_ylim(r_lo, r_hi * 1.02 if r_hi > r_lo else r_lo + 0.1)

        if r_lo <= r0 <= r_hi:
            ax.plot(
                theta_bg,
                np.full_like(theta_bg, r0),
                color="#333333",
                linestyle="--",
                linewidth=2.0,
                zorder=3,
            )

        for tick_diff, r_tick in zip(tick_diffs, tick_radii):
            if not (r_lo <= r_tick <= r_hi):
                continue
            is_zero = abs(tick_diff) < 1e-9
            if is_zero:
                continue
            ax.plot(
                theta_bg,
                np.full_like(theta_bg, r_tick),
                color="#888888",
                linestyle="-",
                linewidth=0.9,
                alpha=0.55,
                zorder=1,
            )

        _apply_mld_diff_radial_ticks(ax, r0, tick_diffs)

        if r0 > r_hi + 1e-9 and d_max <= 0.0:
            ax.text(
                np.deg2rad(0.0),
                r_hi * 0.96,
                f"0 dB ref ↑  (worst {d_min:.1f} dB at center)",
                fontsize=7,
                color="#333333",
                ha="center",
                va="top",
                zorder=5,
                clip_on=False,
            )
        ax.set_theta_zero_location("N")
        ax.set_theta_direction(-1)
        ax.set_thetagrids(degrees, labels=[f"{int(d)}°" for d in degrees])
        ax.set_title(
            "RX sensitivity mld diff vs angle\n"
            f"{band} {phymode} {bandwidth} {coding} {wifi_format} | "
            f"ch={rx_chan} rate={rate}\n"
            f"(plot radius = r0 + diff; r0 = diff 0; rings = min/med/max diff)",
            pad=20,
            fontsize=10,
        )

        ax.legend(
            handles=[
                Line2D([0], [0], color="#333333", linestyle="--", linewidth=2, label="diff = 0 dB"),
                Line2D(
                    [0],
                    [0],
                    marker="o",
                    color="w",
                    markerfacecolor="#2ca02c",
                    label="outside ring (diff > 0, better)",
                ),
                Line2D(
                    [0],
                    [0],
                    marker="o",
                    color="w",
                    markerfacecolor="#d62728",
                    label="inside ring (diff < 0, worse)",
                ),
            ],
            loc="upper right",
            bbox_to_anchor=(1.32, 1.1),
            fontsize=8,
        )

        fname = (
            f"radar_mld_diff_{_safe_filename_part(band)}_{_safe_filename_part(phymode)}"
            f"_{_safe_filename_part(bandwidth)}_{_safe_filename_part(coding)}"
            f"_{_safe_filename_part(wifi_format)}"
            f"_ch{_safe_filename_part(rx_chan)}_rate{_safe_filename_part(rate)}.png"
        )
        out_path = os.path.join(out_dir, fname)
        fig.savefig(out_path, dpi=150, bbox_inches="tight")
        plt.close(fig)
        written.append(out_path)

    return written


def _radar_mld_filename_tag(mld_values: Sequence[str]) -> str:
    """Filename tag for compared mld_en series on one chart."""
    ordered = sorted({str(m).strip() for m in mld_values if str(m).strip() != ""}, key=lambda x: (len(x), x))
    if not ordered:
        return "mld_na"
    if ordered == ["0", "1"]:
        return "mld0v1"
    return "mld" + "_".join(ordered)


def plot_sensitivity_radar(
    rows: Sequence[Dict[str, Any]],
    out_dir: str,
) -> List[str]:
    """
    [Legacy] Polar charts overlaying mld_en=0/1 with radius = −sensitivity_dbm.

    Prefer plot_sensitivity_mld_diff_radar() on wide-table rows after merge_mld_sensitivity_rows.

    One PNG per (band, phymode, bandwidth, coding, wifi_format, rx_chan, rate), with each
    mld_en (e.g. 0 and 1) overlaid on the same axes for comparison.

    Points exist only for angles present in data (missing angles are not interpolated;
    the polyline connects consecutive measured angles). sensitivity_dbm == 0 is drawn
    at the origin on that angle axis.
    """
    import matplotlib.pyplot as plt
    import numpy as np

    # config_key -> mld_en -> rows
    groups: Dict[Tuple[Any, ...], Dict[str, List[Dict[str, Any]]]] = defaultdict(
        lambda: defaultdict(list)
    )
    for row in rows:
        deg = str(row.get("cur_degree") or "").strip()
        sens = row.get("sensitivity_dbm")
        if not deg or sens in (None, ""):
            continue
        try:
            float(deg)
            float(sens)
        except (TypeError, ValueError):
            continue
        mld_en = str(row.get("mld_en") or "").strip() or "na"
        config_key = (
            row.get("band"),
            row.get("phymode"),
            row.get("bandwidth"),
            row.get("coding"),
            row.get("wifi_format"),
            row.get("rx_chan"),
            row.get("rate"),
        )
        groups[config_key][mld_en].append(row)

    if not groups:
        return []

    out_dir = os.path.abspath(out_dir)
    os.makedirs(out_dir, exist_ok=True)
    written: List[str] = []
    series_colors = plt.cm.tab10.colors

    for config_key, by_mld in sorted(groups.items(), key=lambda kv: str(kv[0])):
        (
            band,
            phymode,
            bandwidth,
            coding,
            wifi_format,
            rx_chan,
            rate,
        ) = config_key

        fig, ax = plt.subplots(figsize=(9, 9), subplot_kw={"projection": "polar"})
        all_degrees: set = set()
        mld_keys = sorted(by_mld.keys(), key=lambda x: (x == "na", x))

        for idx, mld_en in enumerate(mld_keys):
            grp = by_mld[mld_en]
            pts = sorted(
                [(float(r["cur_degree"]), float(r["sensitivity_dbm"])) for r in grp],
                key=lambda x: x[0],
            )
            if not pts:
                continue
            degrees = [p[0] for p in pts]
            sens_dbm = [p[1] for p in pts]
            all_degrees.update(degrees)
            theta = np.deg2rad(degrees)
            radius = [_radar_radius_from_sensitivity(s) for s in sens_dbm]
            theta_closed = np.append(theta, theta[0])
            radius_closed = np.append(radius, radius[0])
            color = series_colors[idx % len(series_colors)]
            label = f"mld_en={mld_en}"
            ax.plot(
                theta_closed,
                radius_closed,
                "o-",
                linewidth=2,
                markersize=6,
                color=color,
                label=label,
            )
            ax.fill(theta_closed, radius_closed, alpha=0.12, color=color)

        if not all_degrees:
            plt.close(fig)
            continue

        deg_list = sorted(all_degrees)
        ax.set_theta_zero_location("N")
        ax.set_theta_direction(-1)
        ax.set_thetagrids(deg_list, labels=[f"{int(d)}°" for d in deg_list])
        mld_tag = _radar_mld_filename_tag(mld_keys)
        ax.set_title(
            "RX sensitivity vs angle (mld_en compare)\n"
            f"{band} {phymode} {bandwidth} {coding} {wifi_format} | "
            f"ch={rx_chan} rate={rate}\n"
            f"(radius = −sensitivity_dbm, larger = more sensitive)",
            pad=20,
            fontsize=10,
        )
        ax.legend(loc="upper right", bbox_to_anchor=(1.25, 1.1), fontsize=9)

        fname = (
            f"radar_{_safe_filename_part(band)}_{_safe_filename_part(phymode)}"
            f"_{_safe_filename_part(bandwidth)}_{_safe_filename_part(coding)}"
            f"_{_safe_filename_part(wifi_format)}_{mld_tag}"
            f"_ch{_safe_filename_part(rx_chan)}_rate{_safe_filename_part(rate)}.png"
        )
        out_path = os.path.join(out_dir, fname)
        fig.savefig(out_path, dpi=150, bbox_inches="tight")
        plt.close(fig)
        written.append(out_path)

    return written
