import os
import fnmatch
import re
from openpyxl import load_workbook, Workbook

def merge_all_sheets_to_one(
    root_path,              # 要搜索的根路径
    folder_pattern,         # 文件夹命名匹配规则（通配符/正则）
    output_file,            # 合并后的输出文件路径
    xlsx_pattern="*.xlsx",  # XLSX文件匹配规则（默认匹配所有XLSX）
    use_regex_folder=False, # 文件夹是否用正则匹配
    use_regex_xlsx=False,   # XLSX文件是否用正则匹配
    sheet_name_col="原工作表名称"  # 新增列的列名（标注原sheet名）
):
    """
    1. 搜索指定路径下符合命名格式的文件夹
    2. 收集这些文件夹中的XLSX文件
    3. 合并所有XLSX文件的所有sheet到输出文件的单个sheet：
       - 新增一列标注每条数据的原sheet名称
       - 仅保留一份合并后的表头（包含新增列）
    """
    # 验证根路径是否存在
    if not os.path.exists(root_path):
        print(f"错误：指定的根路径 '{root_path}' 不存在！")
        return

    # 第一步：筛选符合条件的文件夹
    target_folders = []
    for item in os.listdir(root_path):
        item_path = os.path.join(root_path, item)
        if not os.path.isdir(item_path):
            continue
        # 匹配文件夹命名规则
        is_match = re.match(folder_pattern, item) is not None if use_regex_folder else fnmatch.fnmatch(item, folder_pattern)
        if is_match:
            target_folders.append(item_path)
            print(f"找到符合条件的文件夹：{item_path}")

    if not target_folders:
        print("未找到任何符合条件的文件夹！")
        return

    # 第二步：收集所有目标XLSX文件
    xlsx_files = []
    for folder in target_folders:
        try:
            for file_name in os.listdir(folder):
                file_path = os.path.join(folder, file_name)
                if not os.path.isfile(file_path) or not file_name.lower().endswith(".xlsx"):
                    continue
                # 匹配XLSX文件命名规则
                is_xlsx_match = re.match(xlsx_pattern, file_name) is not None if use_regex_xlsx else fnmatch.fnmatch(file_name, xlsx_pattern)
                if is_xlsx_match:
                    xlsx_files.append(file_path)
                    print(f"找到目标XLSX文件：{file_path}")
        except PermissionError:
            print(f"权限错误：无法访问文件夹 {folder}")
        except Exception as e:
            print(f"遍历文件夹 {folder} 时出错：{str(e)}")

    if not xlsx_files:
        print("未在目标文件夹中找到任何符合条件的XLSX文件！")
        return

    # 第三步：合并所有sheet到单个输出sheet
    # 创建输出工作簿，仅保留一个sheet
    wb_output = Workbook()
    ws_output = wb_output.active
    ws_output.title = "合并数据"  # 输出sheet名称
    header_written = False       # 标记是否已写入表头
    total_rows = 0               # 统计合并的行数（不含表头）
    total_files = 0              # 统计处理的XLSX文件数
    total_sheets = 0             # 统计处理的sheet数

    try:
        for xlsx_file in xlsx_files:
            try:
                # 加载输入XLSX文件（只读模式提升大文件性能）
                wb_input = load_workbook(xlsx_file, read_only=True, data_only=True)
                total_files += 1
                print(f"\n处理文件：{xlsx_file}")

                # 遍历当前文件的所有sheet
                for sheet_name in wb_input.sheetnames:
                    total_sheets += 1
                    ws_input = wb_input[sheet_name]
                    # 读取当前sheet的所有行数据（仅值）
                    rows = list(ws_input.iter_rows(values_only=True))

                    if not rows:
                        print(f"  跳过空工作表：{sheet_name}")
                        continue

                    # 处理表头：首次写入时添加"原工作表名称"列
                    if not header_written:
                        # 表头行新增一列（最前面），列名为sheet_name_col
                        header_row = [sheet_name_col] + list(rows[0])
                        ws_output.append(header_row)
                        header_written = True
                        # 数据行：跳过表头，每行开头加sheet名称
                        data_rows = rows[1:]
                    else:
                        # 非首次：直接跳过当前sheet的表头
                        data_rows = rows[1:] if len(rows) > 1 else []

                    # 写入数据行：每行开头添加原sheet名称
                    if data_rows:
                        for row in data_rows:
                            # 跳过空行（可选：根据需求保留/删除）
                            if all(cell is None for cell in row):
                                continue
                            # 行开头插入原sheet名称
                            new_row = [sheet_name] + list(row)
                            ws_output.append(new_row)
                        row_count = len(data_rows)
                        total_rows += row_count
                        print(f"  合并工作表 {sheet_name}：{row_count} 行数据")

                wb_input.close()  # 关闭输入文件，释放资源

            except Exception as e:
                print(f"读取 {xlsx_file} 时出错：{str(e)}，跳过该文件")

        # 保存合并后的文件
        wb_output.save(output_file)
        wb_output.close()

        # 输出统计信息
        print(f"\n=== 合并完成 ===")
        print(f"输出文件：{output_file}（工作表名称：合并数据）")
        print(f"共处理 {len(target_folders)} 个目标文件夹")
        print(f"共处理 {total_files} 个XLSX文件，{total_sheets} 个工作表")
        print(f"总计合并 {total_rows} 行数据（不含表头）")
        print(f"新增列名：{sheet_name_col}（标注每条数据的原工作表名称）")

    except PermissionError:
        print(f"权限错误：无法写入输出文件 {output_file}（可能被Excel占用）")
    except Exception as e:
        print(f"合并过程中出错：{str(e)}")

# ===================== 配置区（请根据你的需求修改）=====================
if __name__ == "__main__":
    # 1. 要搜索的根路径（替换为你的实际路径）
    ROOT_PATH =  r"D:\users\gxu\spur_scan\260309\5G\80m\vht"  # Windows示例
    # ROOT_PATH = "/home/yourname/data_folders"  # Linux/Mac示例
    
    # 2. 文件夹命名匹配规则（示例：匹配以"report_"开头的文件夹）
    FOLDER_PATTERN = "notch_enable0"  # 通配符示例
    # FOLDER_PATTERN = r"^data_\d{4}$"  # 正则示例（如data_2024、data_2025）
    
    # 3. 合并后的输出文件路径
    OUTPUT_FILE = r"D:\users\gxu\spur_scan\260309\5G\80m\vht\5G_80m_vht_notch_enable0.xlsx"  # Windows示例
    # OUTPUT_FILE = "/home/yourname/merged_all.xlsx"  # Linux/Mac示例
    
    # 4. XLSX文件匹配规则（默认匹配所有XLSX，可自定义如"sales_*.xlsx"）
    XLSX_PATTERN = "*sens_202639_*.xlsx"

    USE_REGEX_FOLDER = False
    USE_REGEX_XLSX = False

    SHEET_NAME_COL = "原工作表名称"

    # 执行合并函数
    merge_all_sheets_to_one(
        root_path=ROOT_PATH,
        folder_pattern=FOLDER_PATTERN,
        output_file=OUTPUT_FILE,
        xlsx_pattern=XLSX_PATTERN,
        use_regex_folder=USE_REGEX_FOLDER,
        use_regex_xlsx=USE_REGEX_XLSX,
        sheet_name_col=SHEET_NAME_COL
    )

