import os
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException

# ===================== 配置项（请根据需求修改）=====================
TARGET_PATH = r"D:\users\gxu\spur_scan\260311\6G\80m\vht"  # 要检索的根文件夹路径
FILE_PATTERN = "*spur.xlsx"                           # 查找的文件模式
THRESHOLD_VALUE = 3                                   # 差值超过此阈值的单元格会被标红
BASE_ROW_IDX = 0                                     # 基准行（0=第一行）
START_COL_IDX = 1                                    # 起始列（0=第一列）

def calc_diff_from_spec_col_xlsx(file_path, base_row_idx=0, start_col_idx=0, output_path=None):
    """
    计算XLSX表格所有行与指定行的差值（仅从指定列数开始计算），完全保留原文件格式
    :param file_path: XLSX文件路径
    :param base_row_idx: 基准行（int，从0开始计数）
    :param start_col_idx: 起始列（int，从0开始计数）
    :param output_path: 结果保存路径（None则不保存）
    :return: 包含原列+差值列的DataFrame
    """
    if not file_path.endswith((".xlsx", ".xls")):
        raise ValueError("仅支持Excel文件（.xlsx/.xls）")

    # 使用 openpyxl 直接读取 Excel 文件，保留格式
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        # 读取数据到 DataFrame
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        df = pd.DataFrame(data[1:], columns=data[0])

    except Exception as e:
        print(f"❌ 读取XLSX文件失败 {file_path}：{e}")
        return None

    if base_row_idx < 0 or base_row_idx >= len(df):
        print(f"⚠️ 基准行号{base_row_idx}无效（共{len(df)}行），自动使用第一行")
        base_row_idx = 0
    base_row = df.iloc[base_row_idx]

    total_cols = len(df.columns)
    if start_col_idx < 0 or start_col_idx >= total_cols:
        print(f"⚠️ 起始列号{start_col_idx}无效（共{total_cols}列），自动使用第一列")
        start_col_idx = 0

    target_cols_all = df.columns[start_col_idx:]
    df_target_cols = df[target_cols_all]

    numeric_cols = df_target_cols.select_dtypes(include=['int64', 'float64']).columns
    if len(numeric_cols) == 0:
        print(f"⚠️ 文件 {file_path} 指定列范围內无数值列，无需计算差值")
        return df

    diff_df = df_target_cols[numeric_cols].sub(base_row[numeric_cols], axis=1)
    diff_df.columns = [f"{col}_diff" for col in numeric_cols]

    result_df = pd.concat([df, diff_df], axis=1).fillna(0)

    if output_path:
        if not output_path.endswith(".xlsx"):
            output_path = output_path + ".xlsx"
        try:
            # 复制原文件的格式并写入数据
            wb_save = load_workbook(file_path)
            ws_save = wb_save.active

            # 添加差值列标题
            for i, col_name in enumerate(diff_df.columns):
                col_index = total_cols + i + 1  # Excel列从1开始
                ws_save.cell(row=1, column=col_index).value = col_name

            # 写入差值数据
            for row_idx in range(len(result_df)):
                for col_idx, col_name in enumerate(diff_df.columns):
                    cell_value = result_df.at[row_idx, col_name]
                    ws_save.cell(row=row_idx+2, column=total_cols+col_idx+1).value = cell_value

            wb_save.save(output_path)
            print(f"✅ 差值文件已生成：{output_path}")
        except Exception as e:
            print(f"❌ 保存差值文件失败 {output_path}：{e}")

    return result_df

def mark_exceed_value_red(file_path, threshold=3):
    """
    将Excel文件中大于阈值的数字单元格标红，保留原文件的填充色
    :param file_path: 目标文件的完整路径
    :param threshold: 阈值
    """
    try:
        wb = load_workbook(file_path, data_only=True)
        red_font = Font(color="FF0000")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (int, float)) and cell.value > threshold:
                        cell.font = red_font

        wb.save(file_path)
        print(f"✅ 文件标红完成：{file_path}")

        wb.close()

    except InvalidFileException:
        print(f"❌ 文件格式错误，无法处理：{file_path}")
    except PermissionError:
        print(f"❌ 权限不足，无法访问/修改文件：{file_path}")
    except Exception as e:
        print(f"❌ 处理文件失败 {file_path}：{str(e)}")

def main():
    """主函数：递归检索文件并批量处理"""
    search_pattern = os.path.join(TARGET_PATH, "**", FILE_PATTERN)
    target_files = glob.glob(search_pattern, recursive=True)

    if not target_files:
        print(f"⚠️ 未在路径 {TARGET_PATH} 及其子文件夹下找到符合 {FILE_PATTERN} 格式的xlsx文件")
        return

    print(f"🔍 共找到 {len(target_files)} 个符合条件的文件，开始处理...")

    for file in target_files:
        print(f"\n=== 处理文件：{file} ===")

        # 生成差值文件
        file_dir, file_name = os.path.split(file)
        name, ext = os.path.splitext(file_name)
        diff_output_path = os.path.join(file_dir, f"{name}_diff{ext}")

        result_df = calc_diff_from_spec_col_xlsx(
            file_path=file,
            base_row_idx=BASE_ROW_IDX,
            start_col_idx=START_COL_IDX,
            output_path=diff_output_path
        )

        if result_df is not None:
            mark_exceed_value_red(diff_output_path, threshold=THRESHOLD_VALUE)

    print("\n🎉 所有文件处理完成！")

if __name__ == "__main__":
    if not os.path.exists(TARGET_PATH):
        print(f"❌ 错误：指定的路径 {TARGET_PATH} 不存在，请检查配置项！")
    else:
        main()
