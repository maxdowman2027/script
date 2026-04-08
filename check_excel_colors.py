
import openpyxl
import sys

def check_excel_colors(file_path, sheet_name):
    try:
        # 加载工作簿和工作表
        wb = openpyxl.load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' not found in {file_path}")
            return False

        ws = wb[sheet_name]

        # 找到wifi_format列的索引
        wifi_format_col = None
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value == "wifi_format":
                wifi_format_col = col
                break

        if wifi_format_col is None:
            print(f"Column 'wifi_format' not found in Sheet '{sheet_name}'")
            return False

        print(f"Found 'wifi_format' column at index: {wifi_format_col}")

        # 统计不同wifi_format的填充色
        format_colors = {}
        format_counts = {}

        for row in range(2, ws.max_row + 1):
            # 获取wifi_format值
            format_value = ws.cell(row=row, column=wifi_format_col).value
            if format_value is None:
                continue

            # 获取第一列的填充色
            cell = ws.cell(row=row, column=1)
            if cell.fill.start_color.index != '00000000':
                # 有填充色
                color_code = cell.fill.start_color.index[2:]  # 移除前面的00
                if format_value not in format_colors:
                    format_colors[format_value] = set()
                    format_counts[format_value] = 0
                format_colors[format_value].add(color_code)
                format_counts[format_value] += 1

        # 打印结果
        print("\n统计结果:")
        print("=" * 50)

        for format_value in sorted(format_colors.keys()):
            colors = format_colors[format_value]
            count = format_counts[format_value]
            print(f"Format: {format_value}")
            print(f"  填充色: {', '.join(colors)}")
            print(f"  行数: {count}")

        print("=" * 50)

        # 检查hesu和heer是否有填充色
        print("\n检查hesu和heer格式的填充色:")
        if 'hesu' in format_colors:
            print(f"✓ hesu 格式有填充色: {', '.join(format_colors['hesu'])}")
        else:
            print(f"✗ hesu 格式未找到填充色")

        if 'heer' in format_colors:
            print(f"✓ heer 格式有填充色: {', '.join(format_colors['heer'])}")
        else:
            print(f"✗ heer 格式未找到填充色")

        return True

    except Exception as e:
        print(f"Error: {e}")
        return False


if __name__ == "__main__":
    # 文件路径
    excel_file = r"D:\chip_test\dev\xian_test\Xian-Esp-Test-Scripts\py_script_fpga_tx_wifi7\Log\wifi_tx_rls4\regression_v1.0\merged_tx_result.xlsx"

    # 检查第一个Sheet
    sheet_names = ["channel11_BCC_NSS1", "channel5180_BCC_NSS1"]

    print(f"检查文件: {excel_file}")
    print()

    for sheet_name in sheet_names:
        print(f"\n检查Sheet: {sheet_name}")
        print("-" * 30)
        check_excel_colors(excel_file, sheet_name)
