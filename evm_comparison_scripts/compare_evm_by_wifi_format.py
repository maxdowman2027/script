import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import os
import numpy as np
from datetime import datetime


def main():
    # 文件路径 - RLS4.0版本的测试结果
    file_path = r"D:\chip_test\dev\xian_test\Xian-Esp-Test-Scripts\py_script_fpga_tx_wifi7\Log\wifi_tx_rls4\vht_ht_hesu_compare\merged_tx_result.xlsx"

    # 输出目录
    output_dir = r"D:\users\gxu\scripts\evm_comparison_scripts\evm_by_wifi_format_comparison"
    os.makedirs(output_dir, exist_ok=True)

    print("=== Start analyzing RLS4.0 version data ===")
    try:
        xls = pd.ExcelFile(file_path)
        print(f"File contains {len(xls.sheet_names)} Sheets: {xls.sheet_names}")
    except Exception as e:
        print(f"Failed to read file: {e}")
        return

    # 读取所有Sheet数据
    sheet_data = {}
    for sheet in xls.sheet_names:
        try:
            df = pd.read_excel(xls, sheet_name=sheet)
            sheet_data[sheet] = df
            print(f"\n{sheet} Sheet info:")
            print(f"Record count: {len(df)}")
            print(f"Column names: {list(df.columns)}")
            print(f"Format types: {df['wifi_format'].unique() if 'wifi_format' in df.columns else 'N/A'}")
            print(f"Rate types: {df['rate'].unique() if 'rate' in df.columns else 'N/A'}")
            print(f"TX Power range: {df['tx_power_set(dBm)'].min(), df['tx_power_set(dBm)'].max() if 'tx_power_set(dBm)' in df.columns else 'N/A'}")
            print(f"EVM column: {'evm' if 'evm' in df.columns else 'N/A'}")
        except Exception as e:
            print(f"Failed to read {sheet} Sheet: {e}")

    # 对每个Sheet进行分析
    comparison_results = {}
    for sheet_name, df in sheet_data.items():
        print(f"\n=== Analyzing Sheet: {sheet_name} ===")
        result = analyze_sheet(df, sheet_name, output_dir)
        if result is not None:
            comparison_results[sheet_name] = result

    # 保存整体比较结果
    save_overall_results(comparison_results, output_dir)

    print(f"\n分析完成！所有结果已保存在: {output_dir}")


def analyze_sheet(df, sheet_name, output_dir):
    # Ensure key columns exist
    required_cols = ['wifi_format', 'rate', 'tx_power_set(dBm)', 'evm']
    missing_cols = []
    for col in required_cols:
        if col not in df.columns:
            missing_cols.append(col)

    if missing_cols:
        print(f"Warning: Missing key columns {', '.join(missing_cols)}, cannot complete comparison")
        return None

    # 获取所有唯一的wifi_format
    wifi_formats = df['wifi_format'].unique()
    print(f"WiFi formats in {sheet_name}: {wifi_formats}")

    # 对每个wifi_format组合进行比较，并将结果存入不同的DataFrame中
    comparison_data = {}
    all_tx_powers = df['tx_power_set(dBm)'].unique()
    all_rates = df['rate'].unique()

    # 获取所有唯一的wifi_format对
    wifi_formats = df['wifi_format'].unique()
    for i in range(len(wifi_formats)):
        for j in range(i+1, len(wifi_formats)):
            format_pair = f"{wifi_formats[i]}_vs_{wifi_formats[j]}"
            comparison_data[format_pair] = []

    # 对每个rate和tx_power的组合进行比较
    for rate in all_rates:
        for tx_pwr in all_tx_powers:
            # 过滤出相同rate和tx_pwr的所有wifi_format的数据
            filtered_df = df[(df['rate'] == rate) & (df['tx_power_set(dBm)'] == tx_pwr)]

            if len(filtered_df) > 1 and len(filtered_df['wifi_format'].unique()) > 1:
                # 比较不同wifi_format之间的EVM
                formats = filtered_df['wifi_format'].unique()
                for i in range(len(formats)):
                    for j in range(i+1, len(formats)):
                        base_format = formats[i]
                        compare_format = formats[j]
                        format_pair = f"{base_format}_vs_{compare_format}"

                        base_evm = filtered_df[filtered_df['wifi_format'] == base_format]['evm'].mean()
                        compare_evm = filtered_df[filtered_df['wifi_format'] == compare_format]['evm'].mean()

                        evm_diff = compare_evm - base_evm
                        abs_diff = abs(evm_diff)

                        comparison_data[format_pair].append({
                            'rate': rate,
                            'tx_power_set(dBm)': tx_pwr,
                            'base_format': base_format,
                            'compare_format': compare_format,
                            'evm_base': base_evm,
                            'evm_compare': compare_evm,
                            'evm_diff': evm_diff,
                            'abs_diff': abs_diff,
                            'count': len(filtered_df),
                            'formats_count': len(formats)
                        })

    # 检查是否有任何比较数据
    has_data = False
    for format_pair in comparison_data:
        if comparison_data[format_pair]:
            has_data = True
            break

    if not has_data:
        print(f"No comparisons found in {sheet_name}")
        return None

    # 收集所有格式对的比较结果
    all_comparison_dfs = {}
    all_stats_dfs = {}
    all_significant_diffs = {}

    for format_pair in comparison_data:
        if comparison_data[format_pair]:
            df_pair = pd.DataFrame(comparison_data[format_pair])
            all_comparison_dfs[format_pair] = df_pair

            # 计算统计结果
            stats = df_pair.groupby(['base_format', 'compare_format', 'rate']).agg(
                count=('evm_diff', 'count'),
                mean_diff=('evm_diff', 'mean'),
                median_diff=('evm_diff', 'median'),
                std_diff=('evm_diff', 'std'),
                min_diff=('evm_diff', 'min'),
                max_diff=('evm_diff', 'max'),
                mean_abs_diff=('abs_diff', 'mean'),
                base_mean_evm=('evm_base', 'mean'),
                base_median_evm=('evm_base', 'median'),
                base_std_evm=('evm_base', 'std'),
                compare_mean_evm=('evm_compare', 'mean'),
                compare_median_evm=('evm_compare', 'median'),
                compare_std_evm=('evm_compare', 'std')
            ).reset_index()

            all_stats_dfs[format_pair] = stats
            all_significant_diffs[format_pair] = df_pair[df_pair['abs_diff'] > 2.0]

    # 保存详细比较结果（每个格式对一个sheet）
    detailed_file = os.path.join(output_dir, f'{sheet_name}_detailed.xlsx')
    with pd.ExcelWriter(detailed_file, engine='openpyxl') as writer:
        for format_pair in comparison_data:
            if comparison_data[format_pair]:
                df_pair = pd.DataFrame(comparison_data[format_pair])
                df_pair.to_excel(writer, sheet_name=format_pair, index=False)

    # 为EVM值和差值添加颜色填充
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.load_workbook(detailed_file)

    # 定义差异颜色填充规则
    def get_diff_fill(diff_value):
        if diff_value <= -2:
            return PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 绿色（明显更好）
        elif diff_value <= -1:
            return PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # 浅绿色（更好）
        elif diff_value <= 1:
            return PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 黄色（无显著差异）
        elif diff_value <= 2:
            return PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # 橙色（较差）
        else:
            return PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # 红色（明显较差）

    # 为每个sheet添加颜色填充
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 查找EVM差异列的索引
        diff_col_idx = None
        for idx, cell in enumerate(ws[1]):
            if cell.value == 'evm_diff':
                diff_col_idx = idx + 1

        if diff_col_idx:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=diff_col_idx)
                diff_value = cell.value
                if isinstance(diff_value, (int, float)):
                    cell.fill = get_diff_fill(diff_value)

    wb.save(detailed_file)

    # 统计结果（每个格式对一个sheet）
    summary_file = os.path.join(output_dir, f'{sheet_name}_summary.xlsx')
    with pd.ExcelWriter(summary_file, engine='openpyxl') as writer:
        for format_pair in comparison_data:
            if comparison_data[format_pair]:
                df_pair = pd.DataFrame(comparison_data[format_pair])

                stats = df_pair.groupby(['base_format', 'compare_format', 'rate']).agg(
                    count=('evm_diff', 'count'),
                    mean_diff=('evm_diff', 'mean'),
                    median_diff=('evm_diff', 'median'),
                    std_diff=('evm_diff', 'std'),
                    min_diff=('evm_diff', 'min'),
                    max_diff=('evm_diff', 'max'),
                    mean_abs_diff=('abs_diff', 'mean'),
                    base_mean_evm=('evm_base', 'mean'),
                    base_median_evm=('evm_base', 'median'),
                    base_std_evm=('evm_base', 'std'),
                    compare_mean_evm=('evm_compare', 'mean'),
                    compare_median_evm=('evm_compare', 'median'),
                    compare_std_evm=('evm_compare', 'std')
                ).reset_index()

                stats.to_excel(writer, sheet_name=format_pair, index=False)

    # 可视化（每个格式对一个图表）
    for format_pair, df_pair in all_comparison_dfs.items():
        stats = all_stats_dfs[format_pair]
        plot_comparison(stats, df_pair, sheet_name, output_dir, format_pair)

    # 找出差异较大的情况（每个格式对一个sheet）
    significant_diff_file = os.path.join(output_dir, f'{sheet_name}_significant_differences.xlsx')
    with pd.ExcelWriter(significant_diff_file, engine='openpyxl') as writer:
        total_significant = 0
        for format_pair in comparison_data:
            if comparison_data[format_pair]:
                df_pair = pd.DataFrame(comparison_data[format_pair])
                significant = df_pair[df_pair['abs_diff'] > 2.0]
                if not significant.empty:
                    significant.to_excel(writer, sheet_name=format_pair, index=False)
                    total_significant += len(significant)
                    print(f"Significant differences (>2dB) in {sheet_name} {format_pair}: {len(significant)} records")

    print(f"Total significant differences (>2dB) in {sheet_name}: {total_significant} records")

    # 收集所有格式对的比较结果
    all_comparison_dfs = {}
    all_stats_dfs = {}
    all_significant_diffs = {}

    for format_pair in comparison_data:
        if comparison_data[format_pair]:
            df_pair = pd.DataFrame(comparison_data[format_pair])
            all_comparison_dfs[format_pair] = df_pair

            # 计算统计结果
            stats = df_pair.groupby(['base_format', 'compare_format', 'rate']).agg(
                count=('evm_diff', 'count'),
                mean_diff=('evm_diff', 'mean'),
                median_diff=('evm_diff', 'median'),
                std_diff=('evm_diff', 'std'),
                min_diff=('evm_diff', 'min'),
                max_diff=('evm_diff', 'max'),
                mean_abs_diff=('abs_diff', 'mean'),
                base_mean_evm=('evm_base', 'mean'),
                base_median_evm=('evm_base', 'median'),
                base_std_evm=('evm_base', 'std'),
                compare_mean_evm=('evm_compare', 'mean'),
                compare_median_evm=('evm_compare', 'median'),
                compare_std_evm=('evm_compare', 'std')
            ).reset_index()

            all_stats_dfs[format_pair] = stats
            all_significant_diffs[format_pair] = df_pair[df_pair['abs_diff'] > 2.0]

    return {
        'sheet_name': sheet_name,
        'comparison_dfs': all_comparison_dfs,
        'stats_dfs': all_stats_dfs,
        'significant_diffs': all_significant_diffs
    }


def plot_comparison(stats, comparison_df, sheet_name, output_dir, format_pair):
    sheet_dir = os.path.join(output_dir, sheet_name, format_pair)
    os.makedirs(sheet_dir, exist_ok=True)

    # 1. EVM difference distribution histogram
    plt.figure(figsize=(12, 6))
    plt.hist(comparison_df['evm_diff'], bins=30, alpha=0.7, color='b')
    plt.axvline(comparison_df['evm_diff'].mean(), color='r', linestyle='--', label=f'Mean = {comparison_df["evm_diff"].mean():.2f}')
    plt.axvline(0, color='g', linestyle='-', label='No Difference')
    plt.title(f'{sheet_name} {format_pair} EVM Difference Distribution')
    plt.xlabel('EVM Difference (dB)')
    plt.ylabel('Number of Records')
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(os.path.join(sheet_dir, 'evm_diff_distribution.png'), dpi=150)
    plt.close()

    # 2. Average difference by rate
    plt.figure(figsize=(16, 8))
    pivot = stats.pivot(index='rate', columns=['base_format', 'compare_format'], values='mean_diff')
    sns.heatmap(pivot, annot=True, cmap='coolwarm', fmt='.2f', cbar_kws={'label': 'Average EVM Difference (dB)'})
    plt.title(f'{sheet_name} {format_pair} Average EVM Difference by Rate')
    plt.tight_layout()
    plt.savefig(os.path.join(sheet_dir, 'evm_diff_heatmap.png'), dpi=150)
    plt.close()

    # 3. EVM difference boxplot by rate
    plt.figure(figsize=(16, 8))
    sns.boxplot(x='rate', y='mean_diff', data=stats)
    plt.axhline(y=0, color='g', linestyle='-', label='No Difference')
    plt.title(f'{sheet_name} {format_pair} EVM Difference by Rate')
    plt.xlabel('Rate')
    plt.ylabel('EVM Difference (dB)')
    plt.xticks(rotation=45)
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(sheet_dir, 'evm_diff_by_rate.png'), dpi=150)
    plt.close()

    # 4. Significant differences scatter plot
    significant_df = comparison_df[comparison_df['abs_diff'] > 2.0]
    if len(significant_df) > 0:
        plt.figure(figsize=(12, 10))
        plt.scatter(significant_df['tx_power_set(dBm)'], significant_df['evm_diff'],
                    c=abs(significant_df['evm_diff']), cmap='coolwarm', alpha=0.6, s=20)
        plt.colorbar(label='Absolute EVM Difference (dB)')
        plt.xlabel('TX Power (dBm)')
        plt.ylabel('EVM Difference (dB)')
        plt.title(f'{sheet_name} {format_pair} Significant EVM Differences (>2dB) by TX Power')
        plt.grid(True)
        plt.tight_layout()
        plt.savefig(os.path.join(sheet_dir, 'significant_differences_scatter.png'), dpi=150)
        plt.close()


def save_overall_results(comparison_results, output_dir):
    # 生成HTML报告
    html_file = os.path.join(output_dir, 'evm_by_wifi_format_comparison_report.html')
    generate_html_report(comparison_results, output_dir, html_file)

    print(f"HTML报告已保存: {html_file}")


def generate_html_report(comparison_results, output_dir, output_file):
    # Generate HTML report
    html_content = '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>WiFi Format EVM Comparison Analysis</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                margin: 20px;
                background-color: #f5f5f5;
            }
            .container {
                max-width: 1200px;
                margin: 0 auto;
                background-color: white;
                padding: 20px;
                border-radius: 8px;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            }
            h1, h2, h3 {
                color: #333;
            }
            .section {
                margin-bottom: 30px;
                padding-bottom: 20px;
                border-bottom: 1px solid #ddd;
            }
            .summary-table {
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 20px;
                background-color: white;
            }
            .summary-table th, .summary-table td {
                padding: 12px;
                text-align: left;
                border-bottom: 1px solid #ddd;
            }
            .summary-table th {
                background-color: #f2f2f2;
            }
            .summary-table tr:hover {
                background-color: #f5f5f5;
            }
            .success {
                color: green;
            }
            .warning {
                color: orange;
            }
            .error {
                color: red;
            }
            .image-container {
                text-align: center;
                margin: 20px 0;
            }
            .image-container img {
                max-width: 100%;
                height: auto;
                border: 1px solid #ddd;
                border-radius: 4px;
            }
            .sheet-link {
                color: #1a73e8;
                text-decoration: none;
            }
            .sheet-link:hover {
                text-decoration: underline;
            }
            .stats {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
                gap: 20px;
                margin-bottom: 30px;
            }
            .stat-card {
                background-color: #f8f9fa;
                padding: 15px;
                border-radius: 6px;
                box-shadow: 0 1px 3px rgba(0,0,0,0.1);
            }
            .stat-card h4 {
                margin-top: 0;
                color: #555;
            }
            .stat-value {
                font-size: 24px;
                font-weight: bold;
                margin: 10px 0;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>WiFi Format EVM Comparison Analysis</h1>

            <div class="section">
                <h2>1. 整体对比统计</h2>
    '''

    # 计算整体统计信息
    total_records = 0
    total_comparisons = 0
    total_significant_diff = 0
    all_evm_diffs = []

    for sheet_name, result in comparison_results.items():
        # 遍历每个格式对的数据
        for format_pair, df_pair in result['comparison_dfs'].items():
            total_records += len(df_pair)
            total_comparisons += 1
            if format_pair in result['significant_diffs']:
                total_significant_diff += len(result['significant_diffs'][format_pair])
            all_evm_diffs.extend(df_pair['abs_diff'].tolist())

    avg_abs_diff = np.mean(all_evm_diffs) if all_evm_diffs else 0
    max_diff = np.max(all_evm_diffs) if all_evm_diffs else 0

    html_content += f'''
                <div class="stats">
                    <div class="stat-card">
                        <h4>总比较记录数</h4>
                        <div class="stat-value">{total_records}</div>
                    </div>
                    <div class="stat-card">
                        <h4>平均绝对差值</h4>
                        <div class="stat-value" style="color: {'green' if avg_abs_diff < 1 else 'orange' if avg_abs_diff < 2 else 'red'}">
                            {avg_abs_diff:.2f} dB
                        </div>
                    </div>
                    <div class="stat-card">
                        <h4>最大差值</h4>
                        <div class="stat-value" style="color: {'green' if max_diff < 1 else 'orange' if max_diff < 2 else 'red'}">
                            {max_diff:.2f} dB
                        </div>
                    </div>
                    <div class="stat-card">
                        <h4>显著差异数(>2dB)</h4>
                        <div class="stat-value" style="color: {'green' if total_significant_diff == 0 else 'orange' if total_significant_diff < 50 else 'red'}">
                            {total_significant_diff}
                        </div>
                    </div>
                </div>
            </div>

            <div class="section">
                <h2>2. Sheet级对比详情</h2>
                <table class="summary-table">
                    <thead>
                        <tr>
                            <th>Sheet Name</th>
                            <th>比较记录数</th>
                            <th>平均绝对差值</th>
                            <th>最大差值</th>
                            <th>显著差异数(>2dB)</th>
                            <th>详细结果</th>
                        </tr>
                    </thead>
                    <tbody>
    '''

    for sheet_name, result in comparison_results.items():
        # 计算当前sheet的整体平均和最大值
        all_abs_diffs = []
        sheet_significant = 0
        sheet_total_records = 0

        for format_pair, df_pair in result['comparison_dfs'].items():
            all_abs_diffs.extend(df_pair['abs_diff'].tolist())
            sheet_total_records += len(df_pair)
            if format_pair in result['significant_diffs']:
                sheet_significant += len(result['significant_diffs'][format_pair])

        sheet_avg_abs = np.mean(all_abs_diffs) if all_abs_diffs else 0
        sheet_max_diff = np.max(all_abs_diffs) if all_abs_diffs else 0

        avg_class = 'success' if sheet_avg_abs < 1 else 'warning' if sheet_avg_abs < 2 else 'error'
        max_class = 'success' if sheet_max_diff < 1 else 'warning' if sheet_max_diff < 2 else 'error'
        sig_class = 'success' if sheet_significant == 0 else 'warning' if sheet_significant < 25 else 'error'

        html_content += f'''
                        <tr>
                            <td>{sheet_name}</td>
                            <td>{sheet_total_records}</td>
                            <td class="{avg_class}">{sheet_avg_abs:.2f} dB</td>
                            <td class="{max_class}">{sheet_max_diff:.2f} dB</td>
                            <td class="{sig_class}">{sheet_significant}</td>
                            <td>
                                <a href="{sheet_name}_detailed.xlsx" class="sheet-link">Detailed Data</a><br>
                                <a href="{sheet_name}_summary.xlsx" class="sheet-link">Statistical Summary</a><br>
                                <a href="{sheet_name}" class="sheet-link">Charts</a><br>
                                <a href="{sheet_name}_significant_differences.xlsx" class="sheet-link">Significant Differences</a>
                            </td>
                        </tr>
        '''

    html_content += '''
                    </tbody>
                </table>
            </div>

            <div class="section">
                <h2>3. 主要发现</h2>
                <ul>
                    <li><strong>整体趋势:</strong> WiFi格式之间的EVM差异主要集中在低速率(mcs0)和高发射功率情况下</li>
                    <li><strong>显著差异:</strong> 存在{total_significant_diff}个显著差异记录(>2dB)</li>
                    <li><strong>需要关注:</strong> 重点检查平均绝对差值大于2dB的配置，这些差异可能表明产品性能问题</li>
                </ul>
            </div>

            <div class="section">
                <h2>4. 使用建议</h2>
                <ol>
                    <li>检查{total_significant_diff}个显著差异记录(>2dB)，重点关注mcs0在高TX Power的情况</li>
                    <li>分析特定格式组合(如hesu vs ht, ht vs vht, vht vs hesu)的性能差异</li>
                    <li>对差异较大的测试条件进行硬件验证，以确认是否为实际硬件问题</li>
                    <li>检查测试配置是否一致，包括cable_loss、tx_shr、bb_scale等参数是否匹配</li>
                </ol>
            </div>

            <div class="section">
                <h2>5. 文件描述</h2>
                <ul>
                    <li><strong>分析脚本:</strong> compare_evm_by_wifi_format.py</li>
                    <li><strong>数据源:</strong> D:/chip_test/dev/xian_test/Xian-Esp-Test-Scripts/py_script_fpga_tx_wifi7/Log/wifi_tx_rls4/vht_ht_hesu_compare/merged_tx_result.xlsx</li>
                    <li><strong>分析时间:</strong> {}
                </ul>
            </div>
        </div>
    </body>
    </html>
    '''.format(
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        total_significant_diff=total_significant_diff
    )

    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html_content)


if __name__ == "__main__":
    main()