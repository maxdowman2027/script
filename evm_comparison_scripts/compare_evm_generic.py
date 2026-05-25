import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import os
import numpy as np
from datetime import datetime


def main():
    # 可配置变量 - 直接在这里修改即可使用
    # 文件路径
    file1 = r"D:\chip_test\dev\xian_test\Xian-Esp-Test-Scripts\py_script_fpga_tx_wifi7\Log\wifi_tx_rls4\regression_v3_260424\merged_tx_result.xlsx"
    file2 = r"D:\chip_test\dev\xian_test\Xian-Esp-Test-Scripts\py_script_fpga_tx_wifi7\Log\wifi_tx_rls4\regression_260514_mld\2\merged_tx_result.xlsx"

    # 版本名称
    version1 = "rls4_0424"
    version2 = "rls4_0521"

    # 输出目录
    output_dir = r"D:\chip_test\dev\xian_test\Xian-Esp-Test-Scripts\py_script_fpga_tx_wifi7\Log\rls4_9p_evm_comparison"

    # 创建输出目录
    os.makedirs(output_dir, exist_ok=True)

    # 读取第一个文件
    print(f"=== Start analyzing {version1} data ===")
    try:
        xls1 = pd.ExcelFile(file1)
        print(f"{version1} file contains {len(xls1.sheet_names)} Sheets: {xls1.sheet_names}")
    except Exception as e:
        print(f"Failed to read {version1} file: {e}")
        return

    # 读取第二个文件
    print(f"\n=== Start analyzing {version2} data ===")
    try:
        xls2 = pd.ExcelFile(file2)
        print(f"{version2} file contains {len(xls2.sheet_names)} Sheets: {xls2.sheet_names}")
    except Exception as e:
        print(f"Failed to read {version2} file: {e}")
        return

    # 读取所有Sheet数据
    data1 = {}
    for sheet in xls1.sheet_names:
        try:
            df = pd.read_excel(xls1, sheet_name=sheet)
            data1[sheet] = df
            print(f"\n{version1} {sheet} Sheet info:")
            print(f"Record count: {len(df)}")
            print(f"Column names: {list(df.columns)}")
            print(f"Format types: {df['wifi_format'].unique() if 'wifi_format' in df.columns else 'N/A'}")
            print(f"Rate types: {df['rate'].unique() if 'rate' in df.columns else 'N/A'}")
            print(f"TX Power range: {df['tx_power_set(dBm)'].min(), df['tx_power_set(dBm)'].max() if 'tx_power_set(dBm)' in df.columns else 'N/A'}")
            print(f"EVM columns: {'evm' if 'evm' in df.columns else 'evm_nss0/evm_nss1'}")
        except Exception as e:
            print(f"Failed to read {version1} {sheet} Sheet: {e}")

    data2 = {}
    for sheet in xls2.sheet_names:
        try:
            df = pd.read_excel(xls2, sheet_name=sheet)
            data2[sheet] = df
            print(f"\n{version2} {sheet} Sheet info:")
            print(f"Record count: {len(df)}")
            print(f"Column names: {list(df.columns)}")
            print(f"Format types: {df['wifi_format'].unique() if 'wifi_format' in df.columns else 'N/A'}")
            print(f"Rate types: {df['rate'].unique() if 'rate' in df.columns else 'N/A'}")
            print(f"TX Power range: {df['tx_power_set(dBm)'].min(), df['tx_power_set(dBm)'].max() if 'tx_power_set(dBm)' in df.columns else 'N/A'}")
            print(f"EVM column: {'evm' if 'evm' in df.columns else 'N/A'}")
        except Exception as e:
            print(f"Failed to read {version2} {sheet} Sheet: {e}")

    # Comparison analysis
    comparison_result = []
    for sheet1, df1 in data1.items():
        # Find matching sheet in the second file
        matched_sheet = find_matching_sheet(sheet1, data2.keys())
        if matched_sheet and matched_sheet in data2:
            print(f"\n=== Comparing {sheet1} ({version1}) and {matched_sheet} ({version2}) ===")
            compare_dataframes(df1, data2[matched_sheet], sheet1, matched_sheet, comparison_result, output_dir, version1, version2)
        else:
            print(f"\nWarning: No matching {version2} Sheet found for {version1} {sheet1}")

    # 保存对比结果
    save_comparison_results(comparison_result, output_dir, version1, version2, file1, file2)

    print(f"\n分析完成！所有结果已保存在: {output_dir}")


def find_matching_sheet(old_sheet, new_sheets):
    # 尝试找到匹配的Sheet，考虑可能的命名差异
    old_sheet_lower = old_sheet.lower()
    for new_sheet in new_sheets:
        new_sheet_lower = new_sheet.lower()
        if old_sheet_lower == new_sheet_lower:
            return new_sheet
        if '2g' in old_sheet_lower and '2g' in new_sheet_lower:
            return new_sheet
        if '5g' in old_sheet_lower and '5g' in new_sheet_lower:
            return new_sheet
        if 'vht' in old_sheet_lower and 'vht' in new_sheet_lower:
            return new_sheet
        if 'ht' in old_sheet_lower and 'ht' in new_sheet_lower:
            return new_sheet
    return None


def count_version1_rows_with_key_in_df2(df1, df2, merge_cols):
    """
    Count rows in df1 whose merge key appears at least once in df2 (same merge_cols).
    Each df1 row is counted at most once; avoids inflate from Cartesian product of duplicate keys.
    """
    if df1.empty or not merge_cols:
        return 0
    missing = [c for c in merge_cols if c not in df1.columns or c not in df2.columns]
    if missing:
        return 0
    keys_v2 = df2[merge_cols].drop_duplicates()
    chk = df1[merge_cols].merge(keys_v2, on=merge_cols, how="left", indicator=True)
    return int((chk["_merge"] == "both").sum())


def compare_dataframes(df1, df2, sheet1, sheet2, comparison_result, output_dir, version1, version2):
    # 核心关键列
    core_cols = ['wifi_format', 'rate', 'tx_power_set(dBm)']

    # 可能影响EVM的参数列
    additional_cols = ['giltf', 'heltf', 'short_gi', 'cbw', 'ht_dup', 'suer_dcm', 'afactor', 'pe']

    # 检查核心关键列是否存在
    core_missing_cols = []
    for col in core_cols:
        if col not in df1.columns or col not in df2.columns:
            core_missing_cols.append(col)

    if core_missing_cols:
        print(f"Warning: Missing key columns {', '.join(core_missing_cols)}, cannot complete comparison")
        return

    # 确定两个DataFrame共有的参数列
    shared_additional_cols = []
    for col in additional_cols:
        if col in df1.columns and col in df2.columns:
            shared_additional_cols.append(col)

    print(f"Found {len(shared_additional_cols)} shared additional parameter columns: {', '.join(shared_additional_cols)}")

    # 确定使用哪个EVM列
    evm_col1 = 'evm'
    if 'evm' not in df1.columns:
        if 'evm_nss0' in df1.columns:
            evm_col1 = 'evm_nss0'
        elif 'evm_nss1' in df1.columns:
            evm_col1 = 'evm_nss1'
        else:
            print(f"Warning: No EVM column found in {version1} data")
            return

    if 'evm' not in df2.columns:
        print(f"Warning: No EVM column found in {version2} data")
        return

    # 使用所有共有列进行合并，确保参数一致
    merge_cols = core_cols + shared_additional_cols
    # 添加psdu_crc列（如果存在）
    df1_cols = merge_cols + [evm_col1]
    df2_cols = merge_cols + ['evm']
    if 'psdu_crc' in df1.columns:
        df1_cols.append('psdu_crc')
    if 'psdu_crc' in df2.columns:
        df2_cols.append('psdu_crc')

    merged_df = pd.merge(
        df1[df1_cols],
        df2[df2_cols],
        on=merge_cols,
        how='inner',
        suffixes=(f'_{version1}', f'_{version2}')
    )

    v1_rows_with_match = count_version1_rows_with_key_in_df2(df1, df2, merge_cols)
    inner_join_rows = len(merged_df)
    v1_match_rate_pct = (v1_rows_with_match / len(df1)) * 100 if len(df1) else 0.0

    print(
        f"Inner join rows: {inner_join_rows}; "
        f"version1 rows with key in version2: {v1_rows_with_match}/{len(df1)} "
        f"({v1_match_rate_pct:.2f}%)"
    )

    # Excel / log 中 EVM 可能为字符串（如 '--'）；统一为浮点后再做差，避免 str - float
    col_v1 = f'{evm_col1}_{version1}'
    col_v2 = f'evm_{version2}'
    merged_df[col_v1] = pd.to_numeric(merged_df[col_v1], errors='coerce')
    merged_df[col_v2] = pd.to_numeric(merged_df[col_v2], errors='coerce')

    # 计算EVM差异
    merged_df['evm_diff'] = merged_df[col_v2] - merged_df[col_v1]
    merged_df['abs_diff'] = abs(merged_df['evm_diff'])

    # 保存详细的对比结果，并为EVM值添加填充色
    output_file = os.path.join(output_dir, f'{sheet1}_vs_{sheet2}_detailed.xlsx')
    merged_df.to_excel(output_file, index=False)

    # 为EVM值和差值添加填充色
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # 查找EVM相关列的索引
    evm_col1_idx = None
    evm_col2_idx = None
    evm_diff_col_idx = None
    for idx, cell in enumerate(ws[1]):
        if cell.value == f'{evm_col1}_{version1}':
            evm_col1_idx = idx + 1
        elif cell.value == f'evm_{version2}':
            evm_col2_idx = idx + 1
        elif cell.value == 'evm_diff':
            evm_diff_col_idx = idx + 1

    # 定义EVM颜色填充规则（EVM值越小越好，负数表示dB）
    def get_evm_fill(evm_value):
        if evm_value <= -30:
            return PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 绿色
        elif evm_value <= -25:
            return PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # 浅绿色
        elif evm_value <= -20:
            return PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 黄色
        elif evm_value <= -15:
            return PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # 橙色
        else:
            return PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # 红色

    # 定义差值颜色填充规则
    def get_diff_fill(diff_value):
        if diff_value <= -2:
            return PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 绿色（第二个版本明显更好）
        elif diff_value <= -1:
            return PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # 浅绿色（第二个版本更好）
        elif diff_value <= 1:
            return PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 黄色（无显著差异）
        elif diff_value <= 2:
            return PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # 橙色（第二个版本较差）
        else:
            return PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # 红色（第二个版本明显较差）

    # 为两个版本的EVM值添加填充色
    if evm_col1_idx:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=evm_col1_idx)
            evm_value = cell.value
            if isinstance(evm_value, (int, float)):
                cell.fill = get_evm_fill(evm_value)

    if evm_col2_idx:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=evm_col2_idx)
            evm_value = cell.value
            if isinstance(evm_value, (int, float)):
                cell.fill = get_evm_fill(evm_value)

    # 为差值添加填充色
    if evm_diff_col_idx:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=evm_diff_col_idx)
            diff_value = cell.value
            if isinstance(diff_value, (int, float)):
                cell.fill = get_diff_fill(diff_value)

    # 为psdu_crc为Fail的单元格添加红色填充色
    psdu_crc1_idx = None
    psdu_crc2_idx = None
    for idx, cell in enumerate(ws[1]):
        if f'psdu_crc_{version1}' in str(cell.value):
            psdu_crc1_idx = idx + 1
        if f'psdu_crc_{version2}' in str(cell.value):
            psdu_crc2_idx = idx + 1

    # 定义psdu_crc Fail的填充色（红色）
    fail_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    if psdu_crc1_idx:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=psdu_crc1_idx)
            cell_value = str(cell.value).strip().lower() if cell.value else ''
            if cell_value == 'fail':
                cell.fill = fail_fill

    if psdu_crc2_idx:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=psdu_crc2_idx)
            cell_value = str(cell.value).strip().lower() if cell.value else ''
            if cell_value == 'fail':
                cell.fill = fail_fill

    wb.save(output_file)

    # 统计结果
    stats = merged_df.groupby(['wifi_format', 'rate']).agg(
        count=('evm_diff', 'count'),
        mean_diff=('evm_diff', 'mean'),
        median_diff=('evm_diff', 'median'),
        std_diff=('evm_diff', 'std'),
        min_diff=('evm_diff', 'min'),
        max_diff=('evm_diff', 'max'),
        mean_abs_diff=('abs_diff', 'mean'),
        version1_mean_evm=(f'{evm_col1}_{version1}', 'mean'),
        version1_median_evm=(f'{evm_col1}_{version1}', 'median'),
        version1_std_evm=(f'{evm_col1}_{version1}', 'std'),
        version2_mean_evm=(f'evm_{version2}', 'mean'),
        version2_median_evm=(f'evm_{version2}', 'median'),
        version2_std_evm=(f'evm_{version2}', 'std')
    ).reset_index()

    stats.to_excel(os.path.join(output_dir, f'{sheet1}_vs_{sheet2}_summary.xlsx'), index=False)

    # 可视化
    plot_comparison(stats, merged_df, sheet1, sheet2, output_dir, evm_col1, version1, version2)

    # 更新比较结果列表
    comparison_result.append({
        f'{version1}_sheet': sheet1,
        f'{version2}_sheet': sheet2,
        # Inner join row count (can exceed len(df1) if duplicate merge keys — Cartesian expand)
        'matched_count': inner_join_rows,
        'inner_join_rows': inner_join_rows,
        'version1_rows_with_key_in_v2': v1_rows_with_match,
        'version1_match_rate_pct': v1_match_rate_pct,
        f'total_{version1}_records': len(df1),
        f'total_{version2}_records': len(df2),
        'mean_evm_diff': stats['mean_diff'].mean(),
        'max_evm_diff': stats['max_diff'].max(),
        'min_evm_diff': stats['min_diff'].min(),
        'avg_abs_diff': stats['mean_abs_diff'].mean(),
        f'{version1}_evm_col': evm_col1
    })


def plot_comparison(stats, merged_df, sheet1, sheet2, output_dir, evm_col1, version1, version2):
    sheet_dir = os.path.join(output_dir, f'{sheet1}_vs_{sheet2}')
    os.makedirs(sheet_dir, exist_ok=True)

    # 1. EVM difference distribution histogram
    plt.figure(figsize=(12, 6))
    plt.hist(merged_df['evm_diff'], bins=30, alpha=0.7, color='b')
    plt.axvline(merged_df['evm_diff'].mean(), color='r', linestyle='--', label=f'Mean = {merged_df["evm_diff"].mean():.2f}')
    plt.axvline(0, color='g', linestyle='-', label='No Difference')
    plt.title(f'{sheet1} vs {sheet2} EVM Difference Distribution')
    plt.xlabel('EVM Difference (dB)')
    plt.ylabel('Number of Records')
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(os.path.join(sheet_dir, 'evm_diff_distribution.png'), dpi=150)
    plt.close()

    # 2. Average difference by wifi_format and rate
    plt.figure(figsize=(16, 8))
    pivot = stats.pivot(index='rate', columns='wifi_format', values='mean_diff')
    sns.heatmap(pivot, annot=True, cmap='coolwarm', fmt='.2f', cbar_kws={'label': 'Average EVM Difference (dB)'})
    plt.title(f'{sheet1} vs {sheet2} Average EVM Difference by Format and Rate')
    plt.tight_layout()
    plt.savefig(os.path.join(sheet_dir, 'evm_diff_heatmap.png'), dpi=150)
    plt.close()

    # 3. Version 1 vs Version 2 EVM comparison scatter plot
    plt.figure(figsize=(12, 10))
    plt.scatter(merged_df[f'{evm_col1}_{version1}'], merged_df[f'evm_{version2}'], alpha=0.6, s=20)
    plt.plot([merged_df[f'{evm_col1}_{version1}'].min(), merged_df[f'{evm_col1}_{version1}'].max()],
             [merged_df[f'{evm_col1}_{version1}'].min(), merged_df[f'{evm_col1}_{version1}'].max()], 'r--', label='Ideal Case')
    plt.xlabel(f'{version1} EVM ({evm_col1}) (dB)')
    plt.ylabel(f'{version2} EVM (dB)')
    plt.title(f'{sheet1} vs {sheet2} {version1} vs {version2} EVM Comparison')
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(os.path.join(sheet_dir, 'evm_comparison_scatter.png'), dpi=150)
    plt.close()

    # 4. EVM difference boxplot by wifi_format
    plt.figure(figsize=(16, 8))
    sns.boxplot(x='wifi_format', y='mean_diff', data=stats)
    plt.axhline(y=0, color='g', linestyle='-', label='No Difference')
    plt.title(f'{sheet1} vs {sheet2} EVM Difference by Wifi Format')
    plt.xlabel('Wifi Format')
    plt.ylabel('EVM Difference (dB)')
    plt.xticks(rotation=45)
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(sheet_dir, 'evm_diff_by_format.png'), dpi=150)
    plt.close()

    # 5. EVM difference boxplot by rate
    plt.figure(figsize=(16, 8))
    sns.boxplot(x='rate', y='mean_diff', data=stats)
    plt.axhline(y=0, color='g', linestyle='-', label='No Difference')
    plt.title(f'{sheet1} vs {sheet2} EVM Difference by Rate')
    plt.xlabel('Rate')
    plt.ylabel('EVM Difference (dB)')
    plt.xticks(rotation=45)
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(sheet_dir, 'evm_diff_by_rate.png'), dpi=150)
    plt.close()


def save_comparison_results(comparison_result, output_dir, version1, version2, file1, file2):
    # 保存HTML报告
    html_file = os.path.join(output_dir, 'evm_comparison_report.html')
    generate_html_report(comparison_result, output_dir, html_file, version1, version2, file1, file2)

    print(f"HTML报告已保存: {html_file}")


def generate_html_report(comparison_result, output_dir, output_file, version1, version2, file1, file2):
    # Generate HTML report
    # 使用字符串替换而不是格式化字符串，避免CSS中的%符号导致的问题
    html_content = '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>VER1 vs VER2 Version EVM Comparison Analysis</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                margin: 20px;
                background-color: #f5f5f5;
            }
            .container {
                max-width: 1200px;
                margin: 0 auto;
                background-color: white;
                padding: 20px;
                border-radius: 8px;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            }
            h1, h2, h3 {
                color: #333;
            }
            .section {
                margin-bottom: 30px;
                padding-bottom: 20px;
                border-bottom: 1px solid #ddd;
            }
            .summary-table {
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 20px;
                background-color: white;
            }
            .summary-table th, .summary-table td {
                padding: 12px;
                text-align: left;
                border-bottom: 1px solid #ddd;
            }
            .summary-table th {
                background-color: #f2f2f2;
            }
            .summary-table tr:hover {
                background-color: #f5f5f5;
            }
            .success {
                color: green;
            }
            .warning {
                color: orange;
            }
            .error {
                color: red;
            }
            .image-container {
                text-align: center;
                margin: 20px 0;
            }
            .image-container img {
                max-width: 100%;
                height: auto;
                border: 1px solid #ddd;
                border-radius: 4px;
            }
            .sheet-link {
                color: #1a73e8;
                text-decoration: none;
            }
            .sheet-link:hover {
                text-decoration: underline;
            }
            .stats {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
                gap: 20px;
                margin-bottom: 30px;
            }
            .stat-card {
                background-color: #f8f9fa;
                padding: 15px;
                border-radius: 6px;
                box-shadow: 0 1px 3px rgba(0,0,0,0.1);
            }
            .stat-card h4 {
                margin-top: 0;
                color: #555;
            }
            .stat-value {
                font-size: 24px;
                font-weight: bold;
                margin: 10px 0;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>VER1 与 VER2 版本 EVM 对比分析</h1>
    '''

    # 替换版本名称
    html_content = html_content.replace('VER1', version1).replace('VER2', version2)

    # 计算整体统计信息（匹配率：版本1 中有连接键落在版本2 的行占比，不会超过 100%）
    total_inner_join_rows = 0
    total_v1_rows_with_key_in_v2 = 0
    total_version1_records = 0
    total_version2_records = 0
    max_diff = None
    min_diff = None
    avg_abs_diff = None

    for result in comparison_result:
        total_inner_join_rows += result.get('inner_join_rows', result.get('matched_count', 0))
        total_v1_rows_with_key_in_v2 += result.get('version1_rows_with_key_in_v2', 0)
        total_version1_records += result[f'total_{version1}_records']
        total_version2_records += result[f'total_{version2}_records']

        if max_diff is None or result['max_evm_diff'] > max_diff:
            max_diff = result['max_evm_diff']

        if min_diff is None or result['min_evm_diff'] < min_diff:
            min_diff = result['min_evm_diff']

        if avg_abs_diff is None:
            avg_abs_diff = result['avg_abs_diff']
        else:
            avg_abs_diff = (avg_abs_diff + result['avg_abs_diff']) / 2

    overall_match_rate_pct = (
        (total_v1_rows_with_key_in_v2 / total_version1_records) * 100
        if total_version1_records
        else 0.0
    )

    if avg_abs_diff is None:
        avg_abs_diff = 0.0
    if max_diff is None:
        max_diff = 0.0
    if min_diff is None:
        min_diff = 0.0

    html_content += f'''
            <div class="section">
                <h2>1. 整体对比统计</h2>
                <div class="stats">
                    <div class="stat-card">
                        <h4>版本1 行匹配率</h4>
                        <div class="stat-value">{overall_match_rate_pct:.2f}%</div>
                        <p>版本1 中连接键在版本2 至少出现一次的行数 / 版本1 总行数<br>
                        （{total_v1_rows_with_key_in_v2} / {total_version1_records}）</p>
                    </div>
                    <div class="stat-card">
                        <h4>Inner 配对行数</h4>
                        <div class="stat-value">{total_inner_join_rows}</div>
                        <p>merge 结果行数；若同一连接键在两侧重复，该值可大于版本1 行数（笛卡尔积）</p>
                    </div>
                    <div class="stat-card">
                        <h4>版本记录数</h4>
                        <div class="stat-value" style="font-size:18px">{version1}: {total_version1_records}<br>{version2}: {total_version2_records}</div>
                        <p>各 Sheet 合并对比时的原始行数之和</p>
                    </div>
                    <div class="stat-card">
                        <h4>平均EVM差值</h4>
                        <div class="stat-value" style="color: {'green' if abs(avg_abs_diff) < 1 else 'orange' if abs(avg_abs_diff) < 2 else 'red'}">
                            {avg_abs_diff:.2f} dB
                        </div>
                        <p>各 Sheet 平均绝对差值的算术平均</p>
                    </div>
                    <div class="stat-card">
                        <h4>最大EVM差值</h4>
                        <div class="stat-value" style="color: {'red' if max_diff > 2 else 'orange' if max_diff > 1 else 'green'}">
                            {max_diff:.2f} dB
                        </div>
                        <p>最大值</p>
                    </div>
                    <div class="stat-card">
                        <h4>最小EVM差值</h4>
                        <div class="stat-value" style="color: {'green' if min_diff > -1 else 'orange' if min_diff > -2 else 'red'}">
                            {min_diff:.2f} dB
                        </div>
                        <p>最小值</p>
                    </div>
                </div>
            </div>
    '''

    html_content += f'''
            <div class="section">
                <h2>2. Sheet级对比详情</h2>
                <table class="summary-table">
                    <thead>
                        <tr>
                            <th>{version1} Sheet</th>
                            <th>{version2} Sheet</th>
                            <th>Inner配对行数</th>
                            <th>V1可匹配行数</th>
                            <th>V1行匹配率</th>
                            <th>{version1}记录数</th>
                            <th>{version2}记录数</th>
                            <th>平均差值</th>
                            <th>最大差值</th>
                            <th>最小差值</th>
                            <th>平均绝对差值</th>
                            <th>{version1} EVM列</th>
                            <th>详细结果</th>
                        </tr>
                    </thead>
                    <tbody>
    '''

    for result in comparison_result:
        avg_diff_class = 'success' if abs(result['mean_evm_diff']) < 1 else 'warning' if abs(result['mean_evm_diff']) < 2 else 'error'
        max_diff_class = 'success' if result['max_evm_diff'] <= 2 else 'warning' if result['max_evm_diff'] <= 5 else 'error'
        min_diff_class = 'success' if result['min_evm_diff'] >= -2 else 'warning' if result['min_evm_diff'] >= -5 else 'error'
        avg_abs_class = 'success' if result['avg_abs_diff'] < 1 else 'warning' if result['avg_abs_diff'] < 2 else 'error'

        ij = result.get('inner_join_rows', result['matched_count'])
        v1m = result.get('version1_rows_with_key_in_v2', 0)
        v1pct = result.get('version1_match_rate_pct', 0.0)

        html_content += f'''
                        <tr>
                            <td>{result[f'{version1}_sheet']}</td>
                            <td>{result[f'{version2}_sheet']}</td>
                            <td>{ij}</td>
                            <td>{v1m}</td>
                            <td>{v1pct:.2f}%</td>
                            <td>{result[f'total_{version1}_records']}</td>
                            <td>{result[f'total_{version2}_records']}</td>
                            <td class="{avg_diff_class}">{result['mean_evm_diff']:.2f}</td>
                            <td class="{max_diff_class}">{result['max_evm_diff']:.2f}</td>
                            <td class="{min_diff_class}">{result['min_evm_diff']:.2f}</td>
                            <td class="{avg_abs_class}">{result['avg_abs_diff']:.2f}</td>
                            <td>{result[f'{version1}_evm_col']}</td>
                            <td>
                                <a href="{result[f'{version1}_sheet']}_vs_{result[f'{version2}_sheet']}_detailed.xlsx" class="sheet-link">Detailed Data</a><br>
                                <a href="{result[f'{version1}_sheet']}_vs_{result[f'{version2}_sheet']}_summary.xlsx" class="sheet-link">Statistical Summary</a><br>
                                <a href="{result[f'{version1}_sheet']}_vs_{result[f'{version2}_sheet']}" class="sheet-link">Charts</a>
                            </td>
                        </tr>
        '''

    html_content += '''
                    </tbody>
                </table>
            </div>

            <div class="section">
                <h2>3. 主要发现</h2>
                <ul>
                    <li><strong>版本1 行匹配率:</strong> VER1 记录中 MATCHED_PERCENT% 的连接键在 VER2 中至少存在一行（同一 Sheet 对比内按 merge 键统计；不会超过 100%）。Inner merge 的详细配对行数见「Inner 配对行数」及各 Sheet「Inner配对行数」列（重复键时可为笛卡尔积）。</li>
                    <li><strong>整体EVM趋势:</strong> TREND</li>
                    <li><strong>主要差异来源:</strong> 需要进一步分析特定Rate和Format组合的性能</li>
                </ul>
            </div>

            <div class="section">
                <h2>4. 使用建议</h2>
                <ol>
                    <li>应优先检查具有较大差异的Sheet和Rate/Format组合</li>
                    <li>重点分析平均绝对差值大于2dB的配置</li>
                    <li>检查VER2版本是否覆盖了所有测试场景</li>
                    <li>如有必要，进行硬件验证以确认差异是否合理</li>
                </ol>
            </div>

            <div class="section">
                <h2>5. 文件描述</h2>
                <ul>
                    <li><strong>分析脚本:</strong> compare_evm_generic.py</li>
                    <li><strong>VER1版本文件:</strong> FILE1</li>
                    <li><strong>VER2版本文件:</strong> FILE2</li>
                    <li><strong>生成时间:</strong> TIME</li>
                </ul>
            </div>
        </div>
    </body>
    </html>
    '''

    # 替换剩余的占位符
    matched_percent = overall_match_rate_pct
    trend = f"{version2}版本整体EVM更好" if avg_abs_diff < -0.5 else f"{version2}版本整体EVM更差" if avg_abs_diff > 0.5 else f"{version1}和{version2}版本之间的EVM差异不显著"
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    html_content = html_content.replace('MATCHED_PERCENT', f"{matched_percent:.2f}").replace('TREND', trend)
    html_content = html_content.replace('FILE1', file1).replace('FILE2', file2).replace('TIME', current_time)

    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html_content)


if __name__ == "__main__":
    main()