import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import os
import numpy as np
from datetime import datetime


def main():
    # 文件路径 - 更新为包含hesu格式的新文件
    rls4_file = r"D:\chip_test\dev\xian_test\Xian-Esp-Test-Scripts\py_script_fpga_tx_wifi7\Log\wifi_tx_rls4\vht_ht_hesu_compare\merged_tx_result.xlsx"
    wifi7_file = r"D:\chip_test\dev\xian_test\Xian-Esp-Test-Scripts\py_script_fpga_tx_wifi7\Log\wifi_tx\20260407\vht_ht_hesu\merged_tx_result.xlsx"

    # 输出目录
    output_dir = r"D:\chip_test\dev\xian_test\Xian-Esp-Test-Scripts\py_script_fpga_tx_wifi7\Log\wifi_tx\20260407\vht_ht_hesu\rls4_wifi7_hesu_evm_comparison"
    os.makedirs(output_dir, exist_ok=True)

    print("=== Start analyzing RLS4.0 version data (hesu) ===")
    try:
        rls4_xls = pd.ExcelFile(rls4_file)
        print(f"RLS4.0 version file contains {len(rls4_xls.sheet_names)} Sheets: {rls4_xls.sheet_names}")
    except Exception as e:
        print(f"Failed to read RLS4.0 version file: {e}")
        return

    print("\n=== Start analyzing WiFi7 version data (hesu) ===")
    try:
        wifi7_xls = pd.ExcelFile(wifi7_file)
        print(f"WiFi7 version file contains {len(wifi7_xls.sheet_names)} Sheets: {wifi7_xls.sheet_names}")
    except Exception as e:
        print(f"Failed to read WiFi7 version file: {e}")
        return

    # 读取所有Sheet数据
    rls4_data = {}
    for sheet in rls4_xls.sheet_names:
        try:
            df = pd.read_excel(rls4_xls, sheet_name=sheet)
            rls4_data[sheet] = df
            print(f"\nRLS4.0 {sheet} Sheet info:")
            print(f"Record count: {len(df)}")
            print(f"Column names: {list(df.columns)}")
            print(f"Format types: {df['wifi_format'].unique() if 'wifi_format' in df.columns else 'N/A'}")
            print(f"Rate types: {df['rate'].unique() if 'rate' in df.columns else 'N/A'}")
            print(f"TX Power range: {df['tx_power_set(dBm)'].min(), df['tx_power_set(dBm)'].max() if 'tx_power_set(dBm)' in df.columns else 'N/A'}")
            print(f"EVM columns: {'evm' if 'evm' in df.columns else 'evm_nss0/evm_nss1'}")
        except Exception as e:
            print(f"Failed to read RLS4.0 {sheet} Sheet: {e}")

    wifi7_data = {}
    for sheet in wifi7_xls.sheet_names:
        try:
            df = pd.read_excel(wifi7_xls, sheet_name=sheet)
            wifi7_data[sheet] = df
            print(f"\nWiFi7 {sheet} Sheet info:")
            print(f"Record count: {len(df)}")
            print(f"Column names: {list(df.columns)}")
            print(f"Format types: {df['wifi_format'].unique() if 'wifi_format' in df.columns else 'N/A'}")
            print(f"Rate types: {df['rate'].unique() if 'rate' in df.columns else 'N/A'}")
            print(f"TX Power range: {df['tx_power_set(dBm)'].min(), df['tx_power_set(dBm)'].max() if 'tx_power_set(dBm)' in df.columns else 'N/A'}")
            print(f"EVM column: {'evm' if 'evm' in df.columns else 'N/A'}")
        except Exception as e:
            print(f"Failed to read WiFi7 {sheet} Sheet: {e}")

    # Comparison analysis
    comparison_result = []
    for rls4_sheet, rls4_df in rls4_data.items():
        # Find matching WiFi7 version Sheet
        matched_sheet = find_matching_sheet(rls4_sheet, wifi7_data.keys())
        if matched_sheet and matched_sheet in wifi7_data:
            print(f"\n=== Comparing {rls4_sheet} (RLS4.0) and {matched_sheet} (WiFi7) ===")
            compare_dataframes(rls4_df, wifi7_data[matched_sheet], rls4_sheet, matched_sheet, comparison_result, output_dir)
        else:
            print(f"\nWarning: No matching WiFi7 version Sheet found for RLS4.0 {rls4_sheet}")

    # 保存对比结果
    save_comparison_results(comparison_result, output_dir)

    print(f"\n分析完成！所有结果已保存在: {output_dir}")


def find_matching_sheet(old_sheet, new_sheets):
    # 尝试找到匹配的Sheet，考虑可能的命名差异
    old_sheet_lower = old_sheet.lower()
    for new_sheet in new_sheets:
        new_sheet_lower = new_sheet.lower()
        if old_sheet_lower == new_sheet_lower:
            return new_sheet
        if '2g' in old_sheet_lower and '2g' in new_sheet_lower:
            return new_sheet
        if '5g' in old_sheet_lower and '5g' in new_sheet_lower:
            return new_sheet
        if 'vht' in old_sheet_lower and 'vht' in new_sheet_lower:
            return new_sheet
        if 'ht' in old_sheet_lower and 'ht' in new_sheet_lower:
            return new_sheet
    return None


def compare_dataframes(rls4_df, wifi7_df, rls4_sheet, wifi7_sheet, comparison_result, output_dir):
    # Ensure key columns exist
    required_cols = ['wifi_format', 'rate', 'tx_power_set(dBm)']
    missing_cols = []
    for col in required_cols:
        if col not in rls4_df.columns or col not in wifi7_df.columns:
            missing_cols.append(col)

    if missing_cols:
        print(f"Warning: Missing key columns {', '.join(missing_cols)}, cannot complete comparison")
        return

    # 确定使用哪个EVM列
    rls4_evm_col = 'evm'
    if 'evm' not in rls4_df.columns:
        if 'evm_nss0' in rls4_df.columns:
            rls4_evm_col = 'evm_nss0'
        elif 'evm_nss1' in rls4_df.columns:
            rls4_evm_col = 'evm_nss1'
        else:
            print("Warning: No EVM column found in RLS4.0 data")
            return

    if 'evm' not in wifi7_df.columns:
        print("Warning: No EVM column found in WiFi7 data")
        return

    # Merge data
    merged_df = pd.merge(
        rls4_df[required_cols + [rls4_evm_col]],
        wifi7_df[required_cols + ['evm']],
        on=['wifi_format', 'rate', 'tx_power_set(dBm)'],
        how='inner',
        suffixes=('_rls4', '_wifi7')
    )

    print(f"Found {len(merged_df)} matching records")

    # 计算EVM差异
    merged_df['evm_diff'] = merged_df['evm_wifi7'] - merged_df[f'{rls4_evm_col}_rls4']
    merged_df['abs_diff'] = abs(merged_df['evm_diff'])

    # 保存详细的对比结果，并为EVM值添加填充色
    output_file = os.path.join(output_dir, f'{rls4_sheet}_vs_{wifi7_sheet}_detailed.xlsx')
    merged_df.to_excel(output_file, index=False)

    # 为EVM值和差值添加填充色
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # 查找EVM相关列的索引
    rls4_evm_col_idx = None
    wifi7_evm_col_idx = None
    evm_diff_col_idx = None
    for idx, cell in enumerate(ws[1]):
        if cell.value == f'{rls4_evm_col}_rls4':
            rls4_evm_col_idx = idx + 1
        elif cell.value == 'evm_wifi7':
            wifi7_evm_col_idx = idx + 1
        elif cell.value == 'evm_diff':
            evm_diff_col_idx = idx + 1

    # 定义EVM颜色填充规则（EVM值越小越好，负数表示dB）
    def get_evm_fill(evm_value):
        if evm_value <= -30:
            return PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 绿色
        elif evm_value <= -25:
            return PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # 浅绿色
        elif evm_value <= -20:
            return PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 黄色
        elif evm_value <= -15:
            return PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # 橙色
        else:
            return PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # 红色

    # 定义差值颜色填充规则
    def get_diff_fill(diff_value):
        if diff_value <= -2:
            return PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 绿色（WiFi7 EVM明显更好）
        elif diff_value <= -1:
            return PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # 浅绿色（WiFi7 EVM更好）
        elif diff_value <= 1:
            return PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 黄色（无显著差异）
        elif diff_value <= 2:
            return PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # 橙色（WiFi7 EVM较差）
        else:
            return PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # 红色（WiFi7 EVM明显较差）

    # 为RLS4.0和WiFi7的EVM值添加填充色
    if rls4_evm_col_idx:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=rls4_evm_col_idx)
            evm_value = cell.value
            if isinstance(evm_value, (int, float)):
                cell.fill = get_evm_fill(evm_value)

    if wifi7_evm_col_idx:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=wifi7_evm_col_idx)
            evm_value = cell.value
            if isinstance(evm_value, (int, float)):
                cell.fill = get_evm_fill(evm_value)

    # 为差值添加填充色
    if evm_diff_col_idx:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=evm_diff_col_idx)
            diff_value = cell.value
            if isinstance(diff_value, (int, float)):
                cell.fill = get_diff_fill(diff_value)

    wb.save(output_file)

    # 统计结果
    stats = merged_df.groupby(['wifi_format', 'rate']).agg(
        count=('evm_diff', 'count'),
        mean_diff=('evm_diff', 'mean'),
        median_diff=('evm_diff', 'median'),
        std_diff=('evm_diff', 'std'),
        min_diff=('evm_diff', 'min'),
        max_diff=('evm_diff', 'max'),
        mean_abs_diff=('abs_diff', 'mean'),
        rls4_mean_evm=(f'{rls4_evm_col}_rls4', 'mean'),
        rls4_median_evm=(f'{rls4_evm_col}_rls4', 'median'),
        rls4_std_evm=(f'{rls4_evm_col}_rls4', 'std'),
        wifi7_mean_evm=('evm_wifi7', 'mean'),
        wifi7_median_evm=('evm_wifi7', 'median'),
        wifi7_std_evm=('evm_wifi7', 'std')
    ).reset_index()

    stats.to_excel(os.path.join(output_dir, f'{rls4_sheet}_vs_{wifi7_sheet}_summary.xlsx'), index=False)

    # 可视化
    plot_comparison(stats, merged_df, rls4_sheet, wifi7_sheet, output_dir, rls4_evm_col)

    # 更新比较结果列表
    comparison_result.append({
        'rls4_sheet': rls4_sheet,
        'wifi7_sheet': wifi7_sheet,
        'matched_count': len(merged_df),
        'total_rls4_records': len(rls4_df),
        'total_wifi7_records': len(wifi7_df),
        'mean_evm_diff': stats['mean_diff'].mean(),
        'max_evm_diff': stats['max_diff'].max(),
        'min_evm_diff': stats['min_diff'].min(),
        'avg_abs_diff': stats['mean_abs_diff'].mean(),
        'rls4_evm_col': rls4_evm_col
    })


def plot_comparison(stats, merged_df, rls4_sheet, wifi7_sheet, output_dir, rls4_evm_col):
    sheet_dir = os.path.join(output_dir, f'{rls4_sheet}_vs_{wifi7_sheet}')
    os.makedirs(sheet_dir, exist_ok=True)

    # 1. EVM difference distribution histogram
    plt.figure(figsize=(12, 6))
    plt.hist(merged_df['evm_diff'], bins=30, alpha=0.7, color='b')
    plt.axvline(merged_df['evm_diff'].mean(), color='r', linestyle='--', label=f'Mean = {merged_df["evm_diff"].mean():.2f}')
    plt.axvline(0, color='g', linestyle='-', label='No Difference')
    plt.title(f'{rls4_sheet} vs {wifi7_sheet} EVM Difference Distribution')
    plt.xlabel('EVM Difference (dB)')
    plt.ylabel('Number of Records')
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(os.path.join(sheet_dir, 'evm_diff_distribution.png'), dpi=150)
    plt.close()

    # 2. Average difference by wifi_format and rate
    plt.figure(figsize=(16, 8))
    pivot = stats.pivot(index='rate', columns='wifi_format', values='mean_diff')
    sns.heatmap(pivot, annot=True, cmap='coolwarm', fmt='.2f', cbar_kws={'label': 'Average EVM Difference (dB)'})
    plt.title(f'{rls4_sheet} vs {wifi7_sheet} Average EVM Difference by Format and Rate')
    plt.tight_layout()
    plt.savefig(os.path.join(sheet_dir, 'evm_diff_heatmap.png'), dpi=150)
    plt.close()

    # 3. RLS4.0 vs WiFi7 EVM comparison scatter plot
    plt.figure(figsize=(12, 10))
    plt.scatter(merged_df[f'{rls4_evm_col}_rls4'], merged_df['evm_wifi7'], alpha=0.6, s=20)
    plt.plot([merged_df[f'{rls4_evm_col}_rls4'].min(), merged_df[f'{rls4_evm_col}_rls4'].max()],
             [merged_df[f'{rls4_evm_col}_rls4'].min(), merged_df[f'{rls4_evm_col}_rls4'].max()], 'r--', label='Ideal Case')
    plt.xlabel(f'RLS4.0 EVM ({rls4_evm_col}) (dB)')
    plt.ylabel('WiFi7 EVM (dB)')
    plt.title(f'{rls4_sheet} vs {wifi7_sheet} RLS4.0 vs WiFi7 EVM Comparison')
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(os.path.join(sheet_dir, 'evm_comparison_scatter.png'), dpi=150)
    plt.close()

    # 4. EVM difference boxplot by wifi_format
    plt.figure(figsize=(16, 8))
    sns.boxplot(x='wifi_format', y='mean_diff', data=stats)
    plt.axhline(y=0, color='g', linestyle='-', label='No Difference')
    plt.title(f'{rls4_sheet} vs {wifi7_sheet} EVM Difference by Wifi Format')
    plt.xlabel('Wifi Format')
    plt.ylabel('EVM Difference (dB)')
    plt.xticks(rotation=45)
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(sheet_dir, 'evm_diff_by_format.png'), dpi=150)
    plt.close()

    # 5. EVM difference boxplot by rate
    plt.figure(figsize=(16, 8))
    sns.boxplot(x='rate', y='mean_diff', data=stats)
    plt.axhline(y=0, color='g', linestyle='-', label='No Difference')
    plt.title(f'{rls4_sheet} vs {wifi7_sheet} EVM Difference by Rate')
    plt.xlabel('Rate')
    plt.ylabel('EVM Difference (dB)')
    plt.xticks(rotation=45)
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(sheet_dir, 'evm_diff_by_rate.png'), dpi=150)
    plt.close()


def save_comparison_results(comparison_result, output_dir):
    # 保存HTML报告
    html_file = os.path.join(output_dir, 'evm_comparison_report.html')
    generate_html_report(comparison_result, output_dir, html_file)

    print(f"HTML报告已保存: {html_file}")


def generate_html_report(comparison_result, output_dir, output_file):
    # Generate HTML report
    html_content = '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>RLS4.0 vs WiFi7 Version EVM Comparison Analysis (hesu)</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                margin: 20px;
                background-color: #f5f5f5;
            }
            .container {
                max-width: 1200px;
                margin: 0 auto;
                background-color: white;
                padding: 20px;
                border-radius: 8px;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            }
            h1, h2, h3 {
                color: #333;
            }
            .section {
                margin-bottom: 30px;
                padding-bottom: 20px;
                border-bottom: 1px solid #ddd;
            }
            .summary-table {
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 20px;
                background-color: white;
            }
            .summary-table th, .summary-table td {
                padding: 12px;
                text-align: left;
                border-bottom: 1px solid #ddd;
            }
            .summary-table th {
                background-color: #f2f2f2;
            }
            .summary-table tr:hover {
                background-color: #f5f5f5;
            }
            .success {
                color: green;
            }
            .warning {
                color: orange;
            }
            .error {
                color: red;
            }
            .image-container {
                text-align: center;
                margin: 20px 0;
            }
            .image-container img {
                max-width: 100%;
                height: auto;
                border: 1px solid #ddd;
                border-radius: 4px;
            }
            .sheet-link {
                color: #1a73e8;
                text-decoration: none;
            }
            .sheet-link:hover {
                text-decoration: underline;
            }
            .stats {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
                gap: 20px;
                margin-bottom: 30px;
            }
            .stat-card {
                background-color: #f8f9fa;
                padding: 15px;
                border-radius: 6px;
                box-shadow: 0 1px 3px rgba(0,0,0,0.1);
            }
            .stat-card h4 {
                margin-top: 0;
                color: #555;
            }
            .stat-value {
                font-size: 24px;
                font-weight: bold;
                margin: 10px 0;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>RLS4.0 与 WiFi7 版本 EVM 对比分析 (hesu)</h1>

            <div class="section">
                <h2>1. 整体对比统计</h2>
                <div class="stats">
    '''

    total_matched = 0
    total_rls4_records = 0
    total_wifi7_records = 0
    max_diff = None
    min_diff = None
    avg_abs_diff = None

    for result in comparison_result:
        total_matched += result['matched_count']
        total_rls4_records += result['total_rls4_records']
        total_wifi7_records += result['total_wifi7_records']

        if max_diff is None or result['max_evm_diff'] > max_diff:
            max_diff = result['max_evm_diff']

        if min_diff is None or result['min_evm_diff'] < min_diff:
            min_diff = result['min_evm_diff']

        if avg_abs_diff is None:
            avg_abs_diff = result['avg_abs_diff']
        else:
            avg_abs_diff = (avg_abs_diff + result['avg_abs_diff']) / 2

    html_content += '''
                <div class="stat-card">
                    <h4>匹配记录数</h4>
                    <div class="stat-value">%d</div>
                    <p>RLS4.0总记录数: %d<br>WiFi7总记录数: %d</p>
                </div>
                <div class="stat-card">
                    <h4>平均EVM差值</h4>
                    <div class="stat-value" style="color:%s;">%.2f dB</div>
                    <p>平均值</p>
                </div>
                <div class="stat-card">
                    <h4>最大EVM差值</h4>
                    <div class="stat-value" style="color:%s;">%.2f dB</div>
                    <p>最大值</p>
                </div>
                <div class="stat-card">
                    <h4>最小EVM差值</h4>
                    <div class="stat-value" style="color:%s;">%.2f dB</div>
                    <p>最小值</p>
                </div>
    ''' % (
        total_matched, total_rls4_records, total_wifi7_records,
        'green' if abs(avg_abs_diff) < 1 else 'orange' if abs(avg_abs_diff) < 2 else 'red',
        avg_abs_diff,
        'red' if max_diff > 2 else 'orange' if max_diff > 1 else 'green',
        max_diff,
        'green' if min_diff > -1 else 'orange' if min_diff > -2 else 'red',
        min_diff
    )

    html_content += '''
                </div>
            </div>

            <div class="section">
                <h2>2. Sheet级对比详情</h2>
                <table class="summary-table">
                    <thead>
                        <tr>
                            <th>RLS4.0 Sheet</th>
                            <th>WiFi7 Sheet</th>
                            <th>匹配记录数</th>
                            <th>RLS4.0记录数</th>
                            <th>WiFi7记录数</th>
                            <th>平均差值</th>
                            <th>最大差值</th>
                            <th>最小差值</th>
                            <th>平均绝对差值</th>
                            <th>RLS4.0 EVM列</th>
                            <th>详细结果</th>
                        </tr>
                    </thead>
                    <tbody>
    '''

    for result in comparison_result:
        avg_diff_class = 'success' if abs(result['mean_evm_diff']) < 1 else 'warning' if abs(result['mean_evm_diff']) < 2 else 'error'
        max_diff_class = 'success' if result['max_evm_diff'] <= 2 else 'warning' if result['max_evm_diff'] <= 5 else 'error'
        min_diff_class = 'success' if result['min_evm_diff'] >= -2 else 'warning' if result['min_evm_diff'] >= -5 else 'error'
        avg_abs_class = 'success' if result['avg_abs_diff'] < 1 else 'warning' if result['avg_abs_diff'] < 2 else 'error'

        html_content += '''
                        <tr>
                            <td>%s</td>
                            <td>%s</td>
                            <td>%d</td>
                            <td>%d</td>
                            <td>%d</td>
                            <td class="%s">%.2f</td>
                            <td class="%s">%.2f</td>
                            <td class="%s">%.2f</td>
                            <td class="%s">%.2f</td>
                            <td>%s</td>
                            <td>
                                <a href="%s_vs_%s_detailed.xlsx" class="sheet-link">Detailed Data</a><br>
                                <a href="%s_vs_%s_summary.xlsx" class="sheet-link">Statistical Summary</a><br>
                                <a href="%s_vs_%s" class="sheet-link">Charts</a>
                            </td>
                        </tr>
        ''' % (
            result['rls4_sheet'],
            result['wifi7_sheet'],
            result['matched_count'],
            result['total_rls4_records'],
            result['total_wifi7_records'],
            avg_diff_class, result['mean_evm_diff'],
            max_diff_class, result['max_evm_diff'],
            min_diff_class, result['min_evm_diff'],
            avg_abs_class, result['avg_abs_diff'],
            result['rls4_evm_col'],
            result['rls4_sheet'], result['wifi7_sheet'],
            result['rls4_sheet'], result['wifi7_sheet'],
            result['rls4_sheet'], result['wifi7_sheet']
        )

    html_content += '''
                    </tbody>
                </table>
            </div>

            <div class="section">
                <h2>3. 主要发现</h2>
                <ul>
                    <li><strong>匹配记录百分比:</strong> RLS4.0记录中有%.2f%%在WiFi7版本中找到了匹配项</li>
                    <li><strong>整体EVM趋势:</strong> %s</li>
                    <li><strong>主要差异来源:</strong> 需要进一步分析特定Rate和Format组合的性能，特别是hesu格式</li>
                </ul>
            </div>

            <div class="section">
                <h2>4. 使用建议</h2>
                <ol>
                    <li>应优先检查具有较大差异的Sheet和Rate/Format组合，特别是hesu格式</li>
                    <li>重点分析平均绝对差值大于2dB的配置</li>
                    <li>检查WiFi7版本是否覆盖了所有测试场景</li>
                    <li>如有必要，进行硬件验证以确认差异是否合理</li>
                </ol>
            </div>

            <div class="section">
                <h2>5. 文件描述</h2>
                <ul>
                    <li><strong>分析脚本:</strong> compare_evm_rls4_wifi7_hesu.py</li>
                    <li><strong>RLS4.0版本文件:</strong> D:/chip_test/dev/xian_test/Xian-Esp-Test-Scripts/py_script_fpga_tx_wifi7/Log/wifi_tx_rls4/vht_ht_hesu_compare/merged_tx_result.xlsx</li>
                    <li><strong>WiFi7版本文件:</strong> D:/chip_test/dev/xian_test/Xian-Esp-Test-Scripts/py_script_fpga_tx_wifi7/Log/wifi_tx/20260407/vht_ht_hesu/merged_tx_result.xlsx</li>
                    <li><strong>生成时间:</strong> %s</li>
                </ul>
            </div>
        </div>
    </body>
    </html>
    ''' % (
        (total_matched / total_rls4_records) * 100,
        "WiFi7版本整体EVM更好" if avg_abs_diff < -0.5 else "WiFi7版本整体EVM更差" if avg_abs_diff > 0.5 else "RLS4.0和WiFi7版本之间的EVM差异不显著",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    )

    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html_content)


if __name__ == "__main__":
    main()
