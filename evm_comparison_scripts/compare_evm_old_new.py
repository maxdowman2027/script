import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import os
import numpy as np

def main():
    # 文件路径
    old_file = r"D:\chip_test\dev\chip_tx\eagletest\py_script_fpga_tx_wifi7\Log\wifi_tx\20260407\vht_ht_old\merged_tx_result.xlsx"
    new_file = r"D:\chip_test\dev\xian_test\Xian-Esp-Test-Scripts\py_script_fpga_tx_wifi7\Log\wifi_tx\20260407\vht_ht_2\merged_tx_result.xlsx"

    # 输出目录
    output_dir = os.path.join(os.path.dirname(new_file), 'evm_comparison_results')
    os.makedirs(output_dir, exist_ok=True)

    print("=== Start analyzing old version data ===")
    try:
        old_xls = pd.ExcelFile(old_file)
        print(f"Old version file contains {len(old_xls.sheet_names)} Sheets: {old_xls.sheet_names}")
    except Exception as e:
        print(f"Failed to read old version file: {e}")
        return

    print("\n=== Start analyzing new version data ===")
    try:
        new_xls = pd.ExcelFile(new_file)
        print(f"New version file contains {len(new_xls.sheet_names)} Sheets: {new_xls.sheet_names}")
    except Exception as e:
        print(f"Failed to read new version file: {e}")
        return

    # 读取所有Sheet数据
    old_data = {}
    for sheet in old_xls.sheet_names:
        try:
            df = pd.read_excel(old_xls, sheet_name=sheet)
            old_data[sheet] = df
            print(f"\nOld version {sheet} Sheet info:")
            print(f"Record count: {len(df)}")
            print(f"Column names: {list(df.columns)}")
            print(f"Format types: {df['wifi_format'].unique() if 'wifi_format' in df.columns else 'N/A'}")
            print(f"Rate types: {df['rate'].unique() if 'rate' in df.columns else 'N/A'}")
            print(f"RF Channel range: {df['rf_chan'].min(), df['rf_chan'].max() if 'rf_chan' in df.columns else 'N/A'}")
            print(f"TX Power range: {df['tx_power_set(dBm)'].min(), df['tx_power_set(dBm)'].max() if 'tx_power_set(dBm)' in df.columns else 'N/A'}")
            print(f"EVM range: {df['evm'].min(), df['evm'].max() if 'evm' in df.columns else 'N/A'}")
        except Exception as e:
            print(f"Failed to read old version {sheet} Sheet: {e}")

    new_data = {}
    for sheet in new_xls.sheet_names:
        try:
            df = pd.read_excel(new_xls, sheet_name=sheet)
            new_data[sheet] = df
            print(f"\nNew version {sheet} Sheet info:")
            print(f"Record count: {len(df)}")
            print(f"Column names: {list(df.columns)}")
            print(f"Format types: {df['wifi_format'].unique() if 'wifi_format' in df.columns else 'N/A'}")
            print(f"Rate types: {df['rate'].unique() if 'rate' in df.columns else 'N/A'}")
            print(f"RF Channel range: {df['rf_chan'].min(), df['rf_chan'].max() if 'rf_chan' in df.columns else 'N/A'}")
            print(f"TX Power range: {df['tx_power_set(dBm)'].min(), df['tx_power_set(dBm)'].max() if 'tx_power_set(dBm)' in df.columns else 'N/A'}")
            print(f"EVM range: {df['evm'].min(), df['evm'].max() if 'evm' in df.columns else 'N/A'}")
        except Exception as e:
            print(f"Failed to read new version {sheet} Sheet: {e}")

    # Comparison analysis
    comparison_result = []
    for old_sheet, old_df in old_data.items():
        # Find matching new version Sheet
        matched_sheet = find_matching_sheet(old_sheet, new_data.keys())
        if matched_sheet and matched_sheet in new_data:
            print(f"\n=== Comparing {old_sheet} (Old) and {matched_sheet} (New) ===")
            compare_dataframes(old_df, new_data[matched_sheet], old_sheet, matched_sheet, comparison_result, output_dir)
        else:
            print(f"\nWarning: No matching new version Sheet found for old version {old_sheet}")

    # 保存对比结果
    save_comparison_results(comparison_result, output_dir)

    print(f"\n分析完成！所有结果已保存在: {output_dir}")

def find_matching_sheet(old_sheet, new_sheets):
    # 尝试找到匹配的Sheet，考虑可能的命名差异
    old_sheet_lower = old_sheet.lower()
    for new_sheet in new_sheets:
        new_sheet_lower = new_sheet.lower()
        if old_sheet_lower == new_sheet_lower:
            return new_sheet
        if '2g' in old_sheet_lower and '2g' in new_sheet_lower:
            return new_sheet
        if '5g' in old_sheet_lower and '5g' in new_sheet_lower:
            return new_sheet
        if 'vht' in old_sheet_lower and 'vht' in new_sheet_lower:
            return new_sheet
        if 'ht' in old_sheet_lower and 'ht' in new_sheet_lower:
            return new_sheet
    return None

def compare_dataframes(old_df, new_df, old_sheet, new_sheet, comparison_result, output_dir):
    # Ensure key columns exist
    required_cols = ['wifi_format', 'rate', 'rf_chan', 'tx_power_set(dBm)', 'evm']
    missing_cols = []
    for col in required_cols:
        if col not in old_df.columns or col not in new_df.columns:
            missing_cols.append(col)

    if missing_cols:
        print(f"Warning: Missing key columns {', '.join(missing_cols)}, cannot complete comparison")
        return

    # Merge data
    merged_df = pd.merge(
        old_df[required_cols],
        new_df[required_cols],
        on=['wifi_format', 'rate', 'rf_chan', 'tx_power_set(dBm)'],
        how='inner',
        suffixes=('_old', '_new')
    )

    print(f"Found {len(merged_df)} matching records")

    # 计算EVM差异
    merged_df['evm_diff'] = merged_df['evm_new'] - merged_df['evm_old']
    merged_df['abs_diff'] = abs(merged_df['evm_diff'])

    # 保存详细的对比结果，并为EVM值添加填充色
    output_file = os.path.join(output_dir, f'{old_sheet}_vs_{new_sheet}_detailed.xlsx')
    merged_df.to_excel(output_file, index=False)

    # 为EVM值和差值添加填充色
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # 查找EVM相关列的索引
    evm_old_col = None
    evm_new_col = None
    evm_diff_col = None
    for idx, cell in enumerate(ws[1]):
        if cell.value == 'evm_old':
            evm_old_col = idx + 1
        elif cell.value == 'evm_new':
            evm_new_col = idx + 1
        elif cell.value == 'evm_diff':
            evm_diff_col = idx + 1

    # 定义EVM颜色填充规则（EVM值越小越好，负数表示dB）
    def get_evm_fill(evm_value):
        if evm_value <= -30:
            return PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 绿色
        elif evm_value <= -25:
            return PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # 浅绿色
        elif evm_value <= -20:
            return PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 黄色
        elif evm_value <= -15:
            return PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # 橙色
        else:
            return PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # 红色

    # 定义差值颜色填充规则
    def get_diff_fill(diff_value):
        if diff_value <= -2:
            return PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 绿色（新版本EVM明显更好）
        elif diff_value <= -1:
            return PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # 浅绿色（新版本EVM更好）
        elif diff_value <= 1:
            return PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 黄色（无显著差异）
        elif diff_value <= 2:
            return PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # 橙色（新版本EVM较差）
        else:
            return PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # 红色（新版本EVM明显较差）

    # 为旧版本和新版本的EVM值添加填充色
    if evm_old_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=evm_old_col)
            evm_value = cell.value
            if isinstance(evm_value, (int, float)):
                cell.fill = get_evm_fill(evm_value)

    if evm_new_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=evm_new_col)
            evm_value = cell.value
            if isinstance(evm_value, (int, float)):
                cell.fill = get_evm_fill(evm_value)

    # 为差值添加填充色
    if evm_diff_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=evm_diff_col)
            diff_value = cell.value
            if isinstance(diff_value, (int, float)):
                cell.fill = get_diff_fill(diff_value)

    wb.save(output_file)

    # 统计结果
    stats = merged_df.groupby(['wifi_format', 'rate']).agg(
        count=('evm_diff', 'count'),
        mean_diff=('evm_diff', 'mean'),
        median_diff=('evm_diff', 'median'),
        std_diff=('evm_diff', 'std'),
        min_diff=('evm_diff', 'min'),
        max_diff=('evm_diff', 'max'),
        mean_abs_diff=('abs_diff', 'mean'),
        old_mean_evm=('evm_old', 'mean'),
        old_median_evm=('evm_old', 'median'),
        old_std_evm=('evm_old', 'std'),
        new_mean_evm=('evm_new', 'mean'),
        new_median_evm=('evm_new', 'median'),
        new_std_evm=('evm_new', 'std')
    ).reset_index()

    stats.to_excel(os.path.join(output_dir, f'{old_sheet}_vs_{new_sheet}_summary.xlsx'), index=False)

    # 可视化
    plot_comparison(stats, merged_df, old_sheet, new_sheet, output_dir)

    # 更新比较结果列表
    comparison_result.append({
        'old_sheet': old_sheet,
        'new_sheet': new_sheet,
        'matched_count': len(merged_df),
        'total_old_records': len(old_df),
        'total_new_records': len(new_df),
        'mean_evm_diff': stats['mean_diff'].mean(),
        'max_evm_diff': stats['max_diff'].max(),
        'min_evm_diff': stats['min_diff'].min(),
        'avg_abs_diff': stats['mean_abs_diff'].mean()
    })

def plot_comparison(stats, merged_df, old_sheet, new_sheet, output_dir):
    sheet_dir = os.path.join(output_dir, f'{old_sheet}_vs_{new_sheet}')
    os.makedirs(sheet_dir, exist_ok=True)

    # 1. EVM difference distribution histogram
    plt.figure(figsize=(12, 6))
    plt.hist(merged_df['evm_diff'], bins=30, alpha=0.7, color='b')
    plt.axvline(merged_df['evm_diff'].mean(), color='r', linestyle='--', label=f'Mean = {merged_df["evm_diff"].mean():.2f}')
    plt.axvline(0, color='g', linestyle='-', label='No Difference')
    plt.title(f'{old_sheet} vs {new_sheet} EVM Difference Distribution')
    plt.xlabel('EVM Difference (dB)')
    plt.ylabel('Number of Records')
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(os.path.join(sheet_dir, 'evm_diff_distribution.png'), dpi=150)
    plt.close()

    # 2. Average difference by wifi_format and rate
    plt.figure(figsize=(16, 8))
    pivot = stats.pivot(index='rate', columns='wifi_format', values='mean_diff')
    sns.heatmap(pivot, annot=True, cmap='coolwarm', fmt='.2f', cbar_kws={'label': 'Average EVM Difference (dB)'})
    plt.title(f'{old_sheet} vs {new_sheet} Average EVM Difference by Format and Rate')
    plt.tight_layout()
    plt.savefig(os.path.join(sheet_dir, 'evm_diff_heatmap.png'), dpi=150)
    plt.close()

    # 3. Old vs New EVM comparison scatter plot
    plt.figure(figsize=(12, 10))
    plt.scatter(merged_df['evm_old'], merged_df['evm_new'], alpha=0.6, s=20)
    plt.plot([merged_df['evm_old'].min(), merged_df['evm_old'].max()],
             [merged_df['evm_old'].min(), merged_df['evm_old'].max()], 'r--', label='Ideal Case')
    plt.xlabel('Old Version EVM (dB)')
    plt.ylabel('New Version EVM (dB)')
    plt.title(f'{old_sheet} vs {new_sheet} Old vs New EVM Comparison')
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(os.path.join(sheet_dir, 'evm_comparison_scatter.png'), dpi=150)
    plt.close()

    # 4. EVM difference boxplot by wifi_format
    plt.figure(figsize=(16, 8))
    sns.boxplot(x='wifi_format', y='mean_diff', data=stats)
    plt.axhline(y=0, color='g', linestyle='-', label='No Difference')
    plt.title(f'{old_sheet} vs {new_sheet} EVM Difference by Wifi Format')
    plt.xlabel('Wifi Format')
    plt.ylabel('EVM Difference (dB)')
    plt.xticks(rotation=45)
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(sheet_dir, 'evm_diff_by_format.png'), dpi=150)
    plt.close()

    # 5. EVM difference boxplot by rate
    plt.figure(figsize=(16, 8))
    sns.boxplot(x='rate', y='mean_diff', data=stats)
    plt.axhline(y=0, color='g', linestyle='-', label='No Difference')
    plt.title(f'{old_sheet} vs {new_sheet} EVM Difference by Rate')
    plt.xlabel('Rate')
    plt.ylabel('EVM Difference (dB)')
    plt.xticks(rotation=45)
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(sheet_dir, 'evm_diff_by_rate.png'), dpi=150)
    plt.close()

def save_comparison_results(comparison_result, output_dir):
    # 保存HTML报告
    html_file = os.path.join(output_dir, 'evm_comparison_report.html')
    generate_html_report(comparison_result, output_dir, html_file)

    print(f"HTML报告已保存: {html_file}")

def generate_html_report(comparison_result, output_dir, output_file):
    # Generate HTML report
    html_content = '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>FPGA Old vs New Version EVM Comparison Analysis</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                margin: 20px;
                background-color: #f5f5f5;
            }
            .container {
                max-width: 1200px;
                margin: 0 auto;
                background-color: white;
                padding: 20px;
                border-radius: 8px;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            }
            h1, h2, h3 {
                color: #333;
            }
            .section {
                margin-bottom: 30px;
                padding-bottom: 20px;
                border-bottom: 1px solid #ddd;
            }
            .summary-table {
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 20px;
                background-color: white;
            }
            .summary-table th, .summary-table td {
                padding: 12px;
                text-align: left;
                border-bottom: 1px solid #ddd;
            }
            .summary-table th {
                background-color: #f2f2f2;
            }
            .summary-table tr:hover {
                background-color: #f5f5f5;
            }
            .success {
                color: green;
            }
            .warning {
                color: orange;
            }
            .error {
                color: red;
            }
            .image-container {
                text-align: center;
                margin: 20px 0;
            }
            .image-container img {
                max-width: 100%;
                height: auto;
                border: 1px solid #ddd;
                border-radius: 4px;
            }
            .sheet-link {
                color: #1a73e8;
                text-decoration: none;
            }
            .sheet-link:hover {
                text-decoration: underline;
            }
            .stats {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
                gap: 20px;
                margin-bottom: 30px;
            }
            .stat-card {
                background-color: #f8f9fa;
                padding: 15px;
                border-radius: 6px;
                box-shadow: 0 1px 3px rgba(0,0,0,0.1);
            }
            .stat-card h4 {
                margin-top: 0;
                color: #555;
            }
            .stat-value {
                font-size: 24px;
                font-weight: bold;
                margin: 10px 0;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>FPGA Old vs New Version EVM Comparison Analysis</h1>

            <div class="section">
                <h2>1. Overall Comparison Statistics</h2>
                <div class="stats">
    '''

    total_matched = 0
    total_old_records = 0
    total_new_records = 0
    max_diff = None
    min_diff = None
    avg_abs_diff = None

    for result in comparison_result:
        total_matched += result['matched_count']
        total_old_records += result['total_old_records']
        total_new_records += result['total_new_records']

        if max_diff is None or result['max_evm_diff'] > max_diff:
            max_diff = result['max_evm_diff']

        if min_diff is None or result['min_evm_diff'] < min_diff:
            min_diff = result['min_evm_diff']

        if avg_abs_diff is None:
            avg_abs_diff = result['avg_abs_diff']
        else:
            avg_abs_diff = (avg_abs_diff + result['avg_abs_diff']) / 2

    html_content += '''
                <div class="stat-card">
                    <h4>Matched Records</h4>
                    <div class="stat-value">%d</div>
                    <p>Old version total records: %d<br>New version total records: %d</p>
                </div>
                <div class="stat-card">
                    <h4>Average EVM Difference</h4>
                    <div class="stat-value" style="color:%s;">%.2f dB</div>
                    <p>Mean value</p>
                </div>
                <div class="stat-card">
                    <h4>Maximum EVM Difference</h4>
                    <div class="stat-value" style="color:%s;">%.2f dB</div>
                    <p>Maximum value</p>
                </div>
                <div class="stat-card">
                    <h4>Minimum EVM Difference</h4>
                    <div class="stat-value" style="color:%s;">%.2f dB</div>
                    <p>Minimum value</p>
                </div>
    ''' % (
        total_matched, total_old_records, total_new_records,
        'green' if abs(avg_abs_diff) < 1 else 'orange' if abs(avg_abs_diff) < 2 else 'red',
        avg_abs_diff,
        'red' if max_diff > 2 else 'orange' if max_diff > 1 else 'green',
        max_diff,
        'green' if min_diff > -1 else 'orange' if min_diff > -2 else 'red',
        min_diff
    )

    html_content += '''
                </div>
            </div>

            <div class="section">
                <h2>2. Sheet-level Comparison Details</h2>
                <table class="summary-table">
                    <thead>
                        <tr>
                            <th>Old Version Sheet</th>
                            <th>New Version Sheet</th>
                            <th>Matched Records</th>
                            <th>Old Records</th>
                            <th>New Records</th>
                            <th>Average Difference</th>
                            <th>Maximum Difference</th>
                            <th>Minimum Difference</th>
                            <th>Average Absolute Difference</th>
                            <th>Detailed Results</th>
                        </tr>
                    </thead>
                    <tbody>
    '''

    for result in comparison_result:
        avg_diff_class = 'success' if abs(result['mean_evm_diff']) < 1 else 'warning' if abs(result['mean_evm_diff']) < 2 else 'error'
        max_diff_class = 'success' if result['max_evm_diff'] <= 2 else 'warning' if result['max_evm_diff'] <= 5 else 'error'
        min_diff_class = 'success' if result['min_evm_diff'] >= -2 else 'warning' if result['min_evm_diff'] >= -5 else 'error'
        avg_abs_class = 'success' if result['avg_abs_diff'] < 1 else 'warning' if result['avg_abs_diff'] < 2 else 'error'

        html_content += '''
                        <tr>
                            <td>%s</td>
                            <td>%s</td>
                            <td>%d</td>
                            <td>%d</td>
                            <td>%d</td>
                            <td class="%s">%.2f</td>
                            <td class="%s">%.2f</td>
                            <td class="%s">%.2f</td>
                            <td class="%s">%.2f</td>
                            <td>
                                <a href="%s_vs_%s_detailed.xlsx" class="sheet-link">Detailed Data</a><br>
                                <a href="%s_vs_%s_summary.xlsx" class="sheet-link">Statistical Summary</a><br>
                                <a href="%s_vs_%s" class="sheet-link">Charts</a>
                            </td>
                        </tr>
        ''' % (
            result['old_sheet'],
            result['new_sheet'],
            result['matched_count'],
            result['total_old_records'],
            result['total_new_records'],
            avg_diff_class, result['mean_evm_diff'],
            max_diff_class, result['max_evm_diff'],
            min_diff_class, result['min_evm_diff'],
            avg_abs_class, result['avg_abs_diff'],
            result['old_sheet'], result['new_sheet'],
            result['old_sheet'], result['new_sheet'],
            result['old_sheet'], result['new_sheet']
        )

    html_content += '''
                    </tbody>
                </table>
            </div>

            <div class="section">
                <h2>3. Key Findings</h2>
                <ul>
                    <li><strong>Matched Records Percentage:</strong> %.2f%% of old version records found matching in new version</li>
                    <li><strong>Overall EVM Trend:</strong> %s</li>
                    <li><strong>Main Difference Source:</strong> Need to further analyze the performance of specific Rate and Format combinations</li>
                </ul>
            </div>

            <div class="section">
                <h2>4. Usage Suggestions</h2>
                <ol>
                    <li>Priority should be given to checking Sheets and Rate/Format combinations with larger differences</li>
                    <li>Focus on analyzing configurations with average absolute difference greater than 2dB</li>
                    <li>Check if new version covers all test scenarios</li>
                    <li>Perform hardware verification if necessary to confirm if differences are reasonable</li>
                </ol>
            </div>

            <div class="section">
                <h2>5. File Description</h2>
                <ul>
                    <li><strong>Analysis Script:</strong> compare_evm_old_new.py</li>
                    <li><strong>旧版本文件:</strong> D:/chip_test/dev/chip_tx/eagletest/py_script_fpga_tx_wifi7/Log/wifi_tx/20260407/vht_ht_old/merged_tx_result.xlsx</li>
                    <li><strong>新版本文件:</strong> D:/chip_test/dev/xian_test/Xian-Esp-Test-Scripts/py_script_fpga_tx_wifi7/Log/wifi_tx/20260407/vht_ht_2/merged_tx_result.xlsx</li>
                    <li><strong>生成时间:</strong> %s</li>
                </ul>
            </div>
        </div>
    </body>
    </html>
    ''' % (
        (total_matched / total_old_records) * 100,
        "New version has better overall EVM" if avg_abs_diff < -0.5 else "New version has worse overall EVM" if avg_abs_diff > 0.5 else "EVM difference between old and new versions is not significant",
        pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
    )

    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html_content)

if __name__ == "__main__":
    main()
