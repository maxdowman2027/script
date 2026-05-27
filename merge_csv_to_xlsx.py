#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
合并risc_wifitx格式的CSV文件到XLSX文件，按channel和编码方式（BCC/LDPC）划分Sheet。

另：将指定功率点（默认 15 dBm）的 EVM 透视统计写入单独 XLSX；
按 band、coding（LDPC/BCC）、NSS_STBC 组合分 Sheet；Sheet 内列为 bw_cbw、rate 及各 wifi_format 的平均 EVM，
并在同一 bw_cbw 分组内跨 rate 标出各 wifi_format 列最优（最负）与最差（最不负）EVM 底色。
合并完成后默认调用 txAnalyse_wifi7 生成 TX 多页 PDF（CSV 含 suer_dcm 时图标题含 dcm=），并对跨 rate EVM 均值、功率扫描曲线跳变及 **NSS2 双流 evm_nss0/evm_nss1 差**（单点 + 同 rate 跨 tx_pwr 链均值差）做异常扫描（可 CLI 关闭）。
"""

import os
import sys
import glob
import pandas as pd
import numpy as np
import re
import argparse
import openpyxl


def _infer_band_from_rf_chan(rf):
    """Infer 2G / 5G from rf_chan (channel index or center frequency in MHz)."""
    if rf is None or (isinstance(rf, float) and pd.isna(rf)):
        return "unknown"
    try:
        v = float(rf)
    except (TypeError, ValueError):
        return "unknown"
    if v > 1000:
        return "2G" if v < 3000 else "5G"
    if v <= 14:
        return "2G"
    if v >= 36:
        return "5G"
    return "unknown"


def _stream_cfg_from_sheet(sheet_name):
    """Parse NSS / STBC from sheet name (same convention as grouping keys)."""
    if not sheet_name:
        return "unknown"
    u = str(sheet_name).upper()
    if "STBC" in u:
        return "STBC"
    if "NSS2" in u:
        return "NSS2"
    if "NSS1" in u:
        return "NSS1"
    return "NSS1"


def _fec_label_row(row):
    """LDPC/BCC from fec_coding (0=BCC, 1=LDPC) or sheet name fallback."""
    fc = row.get("fec_coding")
    if fc is not None and not (isinstance(fc, float) and pd.isna(fc)):
        try:
            return "LDPC" if int(float(fc)) != 0 else "BCC"
        except (TypeError, ValueError):
            pass
    ss = str(row.get("_source_sheet", "") or "")
    if "LDPC" in ss:
        return "LDPC"
    if "BCC" in ss:
        return "BCC"
    return "unknown"


def _resolve_evm_column(df):
    for col in ("evm", "evm_aver(dB)", "aver_evmAll"):
        if col in df.columns:
            return col
    if "evm_nss0" in df.columns:
        return "evm_nss0"
    return None


def _is_ht_wifi_format(fmt):
    """True when wifi_format is HT (802.11n), excluding vht/nht/he names."""
    if fmt is None or (isinstance(fmt, float) and pd.isna(fmt)):
        return False
    s = str(fmt).strip().lower()
    s = s.replace("'", "").replace("[", "").replace("]", "").replace(" ", "")
    if not s:
        return False
    if "vht" in s or "nht" in s:
        return False
    return s == "ht"


def _normalize_ht_rate_for_summary(rate, wifi_fmt, stream_cfg):
    """
    HT 模式下的 rate 口径与单流对照一致，便于统计表聚类：
    - STBC：mcs0_stbc 视为 mcs0
    - NSS2：mcs8/9/… 分别视为 mcs0/1/…（MCS 下标减 8）
    """
    if not _is_ht_wifi_format(wifi_fmt):
        return rate
    if rate is None or (isinstance(rate, float) and pd.isna(rate)):
        return rate
    r0 = str(rate).strip()
    sc = str(stream_cfg or "").strip().upper()

    if sc == "STBC":
        m = re.match(r"(?i)mcs(\d+)_stbc\Z", r0)
        if m:
            return f"mcs{m.group(1)}"
        return r0

    if sc == "NSS2":
        m = re.match(r"(?i)mcs(\d+)\Z", r0)
        if m:
            n = int(m.group(1))
            if n >= 8:
                return f"mcs{n - 8}"
        return r0

    return r0


def build_evm_wifi_format_summary(df_all, tx_pwr_dbm=15.0, tx_pwr_tol=0.51):
    """
    For tx_power_set(dBm) ~= tx_pwr_dbm, pivot mean EVM by:
    band (2G/5G), coding (LDPC/BCC), cbw, rate, stream config (NSS1/NSS2/STBC),
    columns = wifi_format.

    HT (wifi_format ht): normalize rate before pivot — STBC strips *_stbc;
    NSS2 maps mcs8+ -> mcs(N-8).
    """
    if df_all is None or df_all.empty:
        return pd.DataFrame()

    df = df_all.copy()
    df.columns = [str(c).strip() for c in df.columns]

    pwr_col = "tx_power_set(dBm)"
    if pwr_col not in df.columns:
        print("统计表: 缺少列 tx_power_set(dBm)，跳过生成 evm 透视表")
        return pd.DataFrame()

    evm_col = _resolve_evm_column(df)
    if not evm_col:
        print("统计表: 未找到 evm / evm_aver(dB) / aver_evmAll / evm_nss0，跳过")
        return pd.DataFrame()

    if "wifi_format" not in df.columns:
        print("统计表: 缺少列 wifi_format，跳过")
        return pd.DataFrame()

    if "rf_chan" in df.columns:
        df["_band"] = df["rf_chan"].map(_infer_band_from_rf_chan)
    else:
        df["_band"] = "unknown"
    df["_coding"] = df.apply(_fec_label_row, axis=1)
    df["_stream_cfg"] = df["_source_sheet"].map(_stream_cfg_from_sheet) if "_source_sheet" in df.columns else "unknown"

    bw_col = "cbw" if "cbw" in df.columns else None
    if bw_col is None:
        print("统计表: 缺少列 cbw，带宽列将填空")
        df["_bw"] = ""
    else:
        df["_bw"] = df[bw_col]

    if "rate" not in df.columns:
        print("统计表: 缺少列 rate，跳过")
        return pd.DataFrame()

    df["_rate"] = df["rate"]
    df["_evm_num"] = pd.to_numeric(df[evm_col], errors="coerce")
    df["_pwr_num"] = pd.to_numeric(df[pwr_col], errors="coerce")

    m = df["_pwr_num"].sub(float(tx_pwr_dbm)).abs() <= float(tx_pwr_tol)
    sub = df.loc[m].dropna(subset=["_evm_num", "wifi_format"])

    if sub.empty:
        print(
            f"统计表: tx_power_set(dBm) 在 {tx_pwr_dbm}±{tx_pwr_tol} dBm 范围内无数据，"
            "跳过透视表（请确认日志是否包含该功率点）"
        )
        return pd.DataFrame()

    sub = sub.copy()
    sub["_rate"] = [
        _normalize_ht_rate_for_summary(r, wf, sc)
        for r, wf, sc in zip(sub["rate"], sub["wifi_format"], sub["_stream_cfg"])
    ]

    idx_cols = ["_band", "_coding", "_bw", "_rate", "_stream_cfg"]
    pt = pd.pivot_table(
        sub,
        values="_evm_num",
        index=idx_cols,
        columns="wifi_format",
        aggfunc="mean",
    )
    pt = pt.round(3)
    out = pt.reset_index()
    rename = {
        "_band": "band",
        "_coding": "coding_LDPC_BCC",
        "_bw": "bw_cbw",
        "_rate": "rate",
        "_stream_cfg": "NSS_STBC",
    }
    out = out.rename(columns=rename)
    # Stable column order: keys then wifi_format columns sorted
    fmt_cols = [c for c in out.columns if c not in rename.values()]
    fmt_cols = sorted(fmt_cols, key=lambda x: str(x).lower())
    out = out[[c for c in rename.values()] + fmt_cols]
    return out


def _concat_csvs_for_analysis(grouped_files):
    """Rebuild merged-frame columns when EVM summary collection was skipped."""
    parts = []
    for sheet_name, files in grouped_files.items():
        for fp in files:
            try:
                d = pd.read_csv(fp)
                d["_source_sheet"] = sheet_name
                parts.append(d)
            except Exception as e:
                print(f"异常检测: 读取 {fp} 失败: {e}")
    if not parts:
        return pd.DataFrame()
    return pd.concat(parts, ignore_index=True)


def _prepare_anomaly_dataframe(df_all):
    if df_all is None or df_all.empty:
        return pd.DataFrame()
    df = df_all.copy()
    df.columns = [str(c).strip() for c in df.columns]
    pwr_col = "tx_power_set(dBm)"
    evm_col = _resolve_evm_column(df)
    if not evm_col or pwr_col not in df.columns:
        return pd.DataFrame()
    if "wifi_format" not in df.columns or "rate" not in df.columns:
        return pd.DataFrame()
    if "rf_chan" in df.columns:
        df["_band"] = df["rf_chan"].map(_infer_band_from_rf_chan)
    else:
        df["_band"] = "unknown"
    df["_coding"] = df.apply(_fec_label_row, axis=1)
    df["_stream_cfg"] = (
        df["_source_sheet"].map(_stream_cfg_from_sheet)
        if "_source_sheet" in df.columns
        else "unknown"
    )
    bw_col = "cbw" if "cbw" in df.columns else None
    df["_bw"] = df[bw_col] if bw_col else ""
    df["_evm_num"] = pd.to_numeric(df[evm_col], errors="coerce")
    df["_pwr_num"] = pd.to_numeric(df[pwr_col], errors="coerce")
    df = df.dropna(subset=["_evm_num", "_pwr_num"])
    df["_rate_norm"] = [
        _normalize_ht_rate_for_summary(r, wf, sc)
        for r, wf, sc in zip(df["rate"], df["wifi_format"], df["_stream_cfg"])
    ]
    return df


def analyze_nss2_evm_stream_imbalance(
    df_all,
    evm_nss_gap_db=3.0,
    max_alert_rows=500,
    cross_pwr_mean_min_distinct_pwr=2,
    max_cross_pwr_alert_rows=100,
):
    """
    NSS2（双流）下比较 evm_nss0 与 evm_nss1：

    1) 单点：|ΔEVM| 过大视为链路边距异常（与工程常用 3 dB 阈值一致，可调）。
    2) 同 rate、跨不同 tx_pwr：对每个功率点先聚合链上 EVM，再对功率求
       mean(evm_nss0)、mean(evm_nss1)；若 |mean0 - mean1| 超过同一阈值，视为
       系统性链路边距异常（可检出单点未超阈但整段功率扫描平均偏一侧的情况）。

    仅处理 _source_sheet 对应 NSS2 的行；需同时存在有效数值的 evm_nss0、evm_nss1。
    """
    if evm_nss_gap_db is None or float(evm_nss_gap_db) <= 0:
        return []

    gap_thr = float(evm_nss_gap_db)
    if df_all is None or df_all.empty:
        return []

    df = df_all.copy()
    df.columns = [str(c).strip() for c in df.columns]
    if "evm_nss0" not in df.columns or "evm_nss1" not in df.columns:
        return []

    if "_source_sheet" not in df.columns:
        return []

    df["_stream_cfg"] = df["_source_sheet"].map(_stream_cfg_from_sheet)
    sub = df[df["_stream_cfg"] == "NSS2"].copy()
    if sub.empty:
        return []

    sub["_e0"] = pd.to_numeric(sub["evm_nss0"], errors="coerce")
    sub["_e1"] = pd.to_numeric(sub["evm_nss1"], errors="coerce")
    sub = sub.dropna(subset=["_e0", "_e1"])
    if sub.empty:
        return [f"NSS2 dual-stream EVM: no rows with both evm_nss0 and evm_nss1 numeric."]

    pwr_col = "tx_power_set(dBm)"
    if pwr_col in sub.columns:
        sub["_pwr_num"] = pd.to_numeric(sub[pwr_col], errors="coerce")
    else:
        sub["_pwr_num"] = np.nan

    if "wifi_format" in sub.columns:
        sub["_wf"] = sub["wifi_format"]
    else:
        sub["_wf"] = ""

    if "rate" in sub.columns:
        sub["_rate"] = sub["rate"]
    else:
        sub["_rate"] = ""

    if "rf_chan" in sub.columns:
        sub["_band"] = sub["rf_chan"].map(_infer_band_from_rf_chan)
    else:
        sub["_band"] = "unknown"
    sub["_coding"] = sub.apply(_fec_label_row, axis=1)
    bw_col = "cbw" if "cbw" in sub.columns else None
    sub["_bw"] = sub[bw_col] if bw_col else ""

    sub["_rate_norm"] = [
        _normalize_ht_rate_for_summary(r, wf, "NSS2")
        for r, wf in zip(sub["_rate"], sub["_wf"])
    ]

    lines = []

    # --- 同 rate、跨 tx_pwr：链上 EVM 先按功率点取均值，再对功率求链均值之差 ---
    min_p = int(cross_pwr_mean_min_distinct_pwr) if cross_pwr_mean_min_distinct_pwr else 2
    if min_p < 2:
        min_p = 2
    cap_agg = int(max_cross_pwr_alert_rows) if max_cross_pwr_alert_rows else 100
    if cap_agg < 1:
        cap_agg = 1

    agg_src = sub.dropna(subset=["_pwr_num"]).copy()
    group_cols = ["_source_sheet", "_band", "_coding", "_bw", "_wf", "_rate_norm"]
    agg_bad_rows = []
    if agg_src.empty:
        lines.append(
            "NSS2 chain EVM (same rate, mean over tx_pwr): skipped (no valid tx_power_set(dBm))."
        )
    else:
        for _key, g in agg_src.groupby(group_cols, dropna=False):
            per_pwr = (
                g.groupby("_pwr_num", observed=True)
                .agg(m0=("_e0", "mean"), m1=("_e1", "mean"))
                .dropna()
            )
            if len(per_pwr) < min_p:
                continue
            m0_all = float(per_pwr["m0"].mean())
            m1_all = float(per_pwr["m1"].mean())
            mean_gap = abs(m0_all - m1_all)
            if mean_gap <= gap_thr:
                continue
            g0 = g.iloc[0]
            agg_bad_rows.append(
                (
                    mean_gap,
                    m0_all,
                    m1_all,
                    len(per_pwr),
                    g0.get("_source_sheet", ""),
                    g0.get("_rate", ""),
                    g0.get("_rate_norm", ""),
                    g0.get("_wf", ""),
                    g0.get("_band", ""),
                    g0.get("_coding", ""),
                    g0.get("_bw", ""),
                )
            )

        agg_bad_rows.sort(key=lambda t: t[0], reverse=True)
        n_agg = len(agg_bad_rows)
        for tup in agg_bad_rows[:cap_agg]:
            (
                mean_gap,
                m0_all,
                m1_all,
                n_pwr,
                sheet_s,
                rate_raw,
                rate_norm,
                wf_s,
                band_s,
                coding_s,
                bw_s,
            ) = tup
            lines.append(
                "[ANOMALY ALERT] NSS2 chain EVM imbalance (same rate, mean over tx_pwr): "
                f"|mean(evm_nss0)-mean(evm_nss1)|={mean_gap:.2f} dB (threshold {gap_thr:.1f} dB); "
                f"mean_evm_nss0={m0_all:.2f} dB, mean_evm_nss1={m1_all:.2f} dB over {n_pwr} distinct tx_pwr; "
                f"rate={rate_raw}; rate_norm={rate_norm}; wifi_format={wf_s}; "
                f"band={band_s}; coding={coding_s}; cbw={bw_s}; sheet={sheet_s}"
            )
        if n_agg > cap_agg:
            lines.append(
                f"[ANOMALY ALERT] NSS2 chain EVM imbalance (same rate, mean over tx_pwr): "
                f"{n_agg - cap_agg} more group(s) omitted (cap {cap_agg})."
            )
        if n_agg == 0:
            lines.append(
                f"NSS2 chain EVM (same rate, mean over tx_pwr, >= {min_p} powers): "
                f"no |mean_nss0-mean_nss1| > {gap_thr:.1f} dB."
            )

    sub["_gap"] = (sub["_e0"] - sub["_e1"]).abs()
    bad = sub[sub["_gap"] > gap_thr].copy()
    if bad.empty:
        lines.append(
            f"NSS2 chain EVM (point-wise evm_nss0 vs evm_nss1): no |ΔEVM| > {gap_thr:.1f} dB."
        )
        return lines

    n_bad = len(bad)
    for _, row in bad.head(max_alert_rows).iterrows():
        pwr_s = (
            f"{float(row['_pwr_num']):.2f}"
            if pd.notna(row["_pwr_num"])
            else "n/a"
        )
        lines.append(
            "[ANOMALY ALERT] NSS2 chain EVM imbalance (evm_nss0 vs evm_nss1): "
            f"|ΔEVM|={float(row['_gap']):.2f} dB (threshold {gap_thr:.1f} dB); "
            f"evm_nss0={float(row['_e0']):.2f} dB, evm_nss1={float(row['_e1']):.2f} dB; "
            f"tx_pwr={pwr_s} dBm; rate={row['_rate']}; wifi_format={row['_wf']}; "
            f"band={row['_band']}; coding={row['_coding']}; cbw={row['_bw']}; "
            f"sheet={row.get('_source_sheet', '')}"
        )
    if n_bad > max_alert_rows:
        lines.append(
            f"[ANOMALY ALERT] NSS2 chain EVM imbalance: "
            f"{n_bad - max_alert_rows} more row(s) omitted (cap {max_alert_rows})."
        )
    return lines


def analyze_evm_tx_anomalies(
    df_all,
    rate_mean_gap_db=2.0,
    curve_jump_db=3.0,
    min_rates_in_group=2,
    min_pwr_points_per_rate_mean=2,
    min_pwr_points_for_curve=3,
):
    """
    在相同 band/coding/cbw/NSS_STBC/wifi_format 分组内：
    - 比较各 rate 在全部 tx_pwr 上的 EVM 均值，若某 rate 明显劣于同组中位水平则报警；
    - 对各 rate 按功率排序的 EVM 曲线，检测相邻功率点跳变过大。
    """
    df = _prepare_anomaly_dataframe(df_all)
    if df.empty:
        return ["(无有效 EVM / rate / tx_power_set(dBm) 数据，跳过 EVM 异常检测)"]

    group_cols = ["_band", "_coding", "_bw", "_stream_cfg", "wifi_format"]
    lines = []
    for key, g in df.groupby(group_cols, dropna=False):
        mbr = g.groupby("_rate_norm", observed=True)["_evm_num"].agg(["mean", "count"])
        mbr = mbr[mbr["count"] >= min_pwr_points_per_rate_mean]
        if len(mbr) >= min_rates_in_group:
            med = float(mbr["mean"].median())
            for r, row in mbr.iterrows():
                mean_r = float(row["mean"])
                if mean_r > med + rate_mean_gap_db:
                    ks = ", ".join(
                        f"{gc}={kv}"
                        for gc, kv in zip(
                            group_cols,
                            key if isinstance(key, tuple) else (key,),
                        )
                    )
                    lines.append(
                        "[ANOMALY ALERT] EVM rate mean vs peers: "
                        f"rate={r} mean EVM={mean_r:.2f} dB vs group median {med:.2f} dB "
                        f"(gap {mean_r - med:.2f} dB, threshold {rate_mean_gap_db} dB). "
                        f"Context: {ks}"
                    )
        for r in g["_rate_norm"].unique():
            sub = g[g["_rate_norm"] == r]
            pt = sub.groupby("_pwr_num", observed=True)["_evm_num"].mean().sort_index()
            if len(pt) < min_pwr_points_for_curve:
                continue
            arr = pt.values.astype(float)
            d = np.diff(arr)
            if d.size == 0:
                continue
            mx = float(np.nanmax(np.abs(d)))
            if mx > curve_jump_db:
                ks = ", ".join(
                    f"{gc}={kv}"
                    for gc, kv in zip(
                        group_cols,
                        key if isinstance(key, tuple) else (key,),
                    )
                )
                lines.append(
                    "[ANOMALY ALERT] EVM vs tx_pwr curve jump: "
                    f"rate={r} max adjacent |ΔEVM|={mx:.2f} dB "
                    f"(threshold {curve_jump_db} dB). Context: {ks}"
                )

    if not lines:
        lines.append(
            "EVM anomaly scan: no cross-rate mean deviation or power-sweep curve jump "
            f"exceeded thresholds (rate gap {rate_mean_gap_db} dB, curve jump {curve_jump_db} dB)."
        )
    return lines


def run_wifi7_tx_plots(csv_paths, plot_path_prefix):
    """
    Call txAnalyse_wifi7.tx_plot_and_analyse (multi-page PDF + txt checks).
    plot_path_prefix: directory + stem, e.g. .../merged_tx_result_; PDF becomes .../merged_tx_result_tx_pdf_*.pdf
    Chart titles: txAnalyse_wifi7._business_config_string; if CSV has suer_dcm (or user_dcm),
    titles include dcm=<value> (e.g. 20MHz | hesu | ch36 | LDPC | dcm=1 | EVM).
    """
    root = os.path.dirname(os.path.abspath(__file__))
    if root not in sys.path:
        sys.path.insert(0, root)
    from txAnalyse_wifi7 import tx_plot_and_analyse

    paths = sorted(csv_paths)
    if not paths:
        print("WiFi7 绘图: 无 CSV 路径，跳过")
        return
    os.makedirs(os.path.dirname(plot_path_prefix), exist_ok=True)
    print(f"WiFi7 绘图: 处理 {len(paths)} 个 CSV，输出前缀 {plot_path_prefix}")
    tx_plot_and_analyse(paths, plot_path_prefix)


# 独立 EVM 统计文件：按 band / LDPC(BCC) / NSS_STBC 拆 Sheet；Sheet 内着色分组键为 bw_cbw
SHEET_SPLIT_KEYS = ("band", "coding_LDPC_BCC", "NSS_STBC")
HIGHLIGHT_GROUP_KEYS = ("bw_cbw",)


def _evm_close(a, b, eps=1e-6):
    try:
        return abs(float(a) - float(b)) <= eps
    except (TypeError, ValueError):
        return False


def _set_summary_column_widths(ws):
    for col in ws.columns:
        letter = col[0].column_letter
        maxlen = min(max(len(str(c.value or "")) for c in col), 55)
        ws.column_dimensions[letter].width = maxlen + 2


def _apply_rate_group_evm_fills(ws, summary_tbl, group_keys):
    """
    在每个 group_keys 分组内按 rate 比较：对每个 wifi_format 列标出最优 / 最差 EVM。
    summary_tbl 须含 rate 列；wifi_format 为除 group_keys 与 rate 外的列。
    """
    fill_best = openpyxl.styles.PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    fill_worst = openpyxl.styles.PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )

    headers = {}
    for j in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=j).value
        if h is not None:
            headers[str(h).strip()] = j

    gset = set(group_keys)
    fmt_cols = [
        c
        for c in summary_tbl.columns
        if c not in gset and str(c) != "rate"
    ]

    for fmt in fmt_cols:
        if fmt not in headers:
            continue
        col_idx = headers[fmt]
        for _, grp in summary_tbl.groupby(list(group_keys), dropna=False):
            ser = pd.to_numeric(grp[fmt], errors="coerce")
            valid = ser.dropna()
            if valid.size == 0:
                continue
            vmin = float(valid.min())
            vmax = float(valid.max())
            for orig_idx in grp.index:
                val = grp.at[orig_idx, fmt]
                if pd.isna(val):
                    continue
                fv = float(val)
                row_excel = int(orig_idx) + 2
                if vmin == vmax:
                    if _evm_close(fv, vmin):
                        ws.cell(row=row_excel, column=col_idx).fill = fill_best
                else:
                    if _evm_close(fv, vmin):
                        ws.cell(row=row_excel, column=col_idx).fill = fill_best
                    elif _evm_close(fv, vmax):
                        ws.cell(row=row_excel, column=col_idx).fill = fill_worst


def _sanitize_evm_summary_sheet_name(key_tuple, max_len=31):
    parts = []
    for p in key_tuple:
        if p is None or (isinstance(p, float) and pd.isna(p)):
            s = "NA"
        else:
            s = str(p).strip()
        for ch in "\\/*?[]:":
            s = s.replace(ch, "_")
        s = s.replace("/", "_")
        parts.append(s if s else "NA")
    raw = "_".join(parts)
    return raw[:max_len]


def _unique_evm_summary_sheet_name(base, used_names):
    name = base[:31]
    if name not in used_names:
        used_names.add(name)
        return name
    i = 2
    while True:
        suf = f"_{i}"
        cand = (base[: max(1, 31 - len(suf))] + suf)[:31]
        if cand not in used_names:
            used_names.add(cand)
            return cand
        i += 1


def write_evm_summary_file(summary_tbl, output_path, summary_tx_pwr_dbm):
    """
    将透视表写入独立 xlsx：按 band、coding_LDPC_BCC、NSS_STBC 分 Sheet；
    各 Sheet 仅含 bw_cbw、rate 与各 wifi_format；着色按 bw_cbw 分组跨 rate。
    """
    if summary_tbl is None or summary_tbl.empty:
        return False

    split_cols = list(SHEET_SPLIT_KEYS)
    missing = [c for c in split_cols if c not in summary_tbl.columns]
    if missing:
        print(f"统计表缺少分 Sheet 列 {missing}，跳过独立 EVM 文件")
        return False

    out_dir = os.path.dirname(os.path.abspath(output_path))
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    chunks = []
    used_sheet_names = set()
    for key_tuple, sub in summary_tbl.groupby(split_cols, dropna=False):
        if sub.empty:
            continue
        display_df = sub.drop(columns=split_cols).reset_index(drop=True)
        base = _sanitize_evm_summary_sheet_name(key_tuple)
        sheet_name = _unique_evm_summary_sheet_name(base, used_sheet_names)
        chunks.append((sheet_name, display_df))

    if not chunks:
        return False

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, display_df in chunks:
            display_df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = openpyxl.load_workbook(output_path)
    for sheet_name, display_df in chunks:
        ws = wb[sheet_name]
        _set_summary_column_widths(ws)
        _apply_rate_group_evm_fills(ws, display_df, HIGHLIGHT_GROUP_KEYS)
    wb.save(output_path)
    wb.close()

    db = float(summary_tx_pwr_dbm)
    pwr_tag = str(int(db)) if db == int(db) else str(db).replace(".", "p")
    print(
        f"EVM 统计（{pwr_tag} dBm）已写入 {len(chunks)} 个 Sheet："
        f"{', '.join(c[0] for c in chunks)}"
    )
    return True


def merge_csv_to_xlsx(input_dir, output_file, crc_fail_file=None,
                      summary_tx_pwr_dbm=15.0, add_evm_summary=True,
                      evm_summary_output_file=None,
                      run_wifi7_plots=True,
                      wifi7_plot_dir=None,
                      run_evm_anomaly_check=True,
                      anomaly_report_file=None,
                      anomaly_rate_mean_gap_db=2.0,
                      anomaly_curve_jump_db=3.0,
                      anomaly_nss2_evm_gap_db=3.0):
    """
    合并指定文件夹中的CSV文件到XLSX文件

    Args:
        input_dir: 包含CSV文件的文件夹路径
        output_file: 输出的XLSX文件路径
        crc_fail_file: 保存psdu_crc为Fail的情况的XLSX文件路径
        summary_tx_pwr_dbm: 统计表筛选的发射功率（dBm），默认 15
        add_evm_summary: 是否生成独立 EVM 统计 xlsx
        evm_summary_output_file: EVM 统计输出路径；默认与合并文件同目录，
            文件名 {merged_basename}_evm_{功率}dBm_stat.xlsx；
            文件内按 band、coding_LDPC_BCC、NSS_STBC 分 Sheet，着色按 bw_cbw 分组跨 rate
        run_wifi7_plots: 是否调用 txAnalyse_wifi7.tx_plot_and_analyse 生成多页 PDF/附带 txt
        wifi7_plot_dir: WiFi7 分析输出目录；默认 {merged_basename}_wifi7_tx_plot（与合并文件同目录）
        run_evm_anomaly_check: 是否扫描跨 rate EVM 均值与功率曲线跳变并写报告
        anomaly_report_file: 异常报告 txt 路径；默认同目录 {basename}_evm_anomaly_report.txt
        anomaly_rate_mean_gap_db: 同配置下某 rate 均值劣于组内中位数的报警阈值（dB）
        anomaly_curve_jump_db: 同一 rate 相邻功率点 |ΔEVM| 报警阈值（dB）
        anomaly_nss2_evm_gap_db: NSS2 下链路边距报警阈值（dB）；None 或 <=0 关闭。
            同时用于：(1) 单点 |evm_nss0 - evm_nss1|；(2) 同 rate、跨不同 tx_pwr 上
            对两链 EVM 分别按功率取均值后再比较 |mean_nss0 - mean_nss1|（至少 2 个有效功率点）。
    """
    # 查找所有risc_wifitx_*.csv文件
    csv_files = glob.glob(os.path.join(input_dir, 'risc_wifitx_*.csv'))

    if not csv_files:
        print(f"未找到符合条件的CSV文件: {input_dir}")
        return

    print(f"找到 {len(csv_files)} 个CSV文件")

    # 按channel、编码方式和NSS/STBC分组
    grouped_files = {}

    for csv_file in csv_files:
        filename = os.path.basename(csv_file)

        # 从文件名中提取channel、编码方式和NSS/STBC
        # 文件名格式示例: risc_wifitx_20m_['11b']_BCC_channel11_GILTF0_2026-0331-175943.csv
        channel_match = re.search(r'channel(\d+)', filename)
        coding_match = re.search(r'(BCC|LDPC)', filename)
        nss_match = re.search(r'(NSS1|NSS2)', filename)
        stbc_match = re.search(r'(STBC)', filename)

        if channel_match and coding_match:
            channel = channel_match.group(1)
            coding = coding_match.group(1)

            sheet_name = f"channel{channel}_{coding}"

            if nss_match:
                sheet_name += f"_{nss_match.group(1)}"
            elif stbc_match:
                sheet_name += f"_{stbc_match.group(1)}"

            if sheet_name not in grouped_files:
                grouped_files[sheet_name] = []

            grouped_files[sheet_name].append(csv_file)

    print(f"按Sheet分组后: {list(grouped_files.keys())}")

    # 创建Excel写入器
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    crc_writer = None
    if crc_fail_file:
        crc_writer = pd.ExcelWriter(crc_fail_file, engine='openpyxl')

    all_for_summary = []

    # 处理每个分组的文件
    for sheet_name, files in grouped_files.items():
        print(f"处理Sheet: {sheet_name} ({len(files)}个文件)")

        # 读取所有CSV文件
        dfs = []
        for f in files:
            try:
                df = pd.read_csv(f)
                dfs.append(df)
            except Exception as e:
                print(f"读取文件 {f} 失败: {e}")
                continue

        if dfs:
            # 合并数据
            merged_df = pd.concat(dfs, ignore_index=True)

            # 调整列顺序，将evm_nss0和evm_nss1列插入到evm列之后
            if 'evm' in merged_df.columns:
                # 获取evm列的索引
                evm_index = merged_df.columns.get_loc('evm')

                # 检查是否有evm_nss0和evm_nss1列
                columns_to_move = []
                if 'evm_nss0' in merged_df.columns:
                    columns_to_move.append('evm_nss0')
                if 'evm_nss1' in merged_df.columns:
                    columns_to_move.append('evm_nss1')

                # 调整列顺序
                if columns_to_move:
                    # 获取所有列的列表
                    columns = list(merged_df.columns)
                    # 移除要移动的列
                    for col in columns_to_move:
                        columns.remove(col)
                    # 插入到evm列之后
                    for i, col in enumerate(columns_to_move):
                        columns.insert(evm_index + 1 + i, col)
                    # 重新排列数据框的列
                    merged_df = merged_df[columns]
            elif 'evm_nss0' in merged_df.columns or 'evm_nss1' in merged_df.columns:
                # 如果没有evm列，但有evm_nss列，则在适当位置添加evm列（可选）
                # 这里我们保持原样，因为用户只要求将evm_nss列放在evm列之后
                pass

            if add_evm_summary or run_evm_anomaly_check:
                part = merged_df.copy()
                part["_source_sheet"] = sheet_name
                all_for_summary.append(part)

            # 写入到Sheet
            merged_df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"成功写入 {len(merged_df)} 行数据到 {sheet_name}")

            # 为不同wifi_format的行添加填充色
            worksheet = writer.sheets[sheet_name]

            # 定义不同wifi_format对应的颜色
            format_colors = {
                '11b': 'FFCCFF',    # 浅粉色
                '11g': 'CCFFFF',    # 浅青色
                '11n': 'FFFFCC',    # 浅黄色
                'ht': 'CCFFCC',     # 浅绿色
                'vht': 'FFCCCC',    # 浅红色
                'he': 'CCCCFF',     # 浅紫色
                'hesu': 'E6CCFF',   # 淡紫色
                'heer': 'D9B3FF',   # 深紫色
                'nht': 'CCE5FF',    # 浅蓝色
                'wifi7': 'FFFFE5'   # 浅橙色
            }

            # 获取wifi_format列的索引（假设在第0列）
            # 如果wifi_format不在第一列，我们需要动态查找
            wifi_format_index = None
            for idx, col in enumerate(worksheet[1]):
                if col.value == 'wifi_format':
                    wifi_format_index = idx
                    break

            if wifi_format_index is not None:
                # 遍历每一行（从第2行开始，因为第1行是表头）
                for row in worksheet.iter_rows(min_row=2, max_row=len(merged_df)+1, min_col=1, max_col=worksheet.max_column):
                    # 获取wifi_format值
                    cell_value = row[wifi_format_index].value
                    # 匹配格式，确保更具体的格式先匹配
                    format_name = None
                    # 先检查更具体的格式
                    specific_formats = ['hesu', 'heer', 'vht', 'nht', 'ht', '11n', '11g', '11b', 'he', 'wifi7']
                    for key in specific_formats:
                        if isinstance(cell_value, str) and key.lower() in cell_value.strip().lower():
                            format_name = key
                            break

                    # 如果找到匹配的格式，设置填充色
                    if format_name and format_name in format_colors:
                        fill = openpyxl.styles.PatternFill(start_color=format_colors[format_name], end_color=format_colors[format_name], fill_type='solid')
                        for cell in row:
                            cell.fill = fill

                # 为evm相关列添加特殊填充色
                evm_columns = ['evm', 'evm_nss0', 'evm_nss1']
                for col_idx in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=1, column=col_idx).value
                    if cell_value in evm_columns:
                        # 为evm相关列添加黄色填充色
                        for row_idx in range(2, worksheet.max_row + 1):
                            evm_fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                            worksheet.cell(row=row_idx, column=col_idx).fill = evm_fill

            # 检查是否需要保存crc失败的情况
            if crc_writer and 'psdu_crc' in merged_df.columns:
                crc_fail_df = merged_df[merged_df['psdu_crc'] == 'Fail']
                if not crc_fail_df.empty:
                    # 写入到Sheet
                    crc_fail_df.to_excel(crc_writer, sheet_name=sheet_name, index=False)
                    print(f"找到 {len(crc_fail_df)} 行psdu_crc为Fail的记录，已写入到 {crc_fail_file}")

                    # 为crc_fail_result表格添加填充色
                    crc_worksheet = crc_writer.sheets[sheet_name]

                    # 定义不同wifi_format对应的颜色
                    format_colors = {
                        '11b': 'FFCCFF',    # 浅粉色
                        '11g': 'CCFFFF',    # 浅青色
                        '11n': 'FFFFCC',    # 浅黄色
                        'ht': 'CCFFCC',     # 浅绿色
                        'vht': 'FFCCCC',    # 浅红色
                        'he': 'CCCCFF',     # 浅紫色
                        'hesu': 'E6CCFF',   # 淡紫色
                        'heer': 'D9B3FF',   # 深紫色
                        'nht': 'CCE5FF',    # 浅蓝色
                        'wifi7': 'FFFFE5'   # 浅橙色
                    }

                    # 查找wifi_format列的索引
                    wifi_format_index = None
                    for idx, cell in enumerate(crc_worksheet[1]):
                        if cell.value == "wifi_format":
                            wifi_format_index = idx
                            break

                    # 设置列宽，让内容更加美观
                    for col in crc_worksheet.columns:
                        max_length = 0
                        column = col[0].column_letter  # 获取列字母
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        crc_worksheet.column_dimensions[column].width = adjusted_width

                    # 为重点列添加红色字体
                    priority_columns = ['tx_power_set(dBm)', 'evm', 'evm_nss0', 'evm_nss1']
                    for col_idx in range(1, crc_worksheet.max_column + 1):
                        cell_value = crc_worksheet.cell(row=1, column=col_idx).value
                        if cell_value in priority_columns:
                            # 将表头字体设置为红色（不加粗）
                            crc_worksheet.cell(row=1, column=col_idx).font = openpyxl.styles.Font(color="FF0000")

                    if wifi_format_index is not None:
                        print(f"在crc_fail_result中找到wifi_format列，索引为: {wifi_format_index}")
                        # 为不同wifi_format的行添加填充色（包括重点列的单元格）
                        for row_idx in range(2, crc_worksheet.max_row + 1):
                            cell_value = crc_worksheet.cell(row=row_idx, column=wifi_format_index + 1).value
                            row_fill = None
                            # 匹配格式，确保更具体的格式先匹配
                            specific_formats = ['hesu', 'heer', 'vht', 'nht', 'ht', '11n', '11g', '11b', 'he', 'wifi7']
                            for key in specific_formats:
                                if isinstance(cell_value, str) and key.lower() in cell_value.strip().lower():
                                    row_fill = format_colors[key]
                                    break

                            if row_fill:
                                # 为所有列的单元格添加wifi_format的填充色
                                fill = openpyxl.styles.PatternFill(start_color=row_fill, end_color=row_fill, fill_type='solid')
                                for col_idx in range(1, crc_worksheet.max_column + 1):
                                    crc_worksheet.cell(row=row_idx, column=col_idx).fill = fill

                        # 为evm相关列添加特殊填充色
                        evm_columns = ['evm', 'evm_nss0', 'evm_nss1']
                        for col_idx in range(1, crc_worksheet.max_column + 1):
                            cell_value = crc_worksheet.cell(row=1, column=col_idx).value
                            if cell_value in evm_columns:
                                # 为evm相关列添加黄色填充色
                                for row_idx in range(2, crc_worksheet.max_row + 1):
                                    evm_fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                                    crc_worksheet.cell(row=row_idx, column=col_idx).fill = evm_fill

    # 初始化Flatness和SpecMargin失败记录写入器
    flatness_writer = None
    specmargin_writer = None
    flatness_fail_file = None
    specmargin_fail_file = None

    # 首先收集所有失败记录，然后再创建ExcelWriter对象
    flatness_fail_data = {}
    specmargin_fail_data = {}

    # 重新遍历每个分组的文件，收集失败记录
    for sheet_name, files in grouped_files.items():
        # 读取并合并该分组的所有CSV文件
        dfs = []
        for f in files:
            try:
                df = pd.read_csv(f)
                dfs.append(df)
            except Exception as e:
                print(f"读取文件 {f} 失败: {e}")
                continue

        if dfs:
            merged_df_sheet = pd.concat(dfs, ignore_index=True)

            # 收集Flatness失败记录
            if 'spectralFlatness_margin' in merged_df_sheet.columns:
                flatness_fail_rows = []
                for index, row in merged_df_sheet.iterrows():
                    flatness_margin = row['spectralFlatness_margin']
                    if isinstance(flatness_margin, str):
                        flatness_values = re.findall(r'[-+]?\d*\.\d+|\d+', flatness_margin)
                        has_negative = False
                        for value in flatness_values:
                            try:
                                if float(value) < 0:
                                    has_negative = True
                                    break
                            except:
                                continue
                        if has_negative:
                            flatness_fail_rows.append(index)

                if flatness_fail_rows:
                    flatness_fail_data[sheet_name] = merged_df_sheet.loc[flatness_fail_rows]

            # 收集SpecMargin失败记录
            if 'spectrumMarginDb' in merged_df_sheet.columns or 'spectrumMarginDb_nss1' in merged_df_sheet.columns:
                specmargin_fail_rows = []
                specmargin_column = 'spectrumMarginDb' if 'spectrumMarginDb' in merged_df_sheet.columns else 'spectrumMarginDb_nss1'

                for index, row in merged_df_sheet.iterrows():
                    spectrum_margin = row[specmargin_column]
                    if isinstance(spectrum_margin, str):
                        specmargin_values = re.findall(r'[-+]?\d*\.\d+|\d+', spectrum_margin)
                        has_negative = False
                        for value in specmargin_values:
                            try:
                                if float(value) < 0:
                                    has_negative = True
                                    break
                            except:
                                continue
                        if has_negative:
                            specmargin_fail_rows.append(index)

                if specmargin_fail_rows:
                    specmargin_fail_data[sheet_name] = merged_df_sheet.loc[specmargin_fail_rows]

    # 只有在有失败记录时才创建ExcelWriter对象
    if flatness_fail_data:
        base_dir = os.path.dirname(output_file)
        base_name = os.path.splitext(os.path.basename(output_file))[0]
        flatness_fail_file = os.path.join(base_dir, f"{base_name}_flatness_fail.xlsx")
        flatness_writer = pd.ExcelWriter(flatness_fail_file, engine='openpyxl')

        # 写入Flatness失败记录
        for sheet_name, df in flatness_fail_data.items():
            df.to_excel(flatness_writer, sheet_name=sheet_name, index=False)
            print(f"找到 {len(df)} 行Flatness失败的记录，已写入到 {flatness_fail_file} 的 {sheet_name} Sheet")

            # 为不同wifi_format的行添加填充色
            worksheet = flatness_writer.sheets[sheet_name]
            format_colors = {
                '11b': 'FFCCFF',    # 浅粉色
                '11g': 'CCFFFF',    # 浅青色
                '11n': 'FFFFCC',    # 浅黄色
                'ht': 'CCFFCC',     # 浅绿色
                'vht': 'FFCCCC',    # 浅红色
                'he': 'CCCCFF',     # 浅紫色
                'hesu': 'E6CCFF',   # 淡紫色
                'heer': 'D9B3FF',   # 深紫色
                'nht': 'CCE5FF',    # 浅蓝色
                'wifi7': 'FFFFE5'   # 浅橙色
            }

            # 查找wifi_format列的索引
            wifi_format_index = None
            for idx, cell in enumerate(worksheet[1]):
                if cell.value == "wifi_format":
                    wifi_format_index = idx
                    break

            # 设置列宽
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column].width = adjusted_width

            # 为重点列添加红色字体
            priority_columns = ['tx_power_set(dBm)', 'evm', 'evm_nss0', 'evm_nss1', 'spectralFlatness_margin']
            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=1, column=col_idx).value
                if cell_value in priority_columns:
                    worksheet.cell(row=1, column=col_idx).font = openpyxl.styles.Font(color="FF0000")

                    if wifi_format_index is not None:
                        # 为不同wifi_format的行添加填充色
                        for row_idx in range(2, worksheet.max_row + 1):
                            cell_value = worksheet.cell(row=row_idx, column=wifi_format_index + 1).value
                            row_fill = None
                            specific_formats = ['hesu', 'heer', 'vht', 'nht', 'ht', '11n', '11g', '11b', 'he', 'wifi7']
                            for key in specific_formats:
                                if isinstance(cell_value, str) and key.lower() in cell_value.strip().lower():
                                    row_fill = format_colors[key]
                                    break

                            if row_fill:
                                fill = openpyxl.styles.PatternFill(start_color=row_fill, end_color=row_fill, fill_type='solid')
                                for col_idx in range(1, worksheet.max_column + 1):
                                    worksheet.cell(row=row_idx, column=col_idx).fill = fill

                        # 为evm相关列添加特殊填充色
                        evm_columns = ['evm', 'evm_nss0', 'evm_nss1']
                        for col_idx in range(1, worksheet.max_column + 1):
                            cell_value = worksheet.cell(row=1, column=col_idx).value
                            if cell_value in evm_columns:
                                # 为evm相关列添加黄色填充色
                                for row_idx in range(2, worksheet.max_row + 1):
                                    evm_fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                                    worksheet.cell(row=row_idx, column=col_idx).fill = evm_fill

    if specmargin_fail_data:
        base_dir = os.path.dirname(output_file)
        base_name = os.path.splitext(os.path.basename(output_file))[0]
        specmargin_fail_file = os.path.join(base_dir, f"{base_name}_specmargin_fail.xlsx")
        specmargin_writer = pd.ExcelWriter(specmargin_fail_file, engine='openpyxl')

        # 写入SpecMargin失败记录
        for sheet_name, df in specmargin_fail_data.items():
            df.to_excel(specmargin_writer, sheet_name=sheet_name, index=False)
            print(f"找到 {len(df)} 行SpecMargin失败的记录，已写入到 {specmargin_fail_file} 的 {sheet_name} Sheet")

            # 为不同wifi_format的行添加填充色
            worksheet = specmargin_writer.sheets[sheet_name]
            format_colors = {
                '11b': 'FFCCFF',    # 浅粉色
                '11g': 'CCFFFF',    # 浅青色
                '11n': 'FFFFCC',    # 浅黄色
                'ht': 'CCFFCC',     # 浅绿色
                'vht': 'FFCCCC',    # 浅红色
                'he': 'CCCCFF',     # 浅紫色
                'hesu': 'E6CCFF',   # 淡紫色
                'heer': 'D9B3FF',   # 深紫色
                'nht': 'CCE5FF',    # 浅蓝色
                'wifi7': 'FFFFE5'   # 浅橙色
            }

            # 查找wifi_format列的索引
            wifi_format_index = None
            for idx, cell in enumerate(worksheet[1]):
                if cell.value == "wifi_format":
                    wifi_format_index = idx
                    break

            # 设置列宽
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column].width = adjusted_width

            # 为重点列添加红色字体
            priority_columns = ['tx_power_set(dBm)', 'evm', 'evm_nss0', 'evm_nss1', 'spectrumMarginDb', 'spectrumMarginDb_nss1']
            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=1, column=col_idx).value
                if cell_value in priority_columns:
                    worksheet.cell(row=1, column=col_idx).font = openpyxl.styles.Font(color="FF0000")

                    if wifi_format_index is not None:
                        # 为不同wifi_format的行添加填充色
                        for row_idx in range(2, worksheet.max_row + 1):
                            cell_value = worksheet.cell(row=row_idx, column=wifi_format_index + 1).value
                            row_fill = None
                            specific_formats = ['hesu', 'heer', 'vht', 'nht', 'ht', '11n', '11g', '11b', 'he', 'wifi7']
                            for key in specific_formats:
                                if isinstance(cell_value, str) and key.lower() in cell_value.strip().lower():
                                    row_fill = format_colors[key]
                                    break

                            if row_fill:
                                fill = openpyxl.styles.PatternFill(start_color=row_fill, end_color=row_fill, fill_type='solid')
                                for col_idx in range(1, worksheet.max_column + 1):
                                    worksheet.cell(row=row_idx, column=col_idx).fill = fill

                        # 为evm相关列添加特殊填充色
                        evm_columns = ['evm', 'evm_nss0', 'evm_nss1']
                        for col_idx in range(1, worksheet.max_column + 1):
                            cell_value = worksheet.cell(row=1, column=col_idx).value
                            if cell_value in evm_columns:
                                # 为evm相关列添加黄色填充色
                                for row_idx in range(2, worksheet.max_row + 1):
                                    evm_fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                                    worksheet.cell(row=row_idx, column=col_idx).fill = evm_fill

    # 为失败记录添加填充色并保存文件
    if flatness_writer:
        for sheet_name, df in flatness_fail_data.items():
            worksheet = flatness_writer.sheets[sheet_name]
            format_colors = {
                '11b': 'FFCCFF',    # 浅粉色
                '11g': 'CCFFFF',    # 浅青色
                '11n': 'FFFFCC',    # 浅黄色
                'ht': 'CCFFCC',     # 浅绿色
                'vht': 'FFCCCC',    # 浅红色
                'he': 'CCCCFF',     # 浅紫色
                'hesu': 'E6CCFF',   # 淡紫色
                'heer': 'D9B3FF',   # 深紫色
                'nht': 'CCE5FF',    # 浅蓝色
                'wifi7': 'FFFFE5'   # 浅橙色
            }

            # 查找wifi_format列的索引
            wifi_format_index = None
            for idx, cell in enumerate(worksheet[1]):
                if cell.value == "wifi_format":
                    wifi_format_index = idx
                    break

            # 设置列宽
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column].width = adjusted_width

            # 为重点列添加红色字体
            priority_columns = ['tx_power_set(dBm)', 'evm', 'evm_nss0', 'evm_nss1', 'spectralFlatness_margin']
            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=1, column=col_idx).value
                if cell_value in priority_columns:
                    worksheet.cell(row=1, column=col_idx).font = openpyxl.styles.Font(color="FF0000")

            if wifi_format_index is not None:
                # 为不同wifi_format的行添加填充色
                for row_idx in range(2, worksheet.max_row + 1):
                    cell_value = worksheet.cell(row=row_idx, column=wifi_format_index + 1).value
                    row_fill = None
                    specific_formats = ['hesu', 'heer', 'vht', 'nht', 'ht', '11n', '11g', '11b', 'he', 'wifi7']
                    for key in specific_formats:
                        if isinstance(cell_value, str) and key.lower() in cell_value.strip().lower():
                            row_fill = format_colors[key]
                            break

                    if row_fill:
                        fill = openpyxl.styles.PatternFill(start_color=row_fill, end_color=row_fill, fill_type='solid')
                        for col_idx in range(1, worksheet.max_column + 1):
                            worksheet.cell(row=row_idx, column=col_idx).fill = fill

                # 为evm相关列添加特殊填充色
                evm_columns = ['evm', 'evm_nss0', 'evm_nss1']
                for col_idx in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=1, column=col_idx).value
                    if cell_value in evm_columns:
                        # 为evm相关列添加黄色填充色
                        for row_idx in range(2, worksheet.max_row + 1):
                            evm_fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                            worksheet.cell(row=row_idx, column=col_idx).fill = evm_fill

        flatness_writer.close()
        print(f"Flatness失败记录已保存到: {flatness_fail_file}")

    if specmargin_writer:
        for sheet_name, df in specmargin_fail_data.items():
            worksheet = specmargin_writer.sheets[sheet_name]
            format_colors = {
                '11b': 'FFCCFF',    # 浅粉色
                '11g': 'CCFFFF',    # 浅青色
                '11n': 'FFFFCC',    # 浅黄色
                'ht': 'CCFFCC',     # 浅绿色
                'vht': 'FFCCCC',    # 浅红色
                'he': 'CCCCFF',     # 浅紫色
                'hesu': 'E6CCFF',   # 淡紫色
                'heer': 'D9B3FF',   # 深紫色
                'nht': 'CCE5FF',    # 浅蓝色
                'wifi7': 'FFFFE5'   # 浅橙色
            }

            # 查找wifi_format列的索引
            wifi_format_index = None
            for idx, cell in enumerate(worksheet[1]):
                if cell.value == "wifi_format":
                    wifi_format_index = idx
                    break

            # 设置列宽
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column].width = adjusted_width

            # 为重点列添加红色字体
            priority_columns = ['tx_power_set(dBm)', 'evm', 'evm_nss0', 'evm_nss1', 'spectrumMarginDb', 'spectrumMarginDb_nss1']
            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=1, column=col_idx).value
                if cell_value in priority_columns:
                    worksheet.cell(row=1, column=col_idx).font = openpyxl.styles.Font(color="FF0000")

            if wifi_format_index is not None:
                # 为不同wifi_format的行添加填充色
                for row_idx in range(2, worksheet.max_row + 1):
                    cell_value = worksheet.cell(row=row_idx, column=wifi_format_index + 1).value
                    row_fill = None
                    specific_formats = ['hesu', 'heer', 'vht', 'nht', 'ht', '11n', '11g', '11b', 'he', 'wifi7']
                    for key in specific_formats:
                        if isinstance(cell_value, str) and key.lower() in cell_value.strip().lower():
                            row_fill = format_colors[key]
                            break

                    if row_fill:
                        fill = openpyxl.styles.PatternFill(start_color=row_fill, end_color=row_fill, fill_type='solid')
                        for col_idx in range(1, worksheet.max_column + 1):
                            worksheet.cell(row=row_idx, column=col_idx).fill = fill

                # 为evm相关列添加特殊填充色
                evm_columns = ['evm', 'evm_nss0', 'evm_nss1']
                for col_idx in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=1, column=col_idx).value
                    if cell_value in evm_columns:
                        # 为evm相关列添加黄色填充色
                        for row_idx in range(2, worksheet.max_row + 1):
                            evm_fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                            worksheet.cell(row=row_idx, column=col_idx).fill = evm_fill

        specmargin_writer.close()
        print(f"SpecMargin失败记录已保存到: {specmargin_fail_file}")

    # 保存文件
    try:
        if add_evm_summary and all_for_summary:
            summary_concat = pd.concat(all_for_summary, ignore_index=True)
            summary_tbl = build_evm_wifi_format_summary(
                summary_concat, tx_pwr_dbm=summary_tx_pwr_dbm
            )
            if not summary_tbl.empty:
                if evm_summary_output_file:
                    evm_out = evm_summary_output_file
                else:
                    base_dir = os.path.dirname(os.path.abspath(output_file))
                    base_name = os.path.splitext(os.path.basename(output_file))[0]
                    db = float(summary_tx_pwr_dbm)
                    pwr_tag = str(int(db)) if db == int(db) else str(db).replace(".", "p")
                    evm_out = os.path.join(
                        base_dir, f"{base_name}_evm_{pwr_tag}dBm_stat.xlsx"
                    )
                if write_evm_summary_file(
                    summary_tbl, evm_out, summary_tx_pwr_dbm
                ):
                    print(
                        f"独立 EVM 统计路径: {evm_out} "
                        f"（共 {len(summary_tbl)} 行透视结果；功率点 {summary_tx_pwr_dbm} dBm）"
                    )

        writer.close()
        print(f"合并完成！文件已保存到: {output_file}")

        if crc_writer:
            crc_writer.close()
            print(f"CRC失败记录已保存到: {crc_fail_file}")

        base_dir = os.path.dirname(os.path.abspath(output_file))
        base_name = os.path.splitext(os.path.basename(output_file))[0]

        if run_wifi7_plots:
            try:
                plot_dir = (
                    wifi7_plot_dir
                    if wifi7_plot_dir
                    else os.path.join(base_dir, f"{base_name}_wifi7_tx_plot")
                )
                plot_prefix = os.path.join(plot_dir, base_name + "_")
                run_wifi7_tx_plots(csv_files, plot_prefix)
                print(f"WiFi7 TX 分析 PDF/TXT 输出目录: {plot_dir}")
            except Exception as ex:
                print(f"WiFi7 绘图/分析失败: {ex}")

        if run_evm_anomaly_check:
            try:
                if all_for_summary:
                    adf = pd.concat(all_for_summary, ignore_index=True)
                else:
                    adf = _concat_csvs_for_analysis(grouped_files)
                lines = analyze_evm_tx_anomalies(
                    adf,
                    rate_mean_gap_db=anomaly_rate_mean_gap_db,
                    curve_jump_db=anomaly_curve_jump_db,
                )
                if anomaly_nss2_evm_gap_db is not None and float(anomaly_nss2_evm_gap_db) > 0:
                    lines.extend(
                        analyze_nss2_evm_stream_imbalance(
                            adf, evm_nss_gap_db=float(anomaly_nss2_evm_gap_db)
                        )
                    )
                rep = anomaly_report_file or os.path.join(
                    base_dir, f"{base_name}_evm_anomaly_report.txt"
                )
                with open(rep, "w", encoding="utf-8") as rf:
                    rf.write("\n".join(lines))
                for ln in lines:
                    if ln.startswith("[ANOMALY ALERT]"):
                        print(ln)
                print(f"EVM 异常检测报告: {rep}")
            except Exception as ex:
                print(f"EVM 异常检测失败: {ex}")
    except Exception as e:
        print(f"保存文件失败: {e}")


def main():
    parser = argparse.ArgumentParser(description="合并 risc_wifitx CSV 到 XLSX，并生成 15dBm EVM 透视统计表")
    parser.add_argument("--input_dir", default=r"D:\chip_test\dev\xian_test\Xian-Esp-Test-Scripts\py_script_fpga_tx_wifi7\Log\wifi_tx_rls4\regression_260526")
    parser.add_argument("--output_file", default=r"D:\chip_test\dev\xian_test\Xian-Esp-Test-Scripts\py_script_fpga_tx_wifi7\Log\wifi_tx_rls4\regression_260526/merged_tx_result.xlsx")
    parser.add_argument("--crc_fail_file", default=r"D:\chip_test\dev\xian_test\Xian-Esp-Test-Scripts\py_script_fpga_tx_wifi7\Log\wifi_tx_rls4\regression_260526/tx_crc_fail_result.xlsx")
    parser.add_argument("--summary_tx_pwr", type=float, default=15.0, help="统计表使用的 tx 功率点 (dBm)")
    parser.add_argument("--no_evm_summary", action="store_true", help="不生成独立 EVM 统计 xlsx")
    parser.add_argument(
        "--evm_summary_out",
        default=None,
        help="EVM 统计输出 xlsx 路径（默认与合并结果同目录：{basename}_evm_{功率}dBm_stat.xlsx）",
    )
    parser.add_argument("--no_wifi7_plots", action="store_true", help="不调用 txAnalyse_wifi7 绘图")
    parser.add_argument(
        "--wifi7_plot_dir",
        default=None,
        help="WiFi7 PDF/TXT 输出目录（默认：与合并文件同目录下 {basename}_wifi7_tx_plot）",
    )
    parser.add_argument("--no_evm_anomaly", action="store_true", help="不写 EVM 跨速率/曲线异常报告")
    parser.add_argument(
        "--anomaly_report",
        default=None,
        help="EVM 异常报告 txt 路径（默认同目录 {basename}_evm_anomaly_report.txt）",
    )
    parser.add_argument(
        "--anomaly_rate_gap",
        type=float,
        default=2.5,
        help="同配置下 rate 均值劣于组内中位数的报警阈值 (dB)，默认 2",
    )
    parser.add_argument(
        "--anomaly_curve_jump",
        type=float,
        default=3.0,
        help="同一 rate 相邻功率点 |ΔEVM| 报警阈值 (dB)，默认 3",
    )
    parser.add_argument(
        "--anomaly_nss2_evm_gap",
        type=float,
        default=3.0,
        help="NSS2 链路边距阈值 (dB)：单点 |evm_nss0-evm_nss1| 与同 rate 跨 tx_pwr 的 |mean_nss0-mean_nss1|，默认 3",
    )
    parser.add_argument(
        "--no_anomaly_nss2",
        action="store_true",
        help="关闭 NSS2 evm_nss0/evm_nss1 链路边距异常检测",
    )
    args = parser.parse_args()

    print(f"输入路径: {args.input_dir}")
    print(f"输出文件: {args.output_file}")
    print(f"CRC失败记录文件: {args.crc_fail_file}")

    merge_csv_to_xlsx(
        args.input_dir,
        args.output_file,
        args.crc_fail_file,
        summary_tx_pwr_dbm=args.summary_tx_pwr,
        add_evm_summary=not args.no_evm_summary,
        evm_summary_output_file=args.evm_summary_out,
        run_wifi7_plots=not args.no_wifi7_plots,
        wifi7_plot_dir=args.wifi7_plot_dir,
        run_evm_anomaly_check=not args.no_evm_anomaly,
        anomaly_report_file=args.anomaly_report,
        anomaly_rate_mean_gap_db=args.anomaly_rate_gap,
        anomaly_curve_jump_db=args.anomaly_curve_jump,
        anomaly_nss2_evm_gap_db=None if args.no_anomaly_nss2 else args.anomaly_nss2_evm_gap,
    )


if __name__ == '__main__':
    main()
