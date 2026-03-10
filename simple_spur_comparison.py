import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font

# 读取两个 CSV 文件
file1_path = r'D:\users\gxu\spur_scan\260309\scan_spur_data\normal\spur_scan_result_2G_coef.csv'
file2_path = r'D:\users\gxu\spur_scan\260309\scan_spur_data\xtal_duty_disable\spur_scan_result_2G_coef.csv'

df1 = pd.read_csv(file1_path)
df2 = pd.read_csv(file2_path)

# 数据预处理
for col in ['phy_mode', 'channel', 'frequency', 'Used_Frequency', 'X_CoefFixed', 'Y_CoefFixed']:
    if col in df1.columns:
        df1[col] = df1[col].replace(['no_spur', 'invalid'], np.nan)
    if col in df2.columns:
        df2[col] = df2[col].replace(['no_spur', 'invalid'], np.nan)

pwr_cols = [col for col in df1.columns if 'pwr' in col or 'avg' in col]
for col in pwr_cols:
    if col in df1.columns:
        df1[col] = df1[col].replace(['#VALUE!', '[', ']', '', ' ', 'no_spur'], np.nan)
        df1[col] = pd.to_numeric(df1[col], errors='coerce')
    if col in df2.columns:
        df2[col] = df2[col].replace(['#VALUE!', '[', ']', '', ' ', 'no_spur'], np.nan)
        df2[col] = pd.to_numeric(df2[col], errors='coerce')

merge_columns = ['phy_mode', 'channel', 'frequency', 'Used_Frequency', 'X_CoefFixed', 'Y_CoefFixed']
merged_df = pd.merge(df1, df2, on=merge_columns, how='inner', suffixes=('_normal', '_xtal'))

merged_df['avg_pwr_diff'] = merged_df['avg pwr_xtal'] - merged_df['avg pwr_normal']

threshold = 5
# 只有当Normal大于XTAL且差异>5dB时才视为异常
merged_df['is_abnormal'] = (merged_df['avg pwr_normal'] > merged_df['avg pwr_xtal']) & (merged_df['avg pwr_normal'] - merged_df['avg pwr_xtal'] > threshold)
merged_df['normal_greater'] = merged_df['avg pwr_normal'] > merged_df['avg pwr_xtal']

output_file = r'D:\users\gxu\spur_scan\260309\scan_spur_data\spur_simple_comparison.xlsx'
merged_df.to_excel(output_file, index=False, engine='openpyxl')

wb = openpyxl.load_workbook(output_file)
ws = wb.active

red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
red_font = Font(color="FF0000", bold=True)
blue_font = Font(color="0000FF", bold=True)

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
    index = i
    if merged_df.iloc[index]['is_abnormal']:
        for cell in row:
            cell.fill = red_fill
            cell.font = red_font
    elif merged_df.iloc[index]['normal_greater']:
        for cell in row:
            cell.font = blue_font

ws.cell(row=1, column=len(merged_df.columns)+1, value='说明')
ws.cell(row=2, column=len(merged_df.columns)+1, value='红色：异常数据（Normal > XTAL且差异>5dB）')
ws.cell(row=3, column=len(merged_df.columns)+1, value='蓝色：Normal > XTAL（需要注意的情况）')
ws.cell(row=4, column=len(merged_df.columns)+1, value='白色：Normal ≤ XTAL（正常情况）')

for col in range(1, ws.max_column+1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12

wb.save(output_file)

print(f"比较结果已保存到: {output_file}")
print(f"有效配置数: {len(merged_df)}")
print(f"异常数据数: {len(merged_df[merged_df['is_abnormal']])}")
print(f"Normal > XTAL（需要注意的情况）: {len(merged_df[merged_df['normal_greater']])}")
print(f"Normal ≤ XTAL（正常情况）: {len(merged_df[~merged_df['normal_greater']])}")

