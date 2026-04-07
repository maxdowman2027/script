import openpyxl

def check_actual_colors():
    file_path = "D:/chip_test/dev/xian_test/Xian-Esp-Test-Scripts/py_script_fpga_tx_wifi7/Log/wifi_tx_rls4/no_he/crc_fail_result.xlsx"

    try:
        wb = openpyxl.load_workbook(file_path)
        sheet_name = "channel11_BCC"
        ws = wb[sheet_name]

        # 查找wifi_format列的索引
        wifi_format_index = None
        for idx, cell in enumerate(ws[1]):
            if cell.value == "wifi_format":
                wifi_format_index = idx
                break

        if wifi_format_index is not None:
            print(f"找到wifi_format列，索引为: {wifi_format_index}")
        else:
            print("未找到wifi_format列")

        colors_used = {}
        for row_idx in range(2, min(11, ws.max_row + 1)):
            cell_value = ws.cell(row=row_idx, column=wifi_format_index + 1).value
            cell_color = ws.cell(row=row_idx, column=1).fill.start_color.index
            colors_used[cell_value] = cell_color

        print("\nFound wifi_format -> color mappings:")
        for format, color in colors_used.items():
            print(f"  {format}: {color}")

        wb.close()
    except Exception as e:
        print(f"Error reading file: {e}")

if __name__ == "__main__":
    check_actual_colors()