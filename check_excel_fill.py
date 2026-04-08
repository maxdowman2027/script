
import openpyxl

def check_fill_color(excel_file):
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        print("检查文件:", excel_file)
        print("-" * 50)

        # 查找EVM相关列
        evm_old_col = None
        evm_new_col = None
        evm_diff_col = None
        for idx, cell in enumerate(ws[1]):
            if cell.value == 'evm_old':
                evm_old_col = idx + 1
            elif cell.value == 'evm_new':
                evm_new_col = idx + 1
            elif cell.value == 'evm_diff':
                evm_diff_col = idx + 1

        if evm_old_col or evm_new_col or evm_diff_col:
            print("找到EVM相关列:")
            if evm_old_col:
                print(f"  evm_old: 第{evm_old_col}列")
            if evm_new_col:
                print(f"  evm_new: 第{evm_new_col}列")
            if evm_diff_col:
                print(f"  evm_diff: 第{evm_diff_col}列")
        else:
            print("未找到EVM相关列")
            return

        print("-" * 50)
        print("检查填充色是否已应用:")

        # 检查前几行的填充色
        found_fill = False
        for row in range(2, min(10, ws.max_row + 1)):
            if evm_old_col:
                cell = ws.cell(row=row, column=evm_old_col)
                if cell.fill.start_color.index != '00000000':
                    print(f"  第{row}行evm_old: 填充色 {cell.fill.start_color.index}")
                    found_fill = True
            if evm_new_col:
                cell = ws.cell(row=row, column=evm_new_col)
                if cell.fill.start_color.index != '00000000':
                    print(f"  第{row}行evm_new: 填充色 {cell.fill.start_color.index}")
                    found_fill = True
            if evm_diff_col:
                cell = ws.cell(row=row, column=evm_diff_col)
                if cell.fill.start_color.index != '00000000':
                    print(f"  第{row}行evm_diff: 填充色 {cell.fill.start_color.index}")
                    found_fill = True

        if found_fill:
            print("-" * 50)
            print("填充色已成功应用")
        else:
            print("-" * 50)
            print("未找到填充色")

        wb.close()

    except Exception as e:
        print(f"检查Excel文件时出错: {e}")

if __name__ == "__main__":
    file1 = r"D:\chip_test\dev\xian_test\Xian-Esp-Test-Scripts\py_script_fpga_tx_wifi7\Log\wifi_tx\20260407\vht_ht_2\evm_comparison_results\channel11_BCC_NSS1_vs_channel11_BCC_NSS1_detailed.xlsx"
    file2 = r"D:\chip_test\dev\xian_test\Xian-Esp-Test-Scripts\py_script_fpga_tx_wifi7\Log\wifi_tx\20260407\vht_ht_2\evm_comparison_results\channel5180_BCC_NSS1_vs_channel5180_BCC_NSS1_detailed.xlsx"

    check_fill_color(file1)
    print("\n" + "-" * 50 + "\n")
    check_fill_color(file2)
